def generar_responsiva_completa_pdf(
    empleado,
    departamento,
    dispositivo,
    marca,
    imei,
    modelo,
    plan,
    serie,
    area,
    estado,
    cargador,
    funcionalidad,
    numero_corporativo,
    valor_equipo,
    ciudad,
    fecha,
    logo_path
):
    """Genera un PDF de responsiva de equipo de trabajo (3 páginas) con logo y textos fijos."""
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    # --- PAGINA 1: RESPONSIVA ---
    pdf.add_page()
    # if logo_path:
    #     pdf.image(logo_path, x=10, y=8, w=28)
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 18, '', ln=True)
    pdf.cell(0, 8, "ASIGNACIÓN DE EQUIPO DE TRABAJO", ln=True, align='C')
    pdf.ln(6)
    pdf.set_font("Arial", size=11)
    pdf.multi_cell(0, 8, f"Yo ____________________________, con tarjeta de identidad número ____________________________, del departamento de {departamento},\nestoy recibiendo por parte de MOTOS (OPERACIONES), el siguiente equipo de trabajo con las siguientes características:")
    pdf.ln(2)
    pdf.set_font("Arial", "B", 11)
    pdf.cell(0, 7, f"Dispositivo: {dispositivo}", ln=True)
    pdf.cell(0, 7, f"Marca: {marca}", ln=True)
    pdf.set_font("Arial", size=11)
    pdf.cell(0, 7, f"Imei: {imei}", ln=True)
    pdf.set_font("Arial", "B", 11)
    pdf.cell(0, 7, f"Modelo: {modelo}    Plan: {plan}", ln=True)
    pdf.cell(0, 7, f"Serie: {serie}    Área: {area} / Departamento: {departamento}", ln=True)
    pdf.cell(0, 7, f"Estado: {estado}    Cargador: {cargador}", ln=True)
    pdf.cell(0, 7, f"Funcionalidad: {funcionalidad}    Número corporativo: {numero_corporativo}", ln=True)
    pdf.ln(2)
    pdf.set_font("Arial", size=11)
    pdf.multi_cell(0, 7, "El cual recibo en perfectas condiciones, y el cual queda bajo mi responsabilidad, comprometiéndome al uso adecuado del mismo, y a entregarlo al retirarme de la empresa. De comprobarse que por negligencia o mal manejo lo dañara, autorizo a la empresa la deducción respectiva del monto correspondiente.")
    pdf.ln(4)
    pdf.cell(0, 7, f"En la ciudad de {ciudad}, a los ___ días de ________ del {fecha}", ln=True)
    pdf.ln(10)
    pdf.cell(80, 7, "_________________________", ln=0, align='C')
    pdf.cell(30, 7, "", ln=0)
    pdf.cell(80, 7, "_________________________", ln=1, align='C')
    pdf.cell(80, 7, "Responsable", ln=0, align='C')
    pdf.cell(30, 7, "", ln=0)
    pdf.cell(80, 7, "Coordinador Lty Seguridad", ln=1, align='C')
    pdf.ln(6)
    pdf.set_font("Arial", "B", 11)
    pdf.cell(0, 7, f"VALOR DE EQUIPO: L{valor_equipo:,.2f}", ln=True)

    # --- PAGINA 2: POLITICAS DE USO ---
    pdf.add_page()
    # if logo_path:
    #     pdf.image(logo_path, x=10, y=8, w=28)
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 18, '', ln=True)
    pdf.cell(0, 8, "POLÍTICAS DE USO", ln=True, align='C')
    pdf.ln(4)
    pdf.set_font("Arial", size=10)
    politicas = (
        "¿Cómo usar el dispositivo móvil propiedad de la compañía?\n"
        "El teléfono móvil propiedad de la empresa KM MOTOS debe usarse solo para fines oficiales. Se permite el uso raro y muy limitado del teléfono móvil para uso personal, sin embargo, todo daño o mal funcionamiento será como propiedad del uso del móvil como una gran herramienta de comunicación para lograr tener una atención al cliente de primera y profesional así logrando mejores resultados de forma general.\n\n"
        "El teléfono móvil propiedad de la empresa debe ser utilizado solo por el empleado elegido. No puede ser utilizado por ningún familiar o amigo del empleado. No puede ser utilizado por ningún miembro de la familia. El dispositivo móvil propiedad de la empresa debe utilizarse cumpliendo con la política. La empresa se reserva el derecho de proceder a retiro si se sospecha que no se está cumpliendo con la política.\n\n"
        "El empleado que lleva el teléfono móvil propiedad de la compañía debe asegurarse de que esté en uso y que no se pierda la línea o señal.\n"
        "No se permiten llamadas internacionales ni suscripciones a ningún tipo de promociones que generen costos extras o en la factura de la empresa.\n"
        "No se permiten juegos o usar el plan para entretenimiento/distracciones.\n\n"
        "Para uso personal\n"
        "Para cualquier actividad comercial; Cualquier indicio de que el empleado está haciendo mal uso del teléfono móvil será causa de una sanción disciplinaria obligatoria.\n"
        "Enviar cualquier material indebido por chat.\n"
        "Cualquier material relacionado con actividades ilícitas.\n"
        "Evitar cualquier mal uso del dispositivo móvil de otros.\n\n"
        "SEGURIDAD Y PROTECCIÓN:\n"
        "El empleado debe asegurarse de que el teléfono móvil de la empresa esté protegido todo el tiempo. Debe protegerlo con la ayuda de una contraseña.\n"
        "Es responsabilidad del empleado proteger el teléfono móvil de cualquier tipo de daño, etc. No debe dejarse en ningún lugar sin supervisión.\n"
        "En caso de daño, el personal de mesa asignada, acudirá al personal encargado responsable del daño del teléfono móvil.\n"
        "La información confidencial y personal no debe almacenarse en el teléfono móvil.\n"
    )
    pdf.multi_cell(0, 6, politicas)

    # --- PAGINA 3: OBJETIVOS ---
    pdf.add_page()
    # if logo_path:
    #     pdf.image(logo_path, x=10, y=8, w=28)
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 18, '', ln=True)
    pdf.cell(0, 8, "OBJETIVOS", ln=True, align='C')
    pdf.ln(4)
    pdf.set_font("Arial", size=10)
    objetivos = (
        "- MEJORAR COMUNICACIÓN GENERAL EN LA EMPRESA\n"
        "- ATENCIÓN AL CLIENTE DE PRIMERA Y DE FORMA INMEDIATA\n"
        "- COMUNICAR CARTERA DE CLIENTES HEREDADAS\n"
        "- AGENDAR NUEVOS CLIENTES\n"
        "- BRINDAR COTIZACIONES Y GESTIONAR VENTAS DE FORMA EFICAZ Y RÁPIDA\n"
        "- AUMENTAR VENTAS EN NUESTRA TIENDA Y EMPRESA\n"
        "- MEJORAR IMAGEN PROFESIONAL DE COMUNICACIÓN\n"
        "- USO DE APLICACIONES QUE UTILIZAMOS EN LA EMPRESA COMO CANALES DE VENTA, COMUNICACIÓN Y APOYO\n"
        "- ENVIAR PROMOCIONES Y DIFUSIÓN FRECUENTE\n"
        "- LLAMAR A PROVEEDORES PARA GESTIÓN RÁPIDA DE VENTAS POR ENCARGO A CLIENTE DETALLE\n"
        "- ACCESO A SISTEMA PARA PRECIOS/VENTAS/COTIZACIONES\n"
        "- ACCESO A PAGINA WEB/FACEBOOK/SISTEMA/WHATSAPP Y CORREO DE NUESTRA TIENDA Y EMPRESA KM MOTOS\n"
        "- MAYOR ALCANCE GENERAL\n"
        "- FACILITAR ACCESOS PARA USO CORRECTO EN LA EMPRESA\n"
        "- MEJORAR NÚMEROS DE VENTAS PERSONALES Y SOBREPASAR METAS DE FORMA CONSTANTE\n"
    )
    pdf.multi_cell(0, 6, objetivos)

    import io
    buffer = io.BytesIO()
    pdf.output(buffer)
    return buffer.getvalue()
def generar_responsiva_pdf(
    empleado,
    identidad,
    departamento,
    dispositivo,
    marca,
    imei,
    modelo,
    plan,
    serie,
    area,
    estado,
    cargador,
    funcionalidad,
    numero_corporativo,
    valor_equipo,
    ciudad,
    fecha,
    responsable,
    coordinador,
    logo_path=None
):
    """Genera un PDF de responsiva de equipo de trabajo con los datos proporcionados."""
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    # Logo
    # if logo_path:
    #     pdf.image(logo_path, x=10, y=8, w=28)
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 18, '', ln=True)  # Espacio tras logo
    pdf.cell(0, 8, "ASIGNACIÓN DE EQUIPO DE TRABAJO", ln=True, align='C')
    pdf.ln(6)
    pdf.set_font("Arial", size=11)
    pdf.multi_cell(0, 8, f"Yo ____________________________, con tarjeta de identidad número ____________________________, del departamento de {departamento},\n\nestoy recibiendo por parte de MOTOS (OPERACIONES), el siguiente equipo de trabajo con las siguientes características:")
    pdf.ln(2)
    pdf.set_font("Arial", "B", 11)
    pdf.cell(0, 7, f"Dispositivo: {dispositivo}", ln=True)
    pdf.cell(0, 7, f"Marca: {marca}", ln=True)
    pdf.set_font("Arial", size=11)
    pdf.cell(0, 7, f"Imei: {imei}", ln=True)
    pdf.set_font("Arial", "B", 11)
    pdf.cell(0, 7, f"Modelo: {modelo}    Plan: {plan}", ln=True)
    pdf.cell(0, 7, f"Serie: {serie}    Área: {area}", ln=True)
    pdf.cell(0, 7, f"Estado: {estado}    Cargador: {cargador}", ln=True)
    pdf.cell(0, 7, f"Funcionalidad: {funcionalidad}    Número corporativo: {numero_corporativo}", ln=True)
    pdf.ln(2)
    pdf.set_font("Arial", size=11)
    pdf.multi_cell(0, 7, "El cual recibo en perfectas condiciones, y el cual queda bajo mi responsabilidad, comprometiéndome al uso adecuado del mismo, y a entregarlo al retirarme de la empresa. De comprobarse que por negligencia o mal manejo lo dañara, autorizo a la empresa la deducción respectiva del monto correspondiente.")
    pdf.ln(4)
    pdf.cell(0, 7, f"En la ciudad de {ciudad}, a los ___ días de ________ del {fecha}", ln=True)
    pdf.ln(10)
    pdf.cell(80, 7, "_________________________", ln=0, align='C')
    pdf.cell(30, 7, "", ln=0)
    pdf.cell(80, 7, "_________________________", ln=1, align='C')
    pdf.cell(80, 7, "Responsable", ln=0, align='C')
    pdf.cell(30, 7, "", ln=0)
    pdf.cell(80, 7, "Coordinador Lty Seguridad", ln=1, align='C')
    pdf.ln(6)
    pdf.set_font("Arial", "B", 11)
    pdf.cell(0, 7, f"VALOR DE EQUIPO: L{valor_equipo:,.2f}", ln=True)
    import io
    buffer = io.BytesIO()
    pdf.output(buffer)
    return buffer.getvalue()


def _texto_seguro_pdf(texto):
    """Limpia el texto para evitar errores de codificación en PDF."""
    if not isinstance(texto, str):
        texto = str(texto)
    # Reemplaza caracteres problemáticos básicos
    return texto.replace('\u2013', '-').replace('\u2014', '-').replace('\u2018', "'").replace('\u2019', "'").replace('\u201c', '"').replace('\u201d', '"')
import streamlit as st
from fpdf import FPDF
st.markdown("---")
# st.image("logo_pantalla_app.png", caption="KM MOTOS", use_column_width=True)

def _leer_tasa_cache():
    """Lee el archivo de cache de tasa si existe y lo retorna, o None si no existe."""
    try:
        if os.path.exists(ARCHIVO_TASA_CACHE):
            with open(ARCHIVO_TASA_CACHE, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return None
import streamlit as st
import os
import json
import pandas as pd
import hashlib
import sqlite3
import re
import unicodedata
from datetime import datetime, timedelta

# Definir ARCHIVO_TASA_CACHE si no existe
ARCHIVO_TASA_CACHE = "tasa_usd_hnl_cache.json"

# Definir _guardar_tasa_cache si no existe
def _guardar_tasa_cache(tasa, fuente):
    payload = {
        "tasa": float(tasa),
        "fuente": str(fuente).strip() or "desconocida",
        "actualizado_en": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    try:
        with open(ARCHIVO_TASA_CACHE, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
    except Exception:
        pass
def _pdf_a_bytes(pdf):
    """Convierte un objeto FPDF a bytes para descarga."""
    import io
    buffer = io.BytesIO()
    pdf.output(buffer)
    return buffer.getvalue()
    """Guarda tasa en disco para persistencia entre reinicios."""
    payload = {
        "tasa": float(tasa),
        "fuente": str(fuente).strip() or "desconocida",
        "actualizado_en": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    try:
        with open(ARCHIVO_TASA_CACHE, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
    except Exception:
        pass


def _obtener_tasa_desde_fuentes():
    """Consulta varias fuentes de tasa USD->HNL y devuelve la primera valida."""
    # ...existing code...
    # El siguiente bloque debe estar correctamente indentado dentro de la función o bloque correspondiente
    # (por ejemplo, dentro de un if o función de agregar plan)
    # Ajuste de indentación:
    #
    # if ...:
    # Definir precio_dispositivo antes de usarlo
    precio_dispositivo = 0.0
    try:
        precio_dispositivo = float(st.session_state.get('precio_dispositivo', 0))
    except Exception:
        precio_dispositivo = 0.0
    nuevo_plan = {
        'numero': numero_limpio,
        'operador': (str(operador).strip() or 'TIGO').upper(),
        'nombre_personal': nombre_personal,
        'area': area,
        'departamento': departamento,
        'perfil_profesional': perfil_profesional,
        'valor_usd': valor_usd,
        'tasa_usd_hnl': st.session_state.tasa_usd_hnl,
        'valor_hnl': valor_lempiras,
        'precio_dispositivo': precio_dispositivo,
        'observaciones': observaciones,
        'dispositivo_asignado': dispositivo_asignado,
        'marca': marca_dispositivo,
        'modelo': modelo_dispositivo,
        'serie_dispositivo': serie_dispositivo,
        'imei1': imei1,
        'imei2': imei2,
        'dispositivo_historial': dispositivo_historial,
        'fecha_creacion': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    # Inicializar historial de asignaciones, incluyendo precio del celular
    asignaciones_historial = [{
        'fecha': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'usuario': st.session_state.usuario_actual,
        'nombre_personal': nombre_personal,
        'dispositivo_asignado': dispositivo_asignado,
        'marca': marca_dispositivo,
        'modelo': modelo_dispositivo,
        'serie_dispositivo': serie_dispositivo,
        'precio_dispositivo': precio_dispositivo
    }]
    nuevo_plan['asignaciones_historial'] = asignaciones_historial
    # ...existing code continues here, properly aligned...


def obtener_tasa_usd_hnl(force_refresh=False, max_horas_cache=6):
    """Obtiene tasa USD->HNL con cache local y multiples fuentes online."""
    cache = _leer_tasa_cache()

    if cache and not force_refresh:
        try:
            actualizado = datetime.strptime(cache["actualizado_en"], "%Y-%m-%d %H:%M:%S")
            if datetime.now() - actualizado <= timedelta(hours=max_horas_cache):
                return cache["tasa"], cache["fuente"], cache["actualizado_en"]
        except Exception:
            pass

    tasa, fuente = _obtener_tasa_desde_fuentes()
    if tasa is not None:
        _guardar_tasa_cache(tasa, fuente)
        # _registrar_tasa_historial(tasa, fuente)  # Comentado: función no definida
        actualizado_en = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        return tasa, fuente, actualizado_en

    # Si no hay internet/fuente, se usa el ultimo valor local guardado.
    if cache:
        return cache["tasa"], f"{cache['fuente']} (cache local)", cache["actualizado_en"]

    return None, None, None

# ============ CONFIGURACIÓN DE TEMA ============
def apply_theme(tema):
    """Aplica el tema seleccionado"""
    if tema == "Oscuro":
        st.markdown("""
        <style>
            .stApp { background-color: #0e1117; color: #c9d1d9; }
            .stSidebar { background-color: #161b22; }
        </style>
        """, unsafe_allow_html=True)
    elif tema == "Azul":
        st.markdown("""
        <style>
            .stApp { background-color: #001f3f; color: #e0f0ff; }
            .stSidebar { background-color: #001a33; }
            .stButton, .stDownloadButton { background-color: #0056b3; color: #ffffff; }
        </style>
        """, unsafe_allow_html=True)
    elif tema == "Verde":
        st.markdown("""
        <style>
            .stApp { background-color: #003300; color: #e6ffe6; }
            .stSidebar { background-color: #002a00; }
            .stButton, .stDownloadButton { background-color: #007f0e; color: #ffffff; }
        </style>
        """, unsafe_allow_html=True)
    elif tema == "Rojo":
        st.markdown("""
        <style>
            .stApp { background-color: #330000; color: #ffe6e6; }
            .stSidebar { background-color: #2a0000; }
            .stButton, .stDownloadButton { background-color: #b30000; color: #ffffff; }
        </style>
        """, unsafe_allow_html=True)
    elif tema == "Claro":
        st.markdown("""
        <style>
            .stApp { background-color: #ffffff; color: #24292e; }
            .stSidebar { background-color: #f6f8fa; }
        </style>
        """, unsafe_allow_html=True)


def apply_main_background(mode="Impacto"):
    """Aplica un fondo visual tipo cielo tecnologico con iconos decorativos."""
    is_sutil = str(mode).strip().lower() == "sutil"
    is_corporativo = str(mode).strip().lower() == "corporativo"

    if is_corporativo:
        grad_1 = "radial-gradient(circle at 16% 14%, rgba(158, 203, 230, 0.10) 0, rgba(158, 203, 230, 0.0) 35%)"
        grad_2 = "radial-gradient(circle at 84% 10%, rgba(202, 224, 241, 0.08) 0, rgba(202, 224, 241, 0.0) 32%)"
        grad_3 = "linear-gradient(180deg, #0d2236 0%, #163951 42%, #1f4f70 100%)"
        icon_opacity = "0.07"
        icon_shadow = "0 3px 12px rgba(0, 0, 0, 0.18)"
        sat_size = "52px"
        wifi_size = "62px"
        phone_size = "70px"
    elif is_sutil:
        grad_1 = "radial-gradient(circle at 18% 16%, rgba(157, 211, 255, 0.12) 0, rgba(157, 211, 255, 0.0) 36%)"
        grad_2 = "radial-gradient(circle at 82% 12%, rgba(181, 229, 255, 0.10) 0, rgba(181, 229, 255, 0.0) 34%)"
        grad_3 = "linear-gradient(180deg, #082b48 0%, #0b4476 38%, #1f6ea4 100%)"
        icon_opacity = "0.10"
        icon_shadow = "0 4px 18px rgba(0, 0, 0, 0.20)"
        sat_size = "58px"
        wifi_size = "70px"
        phone_size = "78px"
    else:
        grad_1 = "radial-gradient(circle at 18% 16%, rgba(157, 211, 255, 0.22) 0, rgba(157, 211, 255, 0.0) 36%)"
        grad_2 = "radial-gradient(circle at 82% 12%, rgba(181, 229, 255, 0.20) 0, rgba(181, 229, 255, 0.0) 34%)"
        grad_3 = "linear-gradient(180deg, #083b67 0%, #0a4f8c 30%, #0e6db8 62%, #2c8fd0 100%)"
        icon_opacity = "0.16"
        icon_shadow = "0 6px 24px rgba(0, 0, 0, 0.28)"
        sat_size = "72px"
        wifi_size = "86px"
        phone_size = "94px"

    st.markdown(
        f"""
        <style>
            .stApp {{
                background: transparent !important;
            }}

            [data-testid="stAppViewContainer"] {{
                background:
                    {grad_1},
                    {grad_2},
                    {grad_3};
                background-attachment: fixed;
            }}

            .sky-tech-overlay {{
                position: fixed;
                inset: 0;
                z-index: -1;
                pointer-events: none;
                overflow: hidden;
            }}

            .sky-tech-overlay span {{
                position: absolute;
                color: rgba(235, 248, 255, {icon_opacity});
                text-shadow: {icon_shadow};
            }}

            .sky-tech-sat {{ top: 12%; left: 8%; font-size: {sat_size}; transform: rotate(-12deg); }}
            .sky-tech-wifi {{ top: 34%; right: 10%; font-size: {wifi_size}; transform: rotate(8deg); }}
            .sky-tech-phone {{ bottom: 10%; left: 44%; font-size: {phone_size}; transform: rotate(-6deg); }}
        </style>
        <div class="sky-tech-overlay">
            <span class="sky-tech-sat">🛰️</span>
            <span class="sky-tech-wifi">📶</span>
            <span class="sky-tech-phone">📱</span>
        </div>
        """,
        unsafe_allow_html=True,
    )


def _serializar_planes_cache(planes):
    """Genera una clave estable para cachear calculos derivados de planes."""
    try:
        return json.dumps(planes, ensure_ascii=False, sort_keys=True, default=str)
    except Exception:
        return json.dumps([], ensure_ascii=False)


@st.cache_data(show_spinner=False)
def _construir_dataframe_planes_cache(planes_serializados):
    """Construye DataFrame base de planes reutilizable entre secciones."""
    planes = json.loads(planes_serializados) if planes_serializados else []
    if not isinstance(planes, list):
        planes = []
    df = pd.DataFrame(planes)
    # Asegura que la columna 'precio_dispositivo' siempre exista
    if 'precio_dispositivo' not in df.columns:
        df['precio_dispositivo'] = 0.0
    # Asegura que las columnas nuevas existan
    for col in ['tienda', 'zona', 'empresa']:
        if col not in df.columns:
            df[col] = ''
    return df


@st.cache_data(show_spinner=False)
def _construir_dashboard_payload(planes_serializados):
    """Prepara metricas y graficas del dashboard sin recalcular en cada interaccion."""
    import plotly.express as px

    planes = json.loads(planes_serializados) if planes_serializados else []
    df = pd.DataFrame(planes)
    if df.empty:
        return None

    for columna in ["area", "departamento", "nombre_personal", "numero", "fecha_creacion"]:
        if columna not in df.columns:
            df[columna] = ""

    if "valor_usd" not in df.columns:
        df["valor_usd"] = 0.0
    df["valor_usd"] = pd.to_numeric(df["valor_usd"], errors="coerce").fillna(0.0)
    df["area"] = df["area"].astype(str).str.strip().replace({"": "Sin Área", "nan": "Sin Área", "None": "Sin Área"})
    df["departamento"] = df["departamento"].astype(str).str.strip().replace(
        {"": "Sin Departamento", "nan": "Sin Departamento", "None": "Sin Departamento", "Nan": "Sin Departamento"}
    )

    payload = {
        "total_lineas": int(len(df)),
        "gasto_total": float(df["valor_usd"].sum()),
        "gasto_promedio": float(df["valor_usd"].mean()) if len(df) else 0.0,
        "areas_unicas": int(df["area"].astype(str).nunique()),
        "deptos_unicos": int(df["departamento"].astype(str).nunique()),
        "lineas_asignadas": int(((df["nombre_personal"].notna()) & (df["nombre_personal"].astype(str).str.strip() != "")).sum()),
    }
    payload["lineas_libres"] = max(payload["total_lineas"] - payload["lineas_asignadas"], 0)

    gasto_area = df.groupby("area", dropna=False)["valor_usd"].sum().sort_values(ascending=False)
    if len(gasto_area) > 10:
        top_areas = gasto_area.head(9)
        otras_areas = float(gasto_area.iloc[9:].sum())
        gasto_area = pd.concat([top_areas, pd.Series([otras_areas], index=["Otras áreas"])])
    gasto_area = gasto_area.sort_values(ascending=True)
    fig_area = px.bar(
        y=gasto_area.index,
        x=gasto_area.values,
        orientation="h",
        labels={"x": "USD", "y": "Área"},
        color_discrete_sequence=["#2aa7d8"],
        title=None,
    )
    fig_area.update_layout(
        height=330,
        margin=dict(l=0, r=0, t=30, b=0),
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        font=dict(color="rgba(255,255,255,0.8)", size=11),
        xaxis=dict(showgrid=False, zeroline=False),
        yaxis=dict(showgrid=False),
    )
    fig_area.update_traces(marker_line_width=0)

    lineas_dept = df["departamento"].value_counts()
    if len(lineas_dept) > 7:
        top_dept = lineas_dept.head(6)
        otros_dept = int(lineas_dept.iloc[6:].sum())
        lineas_dept = pd.concat([top_dept, pd.Series([otros_dept], index=["Otros"])]).astype(int)
    else:
        lineas_dept = lineas_dept.astype(int)
    fig_dept = px.pie(
        values=lineas_dept.values,
        names=lineas_dept.index,
        hole=0.46,
        color_discrete_sequence=px.colors.sequential.Blues[::-1],
        title=None,
    )
    fig_dept.update_layout(
        height=330,
        margin=dict(l=0, r=0, t=30, b=0),
        paper_bgcolor="rgba(0,0,0,0)",
        font=dict(color="rgba(255,255,255,0.8)", size=10),
        legend=dict(font=dict(size=9), yanchor="middle", y=0.5),
    )
    fig_dept.update_traces(textposition="inside", textinfo="percent")

    fechas = pd.to_datetime(df["fecha_creacion"], errors="coerce")
    if fechas.notna().any():
        df_sorted = df.assign(_fecha_sort=fechas).sort_values("_fecha_sort").copy()
    else:
        df_sorted = df.copy()
    df_sorted["acumulado"] = df_sorted["valor_usd"].cumsum()
    if len(df_sorted) > 240:
        df_sorted_full = df_sorted.copy()
        idx_ultimo = df_sorted_full.index[-1]
        paso = max(len(df_sorted_full) // 220, 1)
        df_sorted = df_sorted_full.iloc[::paso].copy()
        if df_sorted.index[-1] != idx_ultimo:
            df_sorted = pd.concat([df_sorted, df_sorted_full.loc[[idx_ultimo]]])
    fig_cumsum = px.line(
        x=range(len(df_sorted)),
        y=df_sorted["acumulado"].values,
        labels={"x": "Planes Registrados", "y": "Gasto Acumulado (USD)"},
        title=None,
        markers=False,
    )
    fig_cumsum.update_layout(
        height=320,
        margin=dict(l=0, r=0, t=30, b=0),
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        font=dict(color="rgba(255,255,255,0.8)", size=11),
        hovermode=False,
    )
    fig_cumsum.update_traces(line=dict(color="#00d9ff", width=3))
    fig_cumsum.update_traces(hovertemplate=None)

    total_registros = max(len(df), 1)
    bins = int(max(8, min(22, round(total_registros ** 0.5))))
    fig_scatter = px.histogram(
        df,
        x="valor_usd",
        nbins=bins,
        labels={"valor_usd": "Valor USD", "count": "Cantidad"},
        title=None,
        color_discrete_sequence=["#6fd0ff"],
    )
    fig_scatter.update_layout(
        height=320,
        margin=dict(l=0, r=0, t=30, b=0),
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        font=dict(color="rgba(255,255,255,0.8)", size=10),
        xaxis=dict(showgrid=False, title="Valor USD"),
        yaxis=dict(showgrid=False, title="Cantidad de líneas"),
        bargap=0.08,
        showlegend=False,
        hovermode=False,
    )
    fig_scatter.update_traces(marker_line_width=0, hovertemplate=None)

    payload["fig_area"] = fig_area
    payload["fig_dept"] = fig_dept
    payload["fig_cumsum"] = fig_cumsum
    payload["fig_scatter"] = fig_scatter
    return payload


def _preferencias_usuario_default():
    return {
        "tema": "Claro",
        "idioma": "Español",
        "fondo_visual": "Impacto",
        "estilo_hero": "iPhone",
        "notificaciones": True,
        "columnas_visibles": ["numero", "nombre_personal", "area", "departamento", "valor_usd"],
    }


def _normalizar_preferencias_usuario(preferencias):
    base = _preferencias_usuario_default()
    if isinstance(preferencias, dict):
        base.update({k: v for k, v in preferencias.items() if k in base})

    if base.get("tema") not in ["Oscuro", "Azul", "Verde", "Rojo", "Claro", "Automático"]:
        base["tema"] = "Claro"
    if base.get("idioma") not in ["Español", "Inglés", "Portugués"]:
        base["idioma"] = "Español"
    if base.get("fondo_visual") not in ["Corporativo", "Sutil", "Impacto"]:
        base["fondo_visual"] = "Impacto"
    if base.get("estilo_hero") not in ["iPhone", "Android"]:
        base["estilo_hero"] = "iPhone"
    if not isinstance(base.get("columnas_visibles"), list):
        base["columnas_visibles"] = _preferencias_usuario_default()["columnas_visibles"]
    base["notificaciones"] = bool(base.get("notificaciones", True))

    return base


def _obtener_clave_dispositivo():
    """Genera una clave estable por navegador/equipo a partir de cabeceras HTTP."""
    headers = {}
    try:
        headers = dict(getattr(st.context, "headers", {}) or {})
    except Exception:
        headers = {}

    ua = str(headers.get("User-Agent", "")).strip().lower()
    accept_language = str(headers.get("Accept-Language", "")).split(",")[0].strip().lower()
    ch_ua = str(headers.get("Sec-CH-UA", "")).strip().lower()
    ch_platform = str(headers.get("Sec-CH-UA-Platform", "")).strip().lower()

    fingerprint = "|".join([ua, accept_language, ch_ua, ch_platform]).strip("|")
    if not fingerprint:
        return "dispositivo_generico"

    hash_id = hashlib.sha1(fingerprint.encode("utf-8", errors="ignore")).hexdigest()[:16]
    return f"dev_{hash_id}"


# ============ PERSISTENCIA SQLITE ============
_APP_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SQLITE_STORAGE_FILE = (os.getenv("APP_SQLITE_PATH", "app_storage.db") or "app_storage.db").strip()
if not os.path.isabs(SQLITE_STORAGE_FILE):
    SQLITE_STORAGE_FILE = os.path.join(_APP_BASE_DIR, SQLITE_STORAGE_FILE)
ESCRIBIR_JSON_COMPAT = False


def _abrir_storage_sqlite():
    storage_dir = os.path.dirname(SQLITE_STORAGE_FILE)
    if storage_dir:
        os.makedirs(storage_dir, exist_ok=True)
    conn = sqlite3.connect(SQLITE_STORAGE_FILE, timeout=30)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    return conn


def _inicializar_storage_sqlite(conn):
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS collection_items (
            collection_name TEXT NOT NULL,
            item_index INTEGER NOT NULL,
            payload TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            PRIMARY KEY (collection_name, item_index)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS collection_meta (
            collection_name TEXT PRIMARY KEY,
            initialized INTEGER NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS dict_items (
            dict_name TEXT NOT NULL,
            item_key TEXT NOT NULL,
            payload TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            PRIMARY KEY (dict_name, item_key)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS dict_meta (
            dict_name TEXT PRIMARY KEY,
            initialized INTEGER NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )


def _cargar_json_legacy(path, default_value):
    if not os.path.exists(path):
        return default_value
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return default_value


def _guardar_json_legacy(path, data):
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, default=str, separators=(",", ":"))
    except Exception:
        pass


def _guardar_lista_sqlite(collection_name, data):
    data = data if isinstance(data, list) else []
    marca = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        with _abrir_storage_sqlite() as conn:
            _inicializar_storage_sqlite(conn)
            conn.execute("DELETE FROM collection_items WHERE collection_name = ?", (collection_name,))
            if data:
                conn.executemany(
                    "INSERT INTO collection_items (collection_name, item_index, payload, updated_at) VALUES (?, ?, ?, ?)",
                    [
                        (
                            collection_name,
                            idx,
                            json.dumps(item, ensure_ascii=False, default=str, separators=(",", ":")),
                            marca,
                        )
                        for idx, item in enumerate(data)
                    ],
                )
            conn.execute(
                "INSERT OR REPLACE INTO collection_meta (collection_name, initialized, updated_at) VALUES (?, 1, ?)",
                (collection_name, marca),
            )
    except Exception:
        pass


def _cargar_lista_sqlite(collection_name):
    try:
        with _abrir_storage_sqlite() as conn:
            _inicializar_storage_sqlite(conn)
            meta = conn.execute(
                "SELECT initialized FROM collection_meta WHERE collection_name = ?",
                (collection_name,),
            ).fetchone()
            if not meta:
                return False, []

            rows = conn.execute(
                "SELECT payload FROM collection_items WHERE collection_name = ? ORDER BY item_index ASC",
                (collection_name,),
            ).fetchall()
            data = []
            for row in rows:
                try:
                    data.append(json.loads(row[0]))
                except Exception:
                    continue
            return True, data
    except Exception:
        return False, []


def _guardar_dict_sqlite(dict_name, data):
    data = data if isinstance(data, dict) else {}
    marca = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        with _abrir_storage_sqlite() as conn:
            _inicializar_storage_sqlite(conn)
            conn.execute("DELETE FROM dict_items WHERE dict_name = ?", (dict_name,))
            if data:
                conn.executemany(
                    "INSERT INTO dict_items (dict_name, item_key, payload, updated_at) VALUES (?, ?, ?, ?)",
                    [
                        (
                            dict_name,
                            str(key),
                            json.dumps(value, ensure_ascii=False, default=str, separators=(",", ":")),
                            marca,
                        )
                        for key, value in data.items()
                    ],
                )
            conn.execute(
                "INSERT OR REPLACE INTO dict_meta (dict_name, initialized, updated_at) VALUES (?, 1, ?)",
                (dict_name, marca),
            )
    except Exception:
        pass


def _cargar_dict_sqlite(dict_name):
    try:
        with _abrir_storage_sqlite() as conn:
            _inicializar_storage_sqlite(conn)
            meta = conn.execute(
                "SELECT initialized FROM dict_meta WHERE dict_name = ?",
                (dict_name,),
            ).fetchone()
            if not meta:
                return False, {}

            rows = conn.execute(
                "SELECT item_key, payload FROM dict_items WHERE dict_name = ? ORDER BY item_key ASC",
                (dict_name,),
            ).fetchall()
            data = {}
            for row in rows:
                try:
                    data[str(row[0])] = json.loads(row[1])
                except Exception:
                    continue
            return True, data
    except Exception:
        return False, {}


def _cargar_lista_persistente(collection_name, legacy_file):
    existe_sqlite, data = _cargar_lista_sqlite(collection_name)
    if existe_sqlite:
        return data

    legacy_data = _cargar_json_legacy(legacy_file, [])
    if not isinstance(legacy_data, list):
        legacy_data = []
    _guardar_lista_sqlite(collection_name, legacy_data)
    return legacy_data


def _guardar_lista_persistente(collection_name, data, legacy_file=None):
    _guardar_lista_sqlite(collection_name, data)
    if ESCRIBIR_JSON_COMPAT and legacy_file:
        _guardar_json_legacy(legacy_file, data)


def _cargar_dict_persistente(dict_name, legacy_file):
    existe_sqlite, data = _cargar_dict_sqlite(dict_name)
    if existe_sqlite:
        return data

    legacy_data = _cargar_json_legacy(legacy_file, {})
    if not isinstance(legacy_data, dict):
        legacy_data = {}
    _guardar_dict_sqlite(dict_name, legacy_data)
    return legacy_data


def _guardar_dict_persistente(dict_name, data, legacy_file=None):
    _guardar_dict_sqlite(dict_name, data)
    if ESCRIBIR_JSON_COMPAT and legacy_file:
        _guardar_json_legacy(legacy_file, data)

# ============ GESTIÓN DE USUARIOS ============
class GestorUsuarios:
    def __init__(self, archivo="usuarios.json"):
        self.archivo = archivo
        self.store_name = "usuarios"
        self.usuarios = self.cargar_usuarios()
    
    def cargar_usuarios(self):
        usuarios = _cargar_dict_persistente(self.store_name, self.archivo)
        if usuarios:

            # Normaliza usuarios antiguos que no tienen el permiso explícito.
            for _, data in usuarios.items():
                if "puede_editar" not in data:
                    data["puede_editar"] = data.get("rol", "usuario") in ["administrador", "superadministrador", "usuario"]
                if "permisos" not in data:
                    puede_total = bool(data.get("puede_editar", False) or data.get("rol") in ["administrador", "superadministrador"])
                    data["permisos"] = {
                        "crear": puede_total,
                        "editar": puede_total,
                        "eliminar": puede_total,
                        "importar": puede_total,
                        "exportar": True,
                    }
                data["preferencias"] = _normalizar_preferencias_usuario(data.get("preferencias", {}))
                prefs_dispositivo = data.get("preferencias_por_dispositivo", {})
                if not isinstance(prefs_dispositivo, dict):
                    prefs_dispositivo = {}
                data["preferencias_por_dispositivo"] = {
                    str(k).strip(): _normalizar_preferencias_usuario(v)
                    for k, v in prefs_dispositivo.items()
                    if str(k).strip()
                }

            # Garantiza cuentas administrativas mínimas si el JSON solo trae usuarios de prueba.
            existe_superadmin = any(u.get("rol") == "superadministrador" for u in usuarios.values())
            existe_admin = any(u.get("rol") == "administrador" for u in usuarios.values())

            if not existe_superadmin:
                usuarios["superadmin"] = {
                    "contraseña": "super123",
                    "rol": "superadministrador",
                    "email": "super@empresa.com",
                    "puede_editar": True,
                    "permisos": {
                        "crear": True,
                        "editar": True,
                        "eliminar": True,
                        "importar": True,
                        "exportar": True,
                    },
                    "preferencias": _preferencias_usuario_default(),
                    "preferencias_por_dispositivo": {},
                }

            if not existe_admin:
                usuarios["admin"] = {
                    "contraseña": "admin123",
                    "rol": "administrador",
                    "email": "admin@empresa.com",
                    "puede_editar": True,
                    "permisos": {
                        "crear": True,
                        "editar": True,
                        "eliminar": True,
                        "importar": True,
                        "exportar": True,
                    },
                    "preferencias": _preferencias_usuario_default(),
                    "preferencias_por_dispositivo": {},
                }

            if not existe_superadmin or not existe_admin:
                _guardar_dict_persistente(self.store_name, usuarios, self.archivo)
            return usuarios

        usuarios_base = {
            "superadmin": {
                "contraseña": "super123",
                "rol": "superadministrador",
                "email": "super@empresa.com",
                "puede_editar": True,
                "permisos": {
                    "crear": True,
                    "editar": True,
                    "eliminar": True,
                    "importar": True,
                    "exportar": True,
                },
                "preferencias": _preferencias_usuario_default(),
                "preferencias_por_dispositivo": {},
            },
            "admin": {
                "contraseña": "admin123",
                "rol": "administrador",
                "email": "admin@empresa.com",
                "puede_editar": True,
                "permisos": {
                    "crear": True,
                    "editar": True,
                    "eliminar": True,
                    "importar": True,
                    "exportar": True,
                },
                "preferencias": _preferencias_usuario_default(),
                "preferencias_por_dispositivo": {},
            }
        }
        _guardar_dict_persistente(self.store_name, usuarios_base, self.archivo)
        return usuarios_base
    
    def guardar_usuarios(self):
        _guardar_dict_persistente(self.store_name, self.usuarios, self.archivo)
    
    def crear_usuario(self, usuario, contrasena, rol="usuario", email="", puede_editar=True):
        if usuario in self.usuarios:
            return False, "El usuario ya existe"

        if rol not in ["usuario", "administrador", "superadministrador"]:
            return False, "Rol no válido"
        
        self.usuarios[usuario] = {
            "contraseña": contrasena,
            "rol": rol,
            "email": email,
            "puede_editar": bool(puede_editar),
            "permisos": {
                "crear": bool(puede_editar),
                "editar": bool(puede_editar),
                "eliminar": bool(puede_editar),
                "importar": bool(puede_editar),
                "exportar": True,
            },
            "preferencias": _preferencias_usuario_default(),
            "preferencias_por_dispositivo": {},
            "fecha_creacion": datetime.now().isoformat()
        }
        self.guardar_usuarios()
        return True, "Usuario creado exitosamente"

    def actualizar_usuario(self, usuario, contrasena=None, rol=None, email=None, puede_editar=None, permisos=None, preferencias=None):
        if usuario not in self.usuarios:
            return False, "Usuario no encontrado"

        if rol and rol not in ["usuario", "administrador", "superadministrador"]:
            return False, "Rol no válido"

        if contrasena:
            self.usuarios[usuario]["contraseña"] = contrasena
        if rol:
            self.usuarios[usuario]["rol"] = rol
        if email is not None:
            self.usuarios[usuario]["email"] = email
        if puede_editar is not None:
            self.usuarios[usuario]["puede_editar"] = bool(puede_editar)
        if permisos is not None and isinstance(permisos, dict):
            self.usuarios[usuario]["permisos"] = permisos
        if preferencias is not None and isinstance(preferencias, dict):
            self.usuarios[usuario]["preferencias"] = _normalizar_preferencias_usuario(preferencias)

        self.guardar_usuarios()
        return True, "Usuario actualizado exitosamente"

    def actualizar_preferencias_usuario(self, usuario, preferencias, clave_dispositivo=None):
        if usuario not in self.usuarios:
            return False, "Usuario no encontrado"

        prefs_norm = _normalizar_preferencias_usuario(preferencias)
        self.usuarios[usuario]["preferencias"] = prefs_norm

        if clave_dispositivo:
            mapa = self.usuarios[usuario].get("preferencias_por_dispositivo", {})
            if not isinstance(mapa, dict):
                mapa = {}
            mapa[str(clave_dispositivo)] = prefs_norm
            self.usuarios[usuario]["preferencias_por_dispositivo"] = mapa

        self.guardar_usuarios()
        return True, "Preferencias actualizadas"
    
    def validar_usuario(self, usuario, contrasena):
        if usuario not in self.usuarios:
            return False, "Usuario no encontrado"
        
        if self.usuarios[usuario]["contraseña"] != contrasena:
            return False, "Contraseña incorrecta"
        
        return True, self.usuarios[usuario]
    
    def obtener_usuarios(self):
        return list(self.usuarios.keys())
    
    def eliminar_usuario(self, usuario):
        if usuario in ["admin", "superadmin"]:
            return False, "No puedes eliminar cuentas administrativas protegidas"
        
        if usuario in self.usuarios:
            del self.usuarios[usuario]
            self.guardar_usuarios()
            return True, "Usuario eliminado"
        
        return False, "Usuario no encontrado"

# Configurar página
st.set_page_config(
    page_title="Dashboard - PLANES CORPORATIVOS KM MOTOS",
    layout="wide",
    initial_sidebar_state="expanded"
)

def aplicar_bloqueo_traduccion_global():
    """Fuerza el idioma espanol y desactiva traductores automáticos en la interfaz."""
    st.markdown(
        """
        <script>
            (function() {
                const aplicar = (doc) => {
                    if (!doc) return;

                    const html = doc.documentElement;
                    if (html) {
                        html.setAttribute('lang', 'es');
                        html.setAttribute('translate', 'no');
                        html.classList.add('notranslate');
                    }

                    if (doc.body) {
                        doc.body.setAttribute('translate', 'no');
                        doc.body.classList.add('notranslate');
                    }

                    let metaGoogle = doc.querySelector('meta[name="google"]');
                    if (!metaGoogle) {
                        metaGoogle = doc.createElement('meta');
                        metaGoogle.setAttribute('name', 'google');
                        doc.head.appendChild(metaGoogle);
                    }
                    metaGoogle.setAttribute('content', 'notranslate');

                    let metaContentLanguage = doc.querySelector('meta[http-equiv="content-language"]');
                    if (!metaContentLanguage) {
                        metaContentLanguage = doc.createElement('meta');
                        metaContentLanguage.setAttribute('http-equiv', 'content-language');
                        doc.head.appendChild(metaContentLanguage);
                    }
                    metaContentLanguage.setAttribute('content', 'es');
                };

                const estaTraducido = (doc) => {
                    if (!doc || !doc.documentElement) return false;
                    const html = doc.documentElement;
                    const body = doc.body;
                    const clasesHtml = (html.className || '').toLowerCase();
                    const clasesBody = body ? (body.className || '').toLowerCase() : '';

                    return (
                        clasesHtml.includes('translated') ||
                        clasesBody.includes('translated') ||
                        !!doc.querySelector('iframe.goog-te-banner-frame') ||
                        !!doc.querySelector('.goog-te-banner-frame') ||
                        !!doc.querySelector('.goog-te-combo')
                    );
                };

                const mostrarAvisoTraduccion = (doc) => {
                    if (!doc || !doc.body) return;
                    const idAviso = 'aviso-traduccion-activa-km';
                    const traducido = estaTraducido(doc);
                    let aviso = doc.getElementById(idAviso);

                    if (!traducido) {
                        if (aviso) aviso.remove();
                        return;
                    }

                    if (!aviso) {
                        aviso = doc.createElement('div');
                        aviso.id = idAviso;
                        aviso.setAttribute('translate', 'no');
                        aviso.className = 'notranslate';
                        aviso.style.cssText = [
                            'position:fixed',
                            'right:12px',
                            'bottom:12px',
                            'z-index:2147483647',
                            'max-width:360px',
                            'padding:10px 12px',
                            'border-radius:12px',
                            'border:1px solid rgba(255,220,120,0.55)',
                            'background:linear-gradient(180deg, rgba(52,35,10,0.96), rgba(35,22,5,0.96))',
                            'color:#ffe8b3',
                            'font:600 13px/1.35 "Segoe UI",sans-serif',
                            'box-shadow:0 6px 18px rgba(0,0,0,0.35)'
                        ].join(';');
                        aviso.textContent = 'Aviso: traduccion del navegador detectada. Desactiva "Traducir pagina" para evitar cambios de botones y textos.';
                        doc.body.appendChild(aviso);
                    }
                };

                const aplicarYVerificar = (doc) => {
                    aplicar(doc);
                    mostrarAvisoTraduccion(doc);
                };

                aplicarYVerificar(document);
                try { aplicarYVerificar(window.parent.document); } catch (e) {}

                setTimeout(() => { aplicarYVerificar(document); }, 250);
                setTimeout(() => { try { aplicarYVerificar(window.parent.document); } catch (e) {} }, 600);
                setInterval(() => {
                    aplicarYVerificar(document);
                    try { aplicarYVerificar(window.parent.document); } catch (e) {}
                }, 2000);

                const observer = new MutationObserver(() => {
                    aplicarYVerificar(document);
                    try { aplicarYVerificar(window.parent.document); } catch (e) {}
                });
                observer.observe(document.documentElement, { childList: true, subtree: true, attributes: true });
            })();
        </script>
        """,
        unsafe_allow_html=True,
    )

aplicar_bloqueo_traduccion_global()

# Inicializar gestor de usuarios (cacheado: solo lee usuarios.json una vez por sesion de servidor)
GESTOR_USUARIOS_CACHE_VERSION = "2026-04-09-sqlite-v1"

@st.cache_resource
def _crear_gestor_usuarios(_version_key):
    return GestorUsuarios()

gestor_usuarios = _crear_gestor_usuarios(GESTOR_USUARIOS_CACHE_VERSION)

# Funciones de persistencia para planes
PLANES_FILE = "planes.json"
PLANES_STORE = "planes"
MOVIMIENTOS_STORE = "movimientos"
EMPLEADOS_STORE = "empleados"
RECORDARME_STORE = "recordarme"

def cargar_planes():
    return _cargar_lista_persistente(PLANES_STORE, PLANES_FILE)


MOVIMIENTOS_FILE = "movimientos.json"
EMPLEADOS_FILE = "empleados.json"
RECORDARME_FILE = "recordarme.json"
DEBUG_IMPORT_FILE = "debug_importacion.log"

def cargar_movimientos():
    return _cargar_lista_persistente(MOVIMIENTOS_STORE, MOVIMIENTOS_FILE)


def guardar_planes():
    try:
        _guardar_lista_persistente(PLANES_STORE, st.session_state.planes, PLANES_FILE)
    except Exception as e:
        st.error(f"❌ No se pudo guardar planes: {e}")


def guardar_movimientos():
    try:
        _guardar_lista_persistente(MOVIMIENTOS_STORE, st.session_state.movimientos, MOVIMIENTOS_FILE)
    except Exception:
        pass


def cargar_empleados():
    return _cargar_lista_persistente(EMPLEADOS_STORE, EMPLEADOS_FILE)


def guardar_empleados():
    try:
        _guardar_lista_persistente(EMPLEADOS_STORE, st.session_state.empleados, EMPLEADOS_FILE)
    except Exception as e:
        st.error(f"❌ No se pudo guardar empleados: {e}")


def cargar_recordarme():
    data = _cargar_dict_persistente(RECORDARME_STORE, RECORDARME_FILE)
    return data if isinstance(data, dict) else {}


def guardar_recordarme():
    try:
        _guardar_dict_persistente(RECORDARME_STORE, st.session_state.credenciales_recordadas, RECORDARME_FILE)
    except Exception as e:
        st.error(f"❌ No se pudo guardar Recordarme: {e}")


def registrar_debug_importacion(evento, detalle):
    try:
        with open(DEBUG_IMPORT_FILE, 'a', encoding='utf-8') as f:
            marca = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            f.write(f"[{marca}] {evento}: {detalle}\n")
    except Exception:
        pass


# Inicializar sesión
if 'planes' not in st.session_state:
    st.session_state.planes = cargar_planes()

if 'usuario_actual' not in st.session_state:
    st.session_state.usuario_actual = None

if 'puede_editar' not in st.session_state:
    st.session_state.puede_editar = False

if 'archivo_cargado' not in st.session_state:
    st.session_state.archivo_cargado = False

if 'tema' not in st.session_state:
    st.session_state.tema = "Oscuro"

if 'movimientos' not in st.session_state:
    st.session_state.movimientos = cargar_movimientos()

if 'mostrar_numeros_corporativos' not in st.session_state:
    st.session_state.mostrar_numeros_corporativos = False

if 'confirmar_limpiar_datos' not in st.session_state:
    st.session_state.confirmar_limpiar_datos = False

if 'empleados' not in st.session_state:
    st.session_state.empleados = cargar_empleados()

if 'credenciales_recordadas' not in st.session_state:
    st.session_state.credenciales_recordadas = cargar_recordarme()

if 'dashboard_modo_rendimiento' not in st.session_state:
    st.session_state.dashboard_modo_rendimiento = True


def registrar_movimiento(tipo, detalle):
    st.session_state.movimientos.append({
        "fecha": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "usuario": st.session_state.usuario_actual or "Anonimo",
        "tipo": tipo,
        "detalle": detalle
    })

    guardar_movimientos()


def tiene_permiso(accion):
    """Valida permisos finos por accion para el usuario en sesion."""
    if st.session_state.get("rol") in ["administrador", "superadministrador"]:
        return True

    usuario = st.session_state.get("usuario_actual")
    data = gestor_usuarios.usuarios.get(usuario, {})
    permisos = data.get("permisos", {})
    return bool(permisos.get(accion, False))


def resumen_numeros_corporativos(planes):
    """Calcula totales de lineas corporativas (totales, asignadas y disponibles)."""
    total_lineas = len(planes)
    lineas_asignadas = 0

    for plan in planes:
        nombre = str(plan.get("nombre_personal", "")).strip()
        if nombre and nombre.lower() not in ["none", "nan"]:
            lineas_asignadas += 1

    lineas_disponibles = max(total_lineas - lineas_asignadas, 0)
    return total_lineas, lineas_asignadas, lineas_disponibles


def _normalizar_texto_columna(valor):
    texto = str(valor).strip().lower()
    if not texto or texto in ["nan", "none"]:
        return ""
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = texto.replace(" ", "_")
    texto = re.sub(r"[^a-z0-9_]+", "", texto)
    return texto


def normalizar_numero_telefonico(numero):
    """Normaliza numeros corporativos a formato +504XXXXXXXX cuando aplica."""
    texto = str(numero).strip()
    if not texto or texto.lower() in ["nan", "none"]:
        return ""

    digitos = "".join(ch for ch in texto if ch.isdigit())
    if digitos.startswith("504") and len(digitos) == 11:
        return f"+{digitos}"
    if len(digitos) == 8:
        return f"+504{digitos}"
    if texto.startswith("+") and digitos:
        return f"+{digitos}"
    return texto


def normalizar_operador(valor):
    """Usa TIGO como operador por defecto cuando el dato viene vacio o invalido."""
    texto = str(valor).strip()
    if not texto or texto.lower() in ["nan", "none"]:
        return "TIGO"
    return texto.upper()


def normalizar_planes_cargados(planes):
    """Corrige registros persistidos para asegurar operador consistente."""
    if not isinstance(planes, list):
        return [], False

    planes_normalizados = []
    hubo_cambios = False

    for plan in planes:
        if not isinstance(plan, dict):
            continue

        plan_normalizado = dict(plan)
        operador_normalizado = normalizar_operador(plan_normalizado.get("operador", "TIGO"))
        if plan_normalizado.get("operador") != operador_normalizado:
            hubo_cambios = True
        plan_normalizado["operador"] = operador_normalizado
        planes_normalizados.append(plan_normalizado)

    return planes_normalizados, hubo_cambios


_planes_normalizados, _planes_corregidos = normalizar_planes_cargados(st.session_state.planes)
st.session_state.planes = _planes_normalizados
if _planes_corregidos:
    guardar_planes()


def preparar_dataframe_importacion(df):
    """Normaliza encabezados para soportar plantillas CSV/Excel/Google Sheets."""
    if df is None or df.empty:
        return df

    def _normalizar_columnas(frame):
        aliases = {
            "numero": "numero",
            "numero_de_linea": "numero",
            "numero_linea": "numero",
            "telefono": "numero",
            "linea": "numero",
            "operador": "operador",
            "empresa": "operador",
            "proveedor": "operador",
            "compania": "operador",
            "carrier": "operador",
            "nombre": "nombre_personal",
            "nombres": "nombre_personal",
            "nombre1": "nombre_personal",
            "nombre2": "nombre_personal",
            "nombre_completo": "nombre_personal",
            "nombre_compl": "nombre_personal",
            "nom": "nombre_personal",
            "asignado_a": "nombre_personal",
            "colaborador": "nombre_personal",
            "nombre_personal": "nombre_personal",
            "area": "area",
            "departamento": "departamento",
            "perfil": "perfil_profesional",
            "perfil_profesional": "perfil_profesional",
            "cargo": "perfil_profesional",
            "valor_usd": "valor_usd",
            "valor": "valor_usd",
            "valor_": "valor_usd",
            "valor__": "valor_usd",
            "valorusd": "valor_usd",
            "valor_en_usd": "valor_usd",
            "costo_usd": "valor_usd",
            "costo": "valor_usd",
            "usd": "valor_usd",
            "dolares": "valor_usd",
            "plan_dolares": "valor_usd",
            "tasa_usd_hnl": "tasa_usd_hnl",
            "tasa": "tasa_usd_hnl",
            "valor_hnl": "valor_hnl",
            "lempiras": "valor_hnl",
            "observaciones": "observaciones",
            "dispositivo_asignado": "dispositivo_asignado",
            "marca": "marca",
            "modelo": "modelo",
            "serie_dispositivo": "serie_dispositivo",
            "imei1": "imei1",
            "imei2": "imei2",
        }

        renames = {}
        for col in frame.columns:
            normalizada = _normalizar_texto_columna(col)
            if not normalizada:
                continue
            renames[col] = aliases.get(normalizada, normalizada)
        frame = frame.rename(columns=renames)
        frame.columns = [str(c).strip().lower() for c in frame.columns]
        return frame

    df = _normalizar_columnas(df)

    # Detectar y eliminar columna índice tipo 'No.' al principio si existe
    try:
        primera_col = str(df.columns[0]).strip().lower()
        if primera_col in ("no", "no.", "n.", "n", "#"):
            vals = df.iloc[:, 0].astype(str).str.strip()
            num_digitos = vals.str.fullmatch(r"\d+").sum()
            if len(vals) > 0 and (num_digitos / len(vals) > 0.7):
                df = df.iloc[:, 1:].copy()
    except Exception:
        pass

    # Mantener compatibilidad: detectar fila de encabezados si 'numero' no existe
    if "numero" not in df.columns:
        max_filas_header = min(12, len(df))
        for i in range(max_filas_header):
            candidatos = [_normalizar_texto_columna(v) for v in df.iloc[i].tolist()]
            if "numero" in candidatos and (
                "area" in candidatos or "departamento" in candidatos or "valor_usd" in candidatos
            ):
                nuevos_headers = [str(v).strip() for v in df.iloc[i].tolist()]
                df = df.iloc[i + 1 :].copy()
                df.columns = nuevos_headers
                df = _normalizar_columnas(df)
                break

    if not df.empty:
        primera_col = df.columns[0]
        primer_valor = df[primera_col].astype(str).str.strip()
        mask_meta = primer_valor.str.startswith("#") | (primer_valor == "---")
        df = df[~mask_meta].reset_index(drop=True)

    # Heurística para detectar columna de nombre si no se normalizó correctamente
    if "nombre_personal" not in df.columns:
        # Preferir columnas cuyo nombre normalizado contenga 'nombre'
        cols_nombre = [c for c in df.columns if 'nombre' in str(c).lower()]
        if cols_nombre:
            df = df.rename(columns={cols_nombre[0]: 'nombre_personal'})
        else:
            # Heurística por contenido: buscar columna con valores que parezcan nombres (>=2 palabras, solo letras)
            def score_nombre(serie):
                s = serie.dropna().astype(str).str.strip()
                if s.empty:
                    return 0.0
                sample = s.head(20)
                def es_nombre(v):
                    v = v.strip()
                    if len(v) < 5:
                        return False
                    if any(ch.isdigit() for ch in v):
                        return False
                    parts = v.split()
                    return len(parts) >= 2 and all(part.isalpha() or all(ch.isalpha() for ch in part) for part in parts)
                matches = sample.apply(lambda v: es_nombre(v))
                return float(matches.sum()) / len(sample)

            mejor = None
            mejor_score = 0.0
            for c in df.columns:
                try:
                    sc = score_nombre(df[c])
                except Exception:
                    sc = 0.0
                if sc > mejor_score:
                    mejor_score = sc
                    mejor = c
            if mejor_score >= 0.6 and mejor is not None:
                df = df.rename(columns={mejor: 'nombre_personal'})

    # Normalizar nombres de columnas finales a minusculas y sin espacios incidentales
    df.columns = [str(c).strip().lower() for c in df.columns]

    return df


def preparar_dataframe_empleados(df):
    """Normaliza encabezados para archivos de empleados sin eliminar filas."""
    if df is None or df.empty:
        return df

    def _normalizar_col(col):
        txt = _normalizar_texto_columna(col)
        aliases = {
            'nombre': 'nombre_personal',
            'nombres': 'nombre_personal',
            'nombre1': 'nombre_personal',
            'nombre2': 'nombre_personal',
            'nombre_completo': 'nombre_personal',
            'nombre_compl': 'nombre_personal',
            'nombre_compl.': 'nombre_personal',
            'nom': 'nombre_personal',
            'perfil': 'perfil_profesional',
            'cargo': 'perfil_profesional',
            'perfil_profesional': 'perfil_profesional',
            'area': 'area',
            'departamento': 'departamento',
            'tienda': 'tienda',
            'zona': 'zona',
            'empresa': 'empresa',
            'operador': 'operador',
        }
        return aliases.get(txt, txt)

    renames = {}
    for col in df.columns:
        renames[col] = _normalizar_col(col)

    df = df.rename(columns=renames)
    df.columns = [str(c).strip().lower() for c in df.columns]
    return df


def analizar_importacion_lineas(df_nuevas, planes_actuales, tasa_default):
    """Calcula resumen de importación sin persistir cambios."""
    resumen = {
        "filas": 0,
        "nuevas": 0,
        "duplicadas": 0,
        "invalidas": 0,
    }

    if df_nuevas is None or df_nuevas.empty:
        return resumen

    resumen["filas"] = len(df_nuevas.index)
    existentes = {
        normalizar_numero_telefonico(p.get("numero", ""))
        for p in planes_actuales
    }

    for _, row in df_nuevas.iterrows():
        plan = construir_plan_desde_fila(row.to_dict(), tasa_default)
        if not plan:
            resumen["invalidas"] += 1
            continue

        numero_norm = normalizar_numero_telefonico(plan.get("numero", ""))
        if not numero_norm or numero_norm in existentes:
            resumen["duplicadas"] += 1
            continue

        resumen["nuevas"] += 1
        existentes.add(numero_norm)

    return resumen


def normalizar_y_deduplicar_planes(planes):
    """Normaliza números y elimina duplicados exactos por número normalizado."""
    vistos = set()
    planes_limpios = []
    cambios_numero = 0
    duplicados_eliminados = 0

    for plan in planes:
        plan_nuevo = dict(plan)
        original = str(plan_nuevo.get("numero", "")).strip()
        normalizado = normalizar_numero_telefonico(original)
        if original != normalizado:
            cambios_numero += 1
        plan_nuevo["numero"] = normalizado

        if not normalizado:
            duplicados_eliminados += 1
            continue

        if normalizado in vistos:
            duplicados_eliminados += 1
            continue

        vistos.add(normalizado)
        planes_limpios.append(plan_nuevo)

    return planes_limpios, cambios_numero, duplicados_eliminados


def procesar_importacion_lineas_nuevas(archivo_nuevas_lineas, hoja_nuevas_lineas):
    if archivo_nuevas_lineas is None:
        st.warning("Selecciona un archivo para importar.")
        return

    try:
        nombre_archivo = archivo_nuevas_lineas.name.lower()
        registrar_debug_importacion(
            "inicio",
            f"archivo={archivo_nuevas_lineas.name}; hoja={hoja_nuevas_lineas}; puede_editar={st.session_state.puede_editar}; planes_actuales={len(st.session_state.planes)}",
        )
        archivo_nuevas_lineas.seek(0)
        if nombre_archivo.endswith(".csv"):
            df_nuevas = pd.read_csv(archivo_nuevas_lineas)
        elif nombre_archivo.endswith((".xlsx", ".xls")):
            if not hoja_nuevas_lineas:
                st.warning("Selecciona la hoja de Excel.")
                return
            archivo_nuevas_lineas.seek(0)
            df_nuevas = pd.read_excel(archivo_nuevas_lineas, sheet_name=hoja_nuevas_lineas)
        else:
            st.error("Formato no soportado.")
            return

        df_nuevas = preparar_dataframe_importacion(df_nuevas)
        registrar_debug_importacion("columnas_detectadas", f"columnas={list(df_nuevas.columns)}; filas={len(df_nuevas.index)}")

        if "numero" not in df_nuevas.columns:
            registrar_debug_importacion("error_columnas", "No se detecto columna numero")
            st.error("No se detecto la columna 'numero'. Verifica encabezados del archivo.")
            return

        tasa_default = float(st.session_state.get("tasa_usd_hnl", 24.0))
        existentes = {
            normalizar_numero_telefonico(p.get("numero", ""))
            for p in st.session_state.planes
        }
        nuevos_planes = []
        duplicados = 0
        invalidas = 0

        for _, row in df_nuevas.iterrows():
            plan = construir_plan_desde_fila(row.to_dict(), tasa_default)
            if not plan:
                invalidas += 1
                continue
            numero_norm = normalizar_numero_telefonico(plan.get("numero", ""))
            if not numero_norm or numero_norm in existentes:
                duplicados += 1
                continue
            plan["numero"] = numero_norm
            nuevos_planes.append(plan)
            existentes.add(numero_norm)

        if nuevos_planes:
            st.session_state.planes.extend(nuevos_planes)
            guardar_planes()
            registrar_movimiento(
                "Importar lineas nuevas", f"{len(nuevos_planes)} lineas agregadas; {duplicados} duplicadas; {invalidas} invalidas"
            )
            st.success(f"Se agregaron {len(nuevos_planes)} lineas nuevas. Duplicadas: {duplicados}. Invalidas: {invalidas}.")
            st.rerun()
        else:
            registrar_debug_importacion("sin_agregados", f"agregadas=0; duplicadas={duplicados}; invalidas={invalidas}")
            registrar_movimiento("Importar lineas nuevas", f"0 lineas agregadas; {duplicados} duplicadas; {invalidas} invalidas")
            st.info(f"No se agregaron lineas nuevas. Duplicadas: {duplicados}. Invalidas: {invalidas}.")
    except Exception as e:
        registrar_debug_importacion("exception", str(e))
        st.error(f"Error en importacion de lineas nuevas: {e}")


def construir_plan_desde_fila(row, tasa_default):
    """Construye un plan minimo a partir de una fila importada."""
    numero = normalizar_numero_telefonico(row.get("numero", ""))
    if not numero:
        return None

    nombre = str(row.get("nombre_personal", "")).strip()
    if nombre.lower() in ["nan", "none"]:
        nombre = ""

    valor_usd = pd.to_numeric(row.get("valor_usd", 0.0), errors="coerce")
    if pd.isna(valor_usd):
        valor_usd = 0.0

    tasa = pd.to_numeric(row.get("tasa_usd_hnl", tasa_default), errors="coerce")
    if pd.isna(tasa) or tasa <= 0:
        tasa = tasa_default

    valor_hnl = pd.to_numeric(row.get("valor_hnl", valor_usd * tasa), errors="coerce")
    if pd.isna(valor_hnl):
        valor_hnl = valor_usd * tasa

    return {
        "numero": numero,
        "operador": normalizar_operador(row.get("operador", "TIGO")),
        "nombre_personal": nombre,
        "area": str(row.get("area", "Sin Área")).strip() or "Sin Área",
        "departamento": str(row.get("departamento", "Sin Departamento")).strip() or "Sin Departamento",
        "perfil_profesional": str(row.get("perfil_profesional", "")).strip(),
        "valor_usd": float(valor_usd),
        "tasa_usd_hnl": float(tasa),
        "valor_hnl": float(valor_hnl),
        "observaciones": str(row.get("observaciones", "")).strip(),
        "dispositivo_asignado": str(row.get("dispositivo_asignado", "")).strip(),
        "marca": str(row.get("marca", "")).strip(),
        "modelo": str(row.get("modelo", "")).strip(),
        "serie_dispositivo": str(row.get("serie_dispositivo", "")).strip(),
        "imei1": str(row.get("imei1", "")).strip(),
        "imei2": str(row.get("imei2", "")).strip(),
        "dispositivo_historial": row.get("dispositivo_historial", []) if isinstance(row.get("dispositivo_historial", []), list) else [],
        "fecha_creacion": str(row.get("fecha_creacion", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))),
    }


def render_hero_principal(usuario, rol):
    """Renderiza cabecera principal tipo hero con animacion suave."""
    st.markdown(f"""
    <div style='width: 230px; height: 470px; margin: 28px auto 20px auto; border-radius: 38px; border: 4px solid #111; background: #111; box-shadow: 0 8px 32px rgba(0,0,0,0.18); position: relative; display: flex; flex-direction: column;'>
        <!-- Notch superior tipo iPhone -->
        <div style='height: 36px; width: 100%; position: relative; display: flex; align-items: center; justify-content: center;'>
            <div style='width: 64px; height: 18px; background: #111; border-radius: 12px; box-shadow: 0 1px 4px rgba(0,0,0,0.18); display: flex; align-items: center; justify-content: center; position: absolute; top: 8px;'>
                <div style='width: 36px; height: 10px; background: #222; border-radius: 6px; margin-right: 6px;'></div>
                <div style='width: 7px; height: 7px; background: #222; border-radius: 50%; border: 2px solid #0f0;'></div>
            </div>
        </div>
        <!-- Pantalla blanca -->
        <div style='background: #fff; border-radius: 24px; margin: 0 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.07); padding: 32px 8px 18px 8px; flex: 1 1 auto; min-height: 0; text-align: center; display: flex; flex-direction: column; justify-content: center;'>
            <div style='font-size: 1.11em; font-weight: bold; letter-spacing: 1px; color: #1a2233; margin-bottom: 2px;'>PLANES CORPORATIVOS</div>
            <div style='font-size: 1em; color: #2d3a4d; margin-bottom: 6px;'>KM MOTOS</div>
            <div style='font-size: 0.93em; color: #3a4a5d; margin-bottom: 8px;'>Control centralizado de líneas y asignaciones corporativas</div>
            <div style='font-size: 0.92em; color: #4a5a6d;'>Usuario: <b>{usuario}</b> | Rol: <b>{rol}</b></div>
        </div>
    </div>
    """, unsafe_allow_html=True)


def render_kpi_cards(total_lineas, gasto_total, areas_diff, deptos_diff):
    """Renderiza KPIs en tarjetas premium estilo glass."""
    st.markdown(f"**Total de Líneas:** {total_lineas}  |  **Gasto Total (USD):** ${gasto_total:,.2f}  |  **Áreas Diferentes:** {areas_diff}  |  **Departamentos:** {deptos_diff}")


def render_semaforo_lineas(total_lineas, lineas_libres):
    """Semaforo visual para disponibilidad de lineas."""
    if total_lineas <= 0:
        ratio_libres = 0
    else:
        ratio_libres = lineas_libres / total_lineas

    if ratio_libres >= 0.20:
        estado = "VERDE"
        color = "#2ecc71"
        mensaje = "Disponibilidad saludable de lineas."
    elif ratio_libres >= 0.10:
        estado = "AMARILLO"
        color = "#f1c40f"
        mensaje = "Disponibilidad moderada; conviene monitorear."
    else:
        st.markdown(f"**Total de Líneas:** {total_lineas}  |  **Gasto Total (USD):** ${gasto_total:,.2f}  |  **Áreas Diferentes:** {areas_diff}  |  **Departamentos:** {deptos_diff}")

    pdf.set_font("Arial", "B", 11)
    pdf.cell(0, 8, "Top Areas por Gasto", ln=True)
    pdf.set_font("Arial", size=10)
    gasto_area = df.groupby('area')['valor_usd'].sum().sort_values(ascending=False).head(10)
    for area, gasto in gasto_area.items():
        pdf.cell(0, 7, f"- {area}: ${gasto:,.2f}", ln=True)

    return _pdf_a_bytes(pdf)


def aplicar_estilo_tabla_profesional():
    """Aplica estilo visual profesional a tablas de Streamlit."""
    st.markdown(
        """
        <style>
            [data-testid="stDataFrame"] {
                border: 1px solid rgba(255, 255, 255, 0.22);
                border-radius: 12px;
                box-shadow: 0 10px 24px rgba(0, 0, 0, 0.20);
                overflow: hidden;
            }

            [data-testid="stDataFrame"] thead tr th {
                background: linear-gradient(180deg, rgba(21, 46, 74, 0.95), rgba(17, 38, 63, 0.95));
                color: #eaf4ff;
                font-weight: 700;
                letter-spacing: 0.2px;
            }
        </style>
        """,
        unsafe_allow_html=True,
    )


def construir_tabla_planes_profesional(df_tabla):
    """Normaliza y ordena columnas para una vista ejecutiva de la tabla de planes."""
    df_vista = df_tabla.copy()

    if 'nombre_personal' in df_vista.columns:
        asignado = df_vista['nombre_personal'].astype(str).str.strip()
        df_vista['estado_linea'] = asignado.apply(
            lambda x: 'Asignada' if x and x.lower() not in ['none', 'nan'] else 'Libre'
        )
    else:
        df_vista['estado_linea'] = 'Libre'

    if 'operador' not in df_vista.columns:
        df_vista['operador'] = 'TIGO'
    else:
        df_vista['operador'] = df_vista['operador'].apply(normalizar_operador)

    columnas_preferidas = [
        'numero',
        'operador',
        'estado_linea',
        'nombre_personal',
        'perfil_profesional',
        'area',
        'departamento',
        'tienda',
        'zona',
        'empresa',
        'dispositivo_asignado',
        'marca',
        'modelo',
        'serie_dispositivo',
        'precio_dispositivo',
        'valor_usd',
        'valor_hnl',
        'fecha_creacion',
    ]

    columnas_existentes = [c for c in columnas_preferidas if c in df_vista.columns]
    df_vista = df_vista[columnas_existentes]
    df_vista.insert(0, 'item', range(1, len(df_vista) + 1))
    return df_vista


def construir_resumen_lineas(df_filtrado, filtro_area=None, filtro_dept=None):
    """Construye un resumen final con total de lineas y total en lempiras del conjunto visible."""
    filtro_area = filtro_area or []
    filtro_dept = filtro_dept or []
    total_lineas = len(df_filtrado)
    filas_resumen = []

    if 'valor_hnl' in df_filtrado.columns:
        hnl_series = pd.to_numeric(df_filtrado['valor_hnl'], errors='coerce').fillna(0)
    else:
        hnl_series = pd.Series([0] * total_lineas, index=df_filtrado.index, dtype='float64')
    total_hnl = float(hnl_series.sum())

    if filtro_area and 'area' in df_filtrado.columns:
        df_area = df_filtrado.copy()
        df_area['_valor_hnl_num'] = hnl_series
        resumen_area = (
            df_area.groupby(
                df_area['area'].fillna('Sin area').astype(str).str.strip().replace('', 'Sin area'),
                dropna=False
            )
            .agg(Lineas=('area', 'size'), Total_HNL=('_valor_hnl_num', 'sum'))
            .sort_index()
        )
        for area, fila in resumen_area.iterrows():
            filas_resumen.append({
                'Resumen': 'Area filtrada',
                'Detalle': area,
                'Lineas': int(fila['Lineas']),
                'Total HNL': float(fila['Total_HNL']),
            })

    if filtro_dept and 'departamento' in df_filtrado.columns:
        df_dept = df_filtrado.copy()
        df_dept['_valor_hnl_num'] = hnl_series
        resumen_dept = (
            df_dept.groupby(
                df_dept['departamento']
                .fillna('Sin departamento')
                .astype(str)
                .str.strip()
                .replace('', 'Sin departamento'),
                dropna=False
            )
            .agg(Lineas=('departamento', 'size'), Total_HNL=('_valor_hnl_num', 'sum'))
            .sort_index()
        )
        for dept, fila in resumen_dept.iterrows():
            filas_resumen.append({
                'Resumen': 'Departamento filtrado',
                'Detalle': dept,
                'Lineas': int(fila['Lineas']),
                'Total HNL': float(fila['Total_HNL']),
            })

    if not (filtro_area or filtro_dept):
        filas_resumen.append({
            'Resumen': 'TOTAL HNL',
            'Detalle': 'Valor total visible en lempiras',
            'Lineas': int(total_lineas),
            'Total HNL': total_hnl,
        })
    else:
        filas_resumen.append({
            'Resumen': 'TOTAL HNL FILTRADO',
            'Detalle': 'Valor total en lempiras del filtro aplicado',
            'Lineas': int(total_lineas),
            'Total HNL': total_hnl,
        })

    filas_resumen.append({
        'Resumen': 'TOTAL FINAL',
        'Detalle': 'Lineas visibles en la tabla',
        'Lineas': int(total_lineas),
        'Total HNL': total_hnl,
    })

    return pd.DataFrame(filas_resumen)


def preparar_tabla_exportacion(df_tabla, incluir_total=True):
    """Prepara la tabla exportable con item correlativo y una fila final de total."""
    df_export = construir_tabla_planes_profesional(df_tabla).copy()

    if incluir_total:
        fila_total = {col: '' for col in df_export.columns}
        fila_total['numero'] = 'TOTAL FILTRADO'
        if 'nombre_personal' in fila_total:
            fila_total['nombre_personal'] = f"{len(df_export)} lineas"
        if 'valor_usd' in df_export.columns:
            total_usd = pd.to_numeric(df_export['valor_usd'], errors='coerce').fillna(0).sum()
            fila_total['valor_usd'] = float(total_usd)
        if 'valor_hnl' in df_export.columns:
            total_hnl = pd.to_numeric(df_export['valor_hnl'], errors='coerce').fillna(0).sum()
            fila_total['valor_hnl'] = float(total_hnl)
        if 'item' in fila_total:
            fila_total['item'] = ''
        df_export = pd.concat([df_export, pd.DataFrame([fila_total])], ignore_index=True)

    return df_export

if 'preferencias' not in st.session_state:
    st.session_state.preferencias = _preferencias_usuario_default().copy()
else:
    st.session_state.preferencias = _normalizar_preferencias_usuario(st.session_state.preferencias)

if '_preferencias_usuario_sync' not in st.session_state:
    st.session_state._preferencias_usuario_sync = ""

if '_clave_dispositivo' not in st.session_state:
    st.session_state._clave_dispositivo = _obtener_clave_dispositivo()

if 'cred_guardada_usuario' not in st.session_state:
    st.session_state.cred_guardada_usuario = ""
if 'cred_guardada_contrasena' not in st.session_state:
    st.session_state.cred_guardada_contrasena = ""
if 'login_usuario_input' not in st.session_state:
    st.session_state.login_usuario_input = ""
if 'login_contrasena_input' not in st.session_state:
    st.session_state.login_contrasena_input = ""
if 'recordarme_prefill' not in st.session_state:
    st.session_state.recordarme_prefill = None
if 'recordarme_feedback' not in st.session_state:
    st.session_state.recordarme_feedback = ""
if 'login_alerta' not in st.session_state:
    st.session_state.login_alerta = ""
if 'recordarme_este_equipo' not in st.session_state:
    st.session_state.recordarme_este_equipo = any(
        isinstance(_cred, dict) and str(_cred.get("contrasena", "")).strip()
        for _cred in st.session_state.credenciales_recordadas.values()
    )
if 'recordarme_autocarga_aplicada' not in st.session_state:
    st.session_state.recordarme_autocarga_aplicada = False
if 'ultima_actividad_sesion' not in st.session_state:
    st.session_state.ultima_actividad_sesion = ""


SESSION_TIMEOUT_MINUTOS = 20


def _marcar_actividad_sesion():
    st.session_state.ultima_actividad_sesion = datetime.now().isoformat()


def _sesion_expirada_por_inactividad():
    """Valida si la sesion actual supera el tiempo maximo de inactividad."""
    if not st.session_state.get("usuario_actual"):
        return False

    ultima_actividad_raw = str(st.session_state.get("ultima_actividad_sesion", "")).strip()
    if not ultima_actividad_raw:
        _marcar_actividad_sesion()
        return False

    try:
        ultima_actividad = datetime.fromisoformat(ultima_actividad_raw)
    except ValueError:
        _marcar_actividad_sesion()
        return False

    return (datetime.now() - ultima_actividad) > timedelta(minutes=SESSION_TIMEOUT_MINUTOS)


def _cerrar_sesion_local(mensaje=""):
    st.session_state.usuario_actual = None
    st.session_state.rol = None
    st.session_state.puede_editar = False
    st.session_state.recordarme_autocarga_aplicada = False
    st.session_state.ultima_actividad_sesion = ""
    if mensaje:
        st.session_state.login_alerta = mensaje


def _obtener_credencial_recordada_preferida(usuario_sugerido=""):
    """Devuelve (usuario, credencial) priorizando coincidencia exacta o la ultima guardada."""
    credenciales_validas = {
        str(_usr): _cred
        for _usr, _cred in st.session_state.credenciales_recordadas.items()
        if isinstance(_cred, dict) and str(_cred.get("contrasena", "")).strip()
    }
    if not credenciales_validas:
        return None, None

    usuario_buscado = (usuario_sugerido or "").strip()
    if usuario_buscado:
        cred = credenciales_validas.get(usuario_buscado)
        if not cred:
            for _usr, _cred in credenciales_validas.items():
                if str(_usr).strip().lower() == usuario_buscado.lower():
                    return _usr, _cred
        if cred:
            return usuario_buscado, cred

    return max(
        credenciales_validas.items(),
        key=lambda item: str(item[1].get("guardado_en", "")),
    )

# ============ SISTEMA DE AUTENTICACIÓN ============
def pantalla_login():
    """Pantalla de login"""
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        # Al abrir login, intenta cargar credenciales automaticamente una sola vez.
        if not st.session_state.get("recordarme_autocarga_aplicada", False):
            if (
                st.session_state.get("recordarme_este_equipo", True)
                and not st.session_state.get("login_usuario_input", "").strip()
                and not st.session_state.get("login_contrasena_input", "").strip()
            ):
                _usr_auto, _cred_auto = _obtener_credencial_recordada_preferida()
                if _usr_auto and _cred_auto:
                    st.session_state.recordarme_prefill = {
                        "usuario": _usr_auto,
                        "contrasena": _cred_auto.get("contrasena", ""),
                    }
                    st.session_state.recordarme_feedback = f"Credenciales cargadas para {_usr_auto}"
            st.session_state.recordarme_autocarga_aplicada = True

        # Prefill debe ejecutarse ANTES de crear los widgets de login.
        _prefill = st.session_state.get("recordarme_prefill")
        if isinstance(_prefill, dict):
            st.session_state.login_usuario_input = _prefill.get("usuario", "")
            st.session_state.login_contrasena_input = _prefill.get("contrasena", "")
            st.session_state.recordarme_prefill = None

        _feedback = st.session_state.get("recordarme_feedback", "")
        if _feedback:
            st.success(_feedback)
            st.session_state.recordarme_feedback = ""

        _alerta_login = st.session_state.get("login_alerta", "")
        if _alerta_login:
            st.warning(_alerta_login)
            st.session_state.login_alerta = ""

        usuario = st.text_input("👤 Usuario:", key="login_usuario_input")
        contrasena = st.text_input("🔑 Contraseña:", type="password", key="login_contrasena_input")
        recordar_equipo = st.toggle(
            "Recordarme este equipo",
            key="recordarme_este_equipo",
            help="Si está activo, al iniciar sesión se guardarán tus credenciales para cargarlas con el botón Recordarme.",
        )

        if st.button("🚀 Iniciar Sesión", width="stretch"):
            valido, info = gestor_usuarios.validar_usuario(usuario, contrasena)
            if valido:
                st.session_state.usuario_actual = usuario
                st.session_state.rol = info.get("rol", "usuario")
                st.session_state.puede_editar = bool(info.get("puede_editar", False) or st.session_state.rol in ["administrador", "superadministrador"])
                clave_dispositivo = st.session_state.get("_clave_dispositivo", "dispositivo_generico")
                preferencias_por_dispositivo = info.get("preferencias_por_dispositivo", {})
                preferencias_usuario = None
                if isinstance(preferencias_por_dispositivo, dict):
                    preferencias_usuario = preferencias_por_dispositivo.get(clave_dispositivo)
                if not isinstance(preferencias_usuario, dict):
                    preferencias_usuario = info.get("preferencias", {})
                preferencias_usuario = _normalizar_preferencias_usuario(preferencias_usuario)
                st.session_state.preferencias = preferencias_usuario.copy()
                st.session_state.tema = preferencias_usuario.get("tema", "Claro")
                st.session_state._tema_aplicado = None
                st.session_state._fondo_aplicado = None
                st.session_state._preferencias_usuario_sync = json.dumps(
                    preferencias_usuario, ensure_ascii=False, sort_keys=True
                )
                st.session_state.cred_guardada_usuario = usuario
                st.session_state.cred_guardada_contrasena = contrasena
                if recordar_equipo:
                    st.session_state.credenciales_recordadas[usuario] = {
                        "contrasena": contrasena,
                        "guardado_en": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    }
                elif usuario in st.session_state.credenciales_recordadas:
                    del st.session_state.credenciales_recordadas[usuario]
                guardar_recordarme()
                _marcar_actividad_sesion()
                st.success("✅ Sesión iniciada correctamente")
                st.rerun()
            else:
                st.error(f"❌ {info}")

        st.caption("Las cuentas de acceso se crean y administran solo desde Configuración por un administrador.")

if st.session_state.get("usuario_actual") and _sesion_expirada_por_inactividad():
    _cerrar_sesion_local("⚠️ Sesion cerrada por inactividad (20 minutos).")
    st.rerun()

if not st.session_state.usuario_actual:
    pantalla_login()
    st.stop()
else:
    _marcar_actividad_sesion()
    # ============ APLICACIÓN PRINCIPAL ============
    total_lineas_sidebar, lineas_asignadas_sidebar, lineas_disponibles_sidebar = resumen_numeros_corporativos(st.session_state.planes)
    
    # Sidebar con info de usuario
    with st.sidebar:
        st.markdown(f"### 👤 {st.session_state.usuario_actual}")
        st.markdown(f"**Rol:** {st.session_state.rol}")
        st.markdown(f"**Puede editar:** {'Sí' if st.session_state.puede_editar else 'No'}")
        st.markdown("---")

        st.markdown(
            """
            <style>
                /* Skin tematico tipo smartphone para el primer expander del sidebar */
                [data-testid="stSidebar"] details[data-testid="stExpander"]:first-of-type {
                    --skin-accent: rgba(86, 207, 255, 0.88);
                    --skin-border: rgba(173, 229, 255, 0.34);
                    border-radius: 36px !important;
                    border: 2px solid var(--skin-border) !important;
                    background:
                        linear-gradient(180deg, rgba(198, 222, 250, 0.72), rgba(152, 185, 222, 0.75)) -3px 116px / 4px 44px no-repeat,
                        linear-gradient(180deg, rgba(198, 222, 250, 0.70), rgba(152, 185, 222, 0.72)) -3px 170px / 4px 28px no-repeat,
                        linear-gradient(180deg, rgba(178, 207, 242, 0.70), rgba(132, 168, 212, 0.72)) calc(100% + 1px) 138px / 4px 56px no-repeat,
                        radial-gradient(circle at 12% 8%, rgba(160, 232, 255, 0.24), rgba(160, 232, 255, 0.0) 42%),
                        radial-gradient(circle at 88% 14%, rgba(126, 204, 255, 0.14), rgba(126, 204, 255, 0.0) 38%),
                        linear-gradient(180deg, rgba(7, 16, 30, 0.97), rgba(5, 10, 20, 0.99)) !important;
                    box-shadow:
                        0 20px 36px rgba(0, 0, 0, 0.45),
                        0 0 0 1px rgba(168, 226, 255, 0.16),
                        inset 0 0 0 1px rgba(255, 255, 255, 0.06),
                        inset 0 -34px 42px rgba(0, 0, 0, 0.24);
                    overflow: visible !important;
                    position: relative;
                    padding: 14px 12px 24px 12px !important;
                    margin-top: 8px !important;
                }

                [data-testid="stSidebar"] details[data-testid="stExpander"]:first-of-type::before {
                    content: "";
                    position: absolute;
                    top: 7px;
                    left: 50%;
                    transform: translateX(-50%);
                    width: 96px;
                    height: 12px;
                    border-radius: 0 0 14px 14px;
                    background:
                        radial-gradient(circle at 79% 50%, rgba(138, 200, 255, 0.92) 0 1.2px, rgba(20, 34, 50, 0.95) 1.3px 3.2px, transparent 3.3px),
                        radial-gradient(circle at 53% 52%, rgba(255, 255, 255, 0.32) 0 0.7px, transparent 0.8px),
                        rgba(3, 7, 14, 0.97);
                    border: 1px solid rgba(255, 255, 255, 0.12);
                    z-index: 5;
                }

                [data-testid="stSidebar"] details[data-testid="stExpander"]:first-of-type::after {
                    content: "";
                    position: absolute;
                    bottom: 8px;
                    left: 50%;
                    transform: translateX(-50%);
                    width: 88px;
                    height: 6px;
                    border-radius: 99px;
                    background: rgba(228, 241, 255, 0.68);
                    box-shadow: 0 0 10px rgba(147, 220, 255, 0.36);
                }

                [data-testid="stSidebar"] details[data-testid="stExpander"]:first-of-type summary {
                    border-radius: 20px !important;
                    background: linear-gradient(180deg, rgba(20, 48, 78, 0.66), rgba(14, 35, 58, 0.72)) !important;
                    border: 1px solid rgba(141, 214, 255, 0.22) !important;
                    margin: 10px 4px 8px 4px !important;
                    box-shadow: inset 0 0 0 1px rgba(214, 241, 255, 0.05);
                    position: relative;
                }

                [data-testid="stSidebar"] details[data-testid="stExpander"]:first-of-type [data-testid="stExpanderDetails"] {
                    border-top: 1px solid rgba(164, 225, 255, 0.18);
                    padding: 12px 4px 0 4px !important;
                    margin-top: 8px;
                }

                [data-testid="stSidebar"] details[data-testid="stExpander"]:first-of-type [data-testid="stMetric"] {
                    border-radius: 14px;
                    padding: 8px 10px;
                    background: linear-gradient(180deg, rgba(15, 30, 49, 0.56), rgba(12, 24, 39, 0.70));
                    border: 1px solid rgba(144, 215, 255, 0.12);
                    margin-bottom: 6px;
                }

                [data-testid="stSidebar"] details[data-testid="stExpander"]:first-of-type button {
                    border-radius: 12px !important;
                    border: 1px solid rgba(149, 219, 255, 0.24) !important;
                    background: linear-gradient(180deg, rgba(28, 58, 91, 0.58), rgba(18, 38, 61, 0.72)) !important;
                }

                [data-testid="stSidebar"] details[data-testid="stExpander"]:first-of-type button:hover {
                    border-color: rgba(166, 231, 255, 0.42) !important;
                    box-shadow: 0 0 0 1px rgba(118, 211, 255, 0.22), 0 8px 16px rgba(0, 0, 0, 0.28);
                }
            </style>
            """,
            unsafe_allow_html=True,
        )

        # Resumen rapido de lineas corporativas en menu izquierdo.
        with st.expander("📞 Numeros Corporativos", expanded=True):
            st.metric("Lineas activas", total_lineas_sidebar)
            st.metric("Asignadas", lineas_asignadas_sidebar)
            st.metric("Libres/Disponibles", lineas_disponibles_sidebar)

            if st.button("📋 Ver/Ocultar lista completa", key="btn_toggle_numeros"):
                st.session_state.mostrar_numeros_corporativos = not st.session_state.mostrar_numeros_corporativos

            st.markdown("#### ⬆️ Importar lineas nuevas (sumar)")
            with st.form("form_importar_nuevas_lineas", clear_on_submit=False):
                archivo_nuevas_lineas = st.file_uploader(
                    "CSV o Excel con lineas nuevas",
                    type=["csv", "xlsx", "xls"],
                    key="uploader_nuevas_lineas",
                    help="Columnas sugeridas: numero, nombre_personal, area, departamento, valor_usd",
                )
                st.caption("Se permiten nombres repetidos. La validación de duplicados se realiza por número telefónico.")

                hoja_nuevas_lineas = None
                if archivo_nuevas_lineas is not None and archivo_nuevas_lineas.name.lower().endswith((".xlsx", ".xls")):
                    try:
                        archivo_nuevas_lineas.seek(0)
                        excel_tmp = pd.ExcelFile(archivo_nuevas_lineas)
                        hoja_nuevas_lineas = st.selectbox(
                            "Hoja de Excel",
                            options=excel_tmp.sheet_names,
                            key="hoja_nuevas_lineas",
                        )
                    except Exception as e:
                        st.error(f"No se pudo leer el Excel: {e}")

                if archivo_nuevas_lineas is not None:
                    try:
                        nombre_archivo_preview = archivo_nuevas_lineas.name.lower()
                        archivo_nuevas_lineas.seek(0)
                        if nombre_archivo_preview.endswith(".csv"):
                            df_preview_import = pd.read_csv(archivo_nuevas_lineas)
                        elif nombre_archivo_preview.endswith((".xlsx", ".xls")) and hoja_nuevas_lineas:
                            archivo_nuevas_lineas.seek(0)
                            df_preview_import = pd.read_excel(archivo_nuevas_lineas, sheet_name=hoja_nuevas_lineas)
                        else:
                            df_preview_import = None

                        if df_preview_import is not None:
                            df_preview_import = preparar_dataframe_importacion(df_preview_import)
                            columnas_preview = ", ".join([str(c) for c in df_preview_import.columns[:8]])
                            requeridas_preview = ["numero", "nombre_personal", "area", "departamento", "valor_usd"]
                            requeridas_ok = all(col in df_preview_import.columns for col in requeridas_preview)
                            filas_preview = len(df_preview_import.index)
                            st.caption(f"Archivo listo: {filas_preview} filas detectadas. Columnas: {columnas_preview}")
                            st.caption(f"Columnas requeridas completas: {'Si' if requeridas_ok else 'No'}")
                            if requeridas_ok:
                                resumen_preview = analizar_importacion_lineas(
                                    df_preview_import,
                                    st.session_state.planes,
                                    float(st.session_state.get("tasa_usd_hnl", 24.0)),
                                )
                                st.caption(
                                    f"Prevalidación: nuevas={resumen_preview['nuevas']} | duplicadas={resumen_preview['duplicadas']} | inválidas={resumen_preview['invalidas']}"
                                )
                    except Exception as e:
                        st.caption(f"No se pudo generar resumen del archivo: {e}")

                submit_importar_lineas = st.form_submit_button(
                    "➕ Importar lineas nuevas",
                    disabled=not st.session_state.puede_editar,
                    use_container_width=True,
                )

            if submit_importar_lineas:
                procesar_importacion_lineas_nuevas(archivo_nuevas_lineas, hoja_nuevas_lineas)
        
        # Preferencias
        st.markdown("### ⚙️ Preferencias")
        tema = st.selectbox("🎨 Tema:", ["Oscuro", "Azul", "Verde", "Rojo", "Claro", "Automático"])
        if st.session_state.get("_tema_aplicado") != tema:
            apply_theme(tema)
            st.session_state._tema_aplicado = tema
        st.session_state.tema = tema

        fondo_visual = st.selectbox(
            "🖼️ Fondo principal:",
            ["Corporativo", "Sutil", "Impacto"],
            index=["Corporativo", "Sutil", "Impacto"].index(st.session_state.preferencias.get("fondo_visual", "Impacto"))
            if st.session_state.preferencias.get("fondo_visual", "Impacto") in ["Corporativo", "Sutil", "Impacto"]
            else 2,
        )
        st.session_state.preferencias["fondo_visual"] = fondo_visual
        if st.session_state.get("_fondo_aplicado") != fondo_visual:
            apply_main_background(fondo_visual)
            st.session_state._fondo_aplicado = fondo_visual

        estilo_hero = st.selectbox(
            "📱 Estilo del smartphone:",
            ["iPhone", "Android"],
            index=["iPhone", "Android"].index(st.session_state.preferencias.get("estilo_hero", "iPhone"))
            if st.session_state.preferencias.get("estilo_hero", "iPhone") in ["iPhone", "Android"]
            else 0,
        )
        st.session_state.preferencias["estilo_hero"] = estilo_hero
        
        idioma = st.selectbox("🌐 Idioma:", ["Español", "Inglés", "Portugués"])
        st.session_state.preferencias["idioma"] = idioma

        st.session_state.dashboard_modo_rendimiento = st.toggle(
            "⚡ Modo rendimiento del dashboard",
            value=bool(st.session_state.get("dashboard_modo_rendimiento", True)),
            help="Activado: graficos en modo estatico para mayor fluidez. Desactivado: interaccion completa.",
        )

        if st.session_state.get("usuario_actual"):
            prefs_norm = _normalizar_preferencias_usuario(st.session_state.preferencias)
            st.session_state.preferencias = prefs_norm
            firma_prefs = json.dumps(prefs_norm, ensure_ascii=False, sort_keys=True)
            if st.session_state.get("_preferencias_usuario_sync") != firma_prefs:
                if hasattr(gestor_usuarios, "actualizar_preferencias_usuario"):
                    ok_prefs, _ = gestor_usuarios.actualizar_preferencias_usuario(
                        st.session_state.usuario_actual,
                        prefs_norm,
                        st.session_state.get("_clave_dispositivo", "dispositivo_generico"),
                    )
                else:
                    ok_prefs, _ = gestor_usuarios.actualizar_usuario(
                        st.session_state.usuario_actual,
                        preferencias=prefs_norm,
                    )
                if ok_prefs:
                    st.session_state._preferencias_usuario_sync = firma_prefs

        # Tasa global de conversión USD -> HNL para toda la app
        st.markdown("### 💱 Tasa USD/HNL")
        if 'tasa_usd_hnl' not in st.session_state:
            cache_tasa = _leer_tasa_cache()
            if cache_tasa:
                st.session_state.tasa_usd_hnl = cache_tasa.get("tasa", 24.0)
                st.session_state.tasa_fuente = cache_tasa.get("fuente", "cache local")
                st.session_state.tasa_actualizada_en = cache_tasa.get(
                    "actualizado_en", datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                )
            else:
                st.session_state.tasa_usd_hnl = 24.0
                st.session_state.tasa_fuente = "Manual por defecto"
                st.session_state.tasa_actualizada_en = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        st.caption("Modo rapido activo: la tasa se actualiza solo cuando presionas el boton de actualizar.")

        if st.button("🔄 Actualizar tasa desde Internet"):
            tasa_actualizada, fuente_actualizada, fecha_actualizada = obtener_tasa_usd_hnl(force_refresh=True)
            if tasa_actualizada:
                # alerta_variacion = _verificar_alerta_variacion(tasa_actualizada)  # Comentado: función no definida
                # _registrar_tasa_historial(tasa_actualizada, fuente_actualizada or "Fuente externa")
                st.session_state.tasa_usd_hnl = tasa_actualizada
                st.session_state.tasa_fuente = fuente_actualizada or "Fuente externa"
                st.session_state.tasa_actualizada_en = fecha_actualizada or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                st.success(
                    f"Tasa actualizada: {tasa_actualizada:.4f} HNL por USD | Fuente: {st.session_state.tasa_fuente}"
                )
                # if alerta_variacion:
                #     st.warning(f"⚠️ {alerta_variacion}")
            else:
                st.error("No se pudo actualizar desde Internet. Se mantiene la ultima tasa guardada.")

        st.caption(
            f"Fuente: {st.session_state.get('tasa_fuente', 'N/D')} | "
            f"Actualizada: {st.session_state.get('tasa_actualizada_en', 'N/D')}"
        )

        tasa_antes_manual = float(st.session_state.tasa_usd_hnl)

        st.number_input(
            "Tasa manual",
            min_value=0.0,
            value=float(st.session_state.tasa_usd_hnl),
            step=0.01,
            key="sidebar_tasa_usd_hnl"
        )
        st.session_state.tasa_usd_hnl = st.session_state.sidebar_tasa_usd_hnl
        if abs(float(st.session_state.tasa_usd_hnl) - tasa_antes_manual) > 1e-9:
            st.session_state.tasa_fuente = "Manual"
            st.session_state.tasa_actualizada_en = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            _guardar_tasa_cache(st.session_state.tasa_usd_hnl, "Manual")
            # _registrar_tasa_historial(st.session_state.tasa_usd_hnl, "Manual")  # Comentado: función no definida
        
        st.markdown("---")
        
        if st.button("🚪 Cerrar Sesión"):
            _cerrar_sesion_local()
            st.rerun()
    
    st.markdown(
        """
        <style>
            /* Navegacion superior: botones profesionales y livianos para buen rendimiento. */
            .nav-principal-marker + div[data-testid="stRadio"] > div {
                background: linear-gradient(180deg, rgba(7, 18, 32, 0.76), rgba(8, 22, 39, 0.68));
                border: 1px solid rgba(149, 201, 238, 0.20);
                border-radius: 16px;
                padding: 10px;
                box-shadow: 0 6px 16px rgba(0, 0, 0, 0.18);
            }

            /* Mantiene el menu principal en ingles y evita traduccion automatica del navegador. */
            .nav-principal-marker + div[data-testid="stRadio"],
            .nav-principal-marker + div[data-testid="stRadio"] * {
                translate: no;
            }

            .nav-principal-marker + div[data-testid="stRadio"] div[role="radiogroup"] {
                display: flex;
                flex-wrap: wrap;
                gap: 8px;
            }

            .nav-principal-marker + div[data-testid="stRadio"] label[data-baseweb="radio"] {
                margin: 0;
                min-height: 46px;
                padding: 0;
                border-radius: 12px;
                border: 1px solid rgba(168, 220, 255, 0.32);
                background: linear-gradient(180deg, rgba(24, 68, 111, 0.60), rgba(14, 43, 72, 0.55));
                box-shadow: 0 4px 10px rgba(7, 28, 49, 0.20);
                transition: border-color 0.14s ease, background 0.14s ease, box-shadow 0.14s ease;
                overflow: hidden;
            }

            .nav-principal-marker + div[data-testid="stRadio"] label[data-baseweb="radio"]:hover {
                border-color: rgba(205, 237, 255, 0.62);
                box-shadow: 0 6px 14px rgba(8, 31, 56, 0.24);
                background: linear-gradient(180deg, rgba(37, 90, 142, 0.68), rgba(21, 57, 94, 0.60));
            }

            .nav-principal-marker + div[data-testid="stRadio"] label[data-baseweb="radio"] > div:first-child {
                display: none;
            }

            .nav-principal-marker + div[data-testid="stRadio"] label[data-baseweb="radio"] > div:last-child {
                padding: 11px 16px;
            }

            .nav-principal-marker + div[data-testid="stRadio"] label[data-baseweb="radio"] p {
                margin: 0;
                color: transparent !important;
                font-size: 17px;
                font-weight: 700;
                letter-spacing: 0.15px;
                text-shadow: none;
                position: relative;
            }

            /* Etiquetas visibles del menu: se renderizan por CSS para evitar traduccion del navegador. */
            .nav-principal-marker + div[data-testid="stRadio"] label[data-baseweb="radio"] p::after {
                color: rgba(236, 246, 255, 0.95);
                content: "";
            }

            .nav-principal-marker + div[data-testid="stRadio"] label[data-baseweb="radio"]:nth-of-type(1) p::after { content: "📋 Gestionar planes"; }
            .nav-principal-marker + div[data-testid="stRadio"] label[data-baseweb="radio"]:nth-of-type(2) p::after { content: "➕ Agregar plan"; }
            .nav-principal-marker + div[data-testid="stRadio"] label[data-baseweb="radio"]:nth-of-type(3) p::after { content: "👥 Empleados"; }
            .nav-principal-marker + div[data-testid="stRadio"] label[data-baseweb="radio"]:nth-of-type(4) p::after { content: "⚙️ Configuracion"; }
            .nav-principal-marker + div[data-testid="stRadio"] label[data-baseweb="radio"]:nth-of-type(5) p::after { content: "📊 Panel de control"; }

            .nav-principal-marker + div[data-testid="stRadio"] label[data-baseweb="radio"]:has(input:checked) {
                border-color: rgba(232, 245, 255, 0.95);
                background: linear-gradient(180deg, rgba(105, 182, 242, 0.90), rgba(58, 133, 203, 0.80));
                box-shadow: 0 0 0 1px rgba(224, 245, 255, 0.45), 0 7px 16px rgba(7, 30, 52, 0.28);
            }

            .nav-principal-marker + div[data-testid="stRadio"] label[data-baseweb="radio"]:has(input:checked) p {
                color: transparent !important;
                text-shadow: none;
            }

            .nav-principal-marker + div[data-testid="stRadio"] label[data-baseweb="radio"]:has(input:checked) p::after {
                color: #ffffff;
                text-shadow: 0 1px 0 rgba(8, 24, 40, 0.18);
            }

            .nav-principal-marker + div[data-testid="stRadio"] label[data-baseweb="radio"]:has(input:focus-visible) {
                outline: 2px solid rgba(189, 231, 255, 0.82);
                outline-offset: 1px;
            }

            @media (max-width: 900px) {
                .nav-principal-marker + div[data-testid="stRadio"] label[data-baseweb="radio"] {
                    width: calc(50% - 5px);
                    min-height: 42px;
                }

                .nav-principal-marker + div[data-testid="stRadio"] label[data-baseweb="radio"] > div:last-child {
                    padding: 10px 13px;
                }

                .nav-principal-marker + div[data-testid="stRadio"] label[data-baseweb="radio"] p {
                    font-size: 15px;
                }
            }

            @media (max-width: 560px) {
                .nav-principal-marker + div[data-testid="stRadio"] > div {
                    border-radius: 14px;
                    padding: 8px;
                }

                .nav-principal-marker + div[data-testid="stRadio"] label[data-baseweb="radio"] {
                    width: 100%;
                    min-height: 40px;
                    border-radius: 10px;
                }

                .nav-principal-marker + div[data-testid="stRadio"] label[data-baseweb="radio"] > div:last-child {
                    padding: 9px 12px;
                }

                .nav-principal-marker + div[data-testid="stRadio"] label[data-baseweb="radio"] p {
                    font-size: 14px;
                }

                .nav-principal-marker + div[data-testid="stRadio"] label[data-baseweb="radio"]:nth-of-type(1) p::after { content: "📋 Gestionar"; }
                .nav-principal-marker + div[data-testid="stRadio"] label[data-baseweb="radio"]:nth-of-type(2) p::after { content: "➕ Agregar"; }
                .nav-principal-marker + div[data-testid="stRadio"] label[data-baseweb="radio"]:nth-of-type(3) p::after { content: "👥 Empleados"; }
                .nav-principal-marker + div[data-testid="stRadio"] label[data-baseweb="radio"]:nth-of-type(4) p::after { content: "⚙️ Config"; }
                .nav-principal-marker + div[data-testid="stRadio"] label[data-baseweb="radio"]:nth-of-type(5) p::after { content: "📊 Panel de control"; }
            }
        </style>
        """,
        unsafe_allow_html=True,
    )

    # Navegacion principal con renderizado condicional.
    # Evita que Streamlit recalcule todas las secciones en cada rerun.
    _secciones_principales = {
        "gestionar": "📋 Gestionar planes",
        "agregar": "➕ Agregar plan",
        "empleados": "👥 Empleados",
        "configuracion": "⚙️ Configuracion",
        "dashboard": "📊 Panel de control",
    }
    if st.session_state.get("vista_principal") not in _secciones_principales:
        st.session_state.vista_principal = "gestionar"
    st.markdown('<div class="nav-principal-marker"></div>', unsafe_allow_html=True)
    st.markdown(
        """
        <script>
            (function() {
                const applyNoTranslateToMainMenu = () => {
                    const marker = window.parent.document.querySelector('.nav-principal-marker');
                    if (!marker) return;
                    const radioContainer = marker.nextElementSibling;
                    if (!radioContainer) return;

                    radioContainer.classList.add('notranslate');
                    radioContainer.setAttribute('translate', 'no');
                    radioContainer.setAttribute('lang', 'es');

                    radioContainer.querySelectorAll('*').forEach((node) => {
                        node.classList.add('notranslate');
                        node.setAttribute('translate', 'no');
                        node.setAttribute('lang', 'es');
                    });
                };

                applyNoTranslateToMainMenu();
                setTimeout(applyNoTranslateToMainMenu, 350);
                setTimeout(applyNoTranslateToMainMenu, 900);
                window.parent.addEventListener('resize', applyNoTranslateToMainMenu, { passive: true });

                const marker = window.parent.document.querySelector('.nav-principal-marker');
                const radioContainer = marker ? marker.nextElementSibling : null;
                if (radioContainer) {
                    const observer = new MutationObserver(() => applyNoTranslateToMainMenu());
                    observer.observe(radioContainer, { childList: true, subtree: true, characterData: true });
                }
            })();
        </script>
        """,
        unsafe_allow_html=True,
    )
    vista_actual = st.radio(
        "Secciones principales",
        options=list(_secciones_principales.keys()),
        format_func=lambda clave: _secciones_principales.get(clave, clave),
        horizontal=True,
        key="vista_principal",
        label_visibility="collapsed",
    )


# --- Mostrar el banner visual del celular solo al final de la app ---

# --- Banner visual del celular al final de la app ---

# --- Banner visual del celular al final de la app ---

# --- Banner visual del celular al final de la app ---

# --- Banner visual del celular al final de la app ---

# --- Banner visual del celular al final de la app ---

# --- Banner visual del celular al final de la app (última línea) ---


es_editor = bool(st.session_state.puede_editar or st.session_state.rol in ["administrador", "superadministrador"])
permiso_crear = tiene_permiso("crear")
permiso_editar = tiene_permiso("editar")
permiso_eliminar = tiene_permiso("eliminar")
permiso_importar = tiene_permiso("importar")
permiso_exportar = tiene_permiso("exportar")


def render_bloque_operador_tigo():
    """Muestra branding del operador corporativo actual (TIGO)."""
    col_logo, col_texto = st.columns([1, 7])
    with col_logo:
        st.markdown(
            """
            <div style="
                width:72px;
                height:72px;
                border-radius:18px;
                display:flex;
                align-items:center;
                justify-content:center;
                background:linear-gradient(180deg, #00b8ff, #008fd1);
                color:#ffffff;
                font-weight:800;
                font-size:22px;
                letter-spacing:1px;
                box-shadow:0 8px 18px rgba(0, 111, 168, 0.28);
            ">TIGO</div>
            """,
            unsafe_allow_html=True,
        )
    with col_texto:
        st.markdown("**Operador de servicio corporativo:** TIGO")
        st.caption("El sistema usa TIGO como operador por defecto en importaciones y nuevos registros.")

# ============ TAB 1: DASHBOARD ============
if vista_actual == "dashboard":
    st.markdown(
        """<style>
        .kpi-big { font-size: 2.8em; font-weight: 700; color: #00d9ff; text-shadow: 0 0 10px rgba(0, 217, 255, 0.4); }
        .kpi-label { font-size: 0.95em; color: rgba(255,255,255,0.76); font-weight: 500; }
        .kpi-container { background: linear-gradient(135deg, rgba(30,60,120,0.5), rgba(10,35,80,0.5)); border: 1px solid rgba(0,217,255,0.3); border-radius: 12px; padding: 20px; margin: 10px 0; }
        </style>""",
        unsafe_allow_html=True,
    )
    
    if st.session_state.planes:
        dashboard_payload = _construir_dashboard_payload(_serializar_planes_cache(st.session_state.planes))
        chart_config = {"displayModeBar": False}
        if st.session_state.get("dashboard_modo_rendimiento", True):
            chart_config["staticPlot"] = True

        total_lineas = int(dashboard_payload["total_lineas"])
        gasto_total = float(dashboard_payload["gasto_total"])
        gasto_promedio = float(dashboard_payload["gasto_promedio"])
        areas_unicas = int(dashboard_payload["areas_unicas"])
        deptos_unicos = int(dashboard_payload["deptos_unicos"])
        lineas_asignadas = int(dashboard_payload["lineas_asignadas"])
        lineas_libres = int(dashboard_payload["lineas_libres"])

        st.markdown("### 📊 Indicadores Principales")
        col_kpi1, col_kpi2, col_kpi3, col_kpi4 = st.columns(4)
        with col_kpi1:
            st.markdown(
                f'<div class="kpi-container"><div class="kpi-label">Total Líneas</div><div class="kpi-big">{total_lineas}</div></div>',
                unsafe_allow_html=True,
            )
        with col_kpi2:
            st.markdown(
                f'<div class="kpi-container"><div class="kpi-label">Gasto Total USD</div><div class="kpi-big">${gasto_total:,.0f}</div></div>',
                unsafe_allow_html=True,
            )
        with col_kpi3:
            st.markdown(
                f'<div class="kpi-container"><div class="kpi-label">Líneas Asignadas</div><div class="kpi-big">{lineas_asignadas}</div></div>',
                unsafe_allow_html=True,
            )
        with col_kpi4:
            st.markdown(
                f'<div class="kpi-container"><div class="kpi-label">Líneas Libres</div><div class="kpi-big">{lineas_libres}</div></div>',
                unsafe_allow_html=True,
            )

        st.markdown("---")

        col_chart1, col_chart2 = st.columns(2)

        with col_chart1:
            st.markdown("### 💰 Gasto por Área")
            st.plotly_chart(dashboard_payload["fig_area"], use_container_width=True, config=chart_config)

        with col_chart2:
            st.markdown("### 👥 Distribución por Departamento")
            st.plotly_chart(dashboard_payload["fig_dept"], use_container_width=True, config=chart_config)

        st.markdown("---")

        col_chart3, col_chart4 = st.columns(2)

        with col_chart3:
            st.markdown("### 📈 Tendencia Acumulada de Gasto")
            st.plotly_chart(dashboard_payload["fig_cumsum"], use_container_width=True, config=chart_config)

        with col_chart4:
            st.markdown("### 💵 Histograma de Valores USD")
            st.plotly_chart(dashboard_payload["fig_scatter"], use_container_width=True, config=chart_config)

        st.markdown("---")
        st.markdown("### 📊 Métricas Adicionales")
        metrics_col1, metrics_col2, metrics_col3, metrics_col4 = st.columns(4)
        with metrics_col1:
            st.metric(label="Promedio por Línea (USD)", value=f"${gasto_promedio:,.2f}")
        with metrics_col2:
            st.metric(label="Áreas Únicas", value=areas_unicas)
        with metrics_col3:
            st.metric(label="Departamentos", value=deptos_unicos)
        with metrics_col4:
            tasa_ocupacion = (lineas_asignadas / total_lineas * 100) if total_lineas > 0 else 0
            st.metric(label="Ocupación (%)", value=f"{tasa_ocupacion:.1f}%")

    else:
        st.info("📌 No hay planes registrados aún. ¡Comienza agregando uno!")

# ============ TAB 2: AGREGAR PLAN ============
if vista_actual == "agregar":
    st.subheader("➕ Agregar Nuevo Plan Corporativo")

    if not permiso_crear:
        st.warning("🔒 Tu usuario tiene acceso de solo lectura. No puedes crear ni editar planes.")

    col1, col2 = st.columns(2)
    
    with col1:
        numero = st.text_input("📞 Número Corporativo", placeholder="Ej: +504-2234-5678")
        st.caption("Formato sugerido: +504-2234-5678")
        st.caption("Se permiten nombres repetidos; el número corporativo debe ser único.")
        nombre_personal = st.text_input("👤 Nombre del Personal", placeholder="Ej: Juan Pérez")
        area = st.text_input("🏢 Área", placeholder="Ej: Ventas")
        operador = st.text_input("📶 Operador", value="TIGO")
        perfil_profesional = st.text_input("🎓 Perfil Profesional", placeholder="Ej: Ingeniero de Soporte")
        dispositivo_asignado = st.text_input("📱 Dispositivo Asignado", placeholder="Ej: Samsung Galaxy S23")
        marca_dispositivo = st.text_input("🏷️ Marca", placeholder="Ej: Samsung")
        modelo_dispositivo = st.text_input("🆔 Modelo", placeholder="Ej: S23")
        serie_dispositivo = st.text_input("🔢 Serie del Dispositivo", placeholder="Ej: SN-ABC123456")

    with col2:
        departamento = st.text_input("🏛️ Departamento", placeholder="Ej: Comercial")
        valor_usd = st.number_input("💵 Valor del Plan (USD)", min_value=0.0, step=0.01)
        imei1 = st.text_input("📳 IMEI 1", placeholder="Ej: 123456789012345")
        imei2 = st.text_input("📳 IMEI 2", placeholder="Ej: 543210987654321")
        observaciones = st.text_area("📝 Observaciones", height=100)
        motivo_cambio_dispositivo = st.text_input("📝 Motivo de cambio de dispositivo (si aplica)", "")

        valor_lempiras = valor_usd * st.session_state.tasa_usd_hnl
        st.info(f"Total en Lempiras: L {valor_lempiras:,.2f} (Tasa: {st.session_state.tasa_usd_hnl:,.2f})")

    # Botón guardar
    if st.button("💾 Guardar Plan", width="stretch", type="primary", disabled=not permiso_crear):
        if numero and nombre_personal and area and departamento and perfil_profesional:
            numero_limpio = normalizar_numero_telefonico(numero)
            if not re.match(r"^\+?[0-9\-\s]{7,20}$", numero_limpio):
                st.error("❌ El número corporativo tiene formato inválido.")
            existentes = {normalizar_numero_telefonico(p.get('numero', '')) for p in st.session_state.planes}
            if numero_limpio in existentes:
                st.error("❌ Este número corporativo ya existe en el sistema.")
            elif re.match(r"^\+?[0-9\-\s]{7,20}$", numero_limpio):
                dispositivo_historial = []
                if motivo_cambio_dispositivo:
                    dispositivo_historial.append({
                        'fecha': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        'usuario': st.session_state.usuario_actual,
                        'motivo': motivo_cambio_dispositivo,
                        'dispositivo_asignado': dispositivo_asignado,
                        'marca': marca_dispositivo,
                        'modelo': modelo_dispositivo,
                        'serie_dispositivo': serie_dispositivo,
                        'imei1': imei1,
                        'imei2': imei2
                    })

                nuevo_plan = {
                    'numero': numero_limpio,
                    'operador': (str(operador).strip() or 'TIGO').upper(),
                    'nombre_personal': nombre_personal,
                    'area': area,
                    'departamento': departamento,
                    'perfil_profesional': perfil_profesional,
                    'valor_usd': valor_usd,
                    'tasa_usd_hnl': st.session_state.tasa_usd_hnl,
                    'valor_hnl': valor_lempiras,
                    'observaciones': observaciones,
                    'dispositivo_asignado': dispositivo_asignado,
                    'marca': marca_dispositivo,
                    'modelo': modelo_dispositivo,
                    'serie_dispositivo': serie_dispositivo,
                    'imei1': imei1,
                    'imei2': imei2,
                    'dispositivo_historial': dispositivo_historial,
                    'fecha_creacion': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                }
                # Inicializar historial de asignaciones, incluyendo precio del celular
                asignaciones_historial = [{
                    'fecha': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    'usuario': st.session_state.usuario_actual,
                    'nombre_personal': nombre_personal
                }]
                nuevo_plan['asignaciones_historial'] = asignaciones_historial
                st.session_state.planes.append(nuevo_plan)
                guardar_planes()
                registrar_movimiento("Agregar Plan", f"{numero} - {nombre_personal} - {area} - {departamento} - Dispositivo: {dispositivo_asignado}")
                st.success("✅ Plan agregado exitosamente!")
                st.balloons()
        else:
            st.error("❌ Por favor completa los campos obligatorios: Número, Nombre, Área, Departamento y Perfil Profesional")

# ============ TAB 3: GESTIONAR PLANES ============

if vista_actual == "gestionar":
    # --- Estilo visual verde para la tabla de gestionar ---
    st.markdown(
        """
        <style>
        [data-testid="stDataFrame"] {
            background: #2ecc40;
            border: 3px solid #145c1b;
            border-radius: 18px;
            box-shadow: 0 6px 24px 0 rgba(46,204,64,0.18), 0 0 0 2px #2ecc4055;
            overflow: hidden;
        }
        [data-testid="stDataFrame"] thead tr th {
            background: #145c1b;
            color: #fff;
            font-weight: 900;
            font-size: 1.13em;
            letter-spacing: 0.7px;
            border-bottom: 3px solid #2ecc40;
            text-shadow: 0 1px 0 #fff8;
            text-align: center;
            box-shadow: 0 4px 16px 0 rgba(20,92,27,0.25), 0 1.5px 0 #2ecc40;
        }
        [data-testid="stDataFrame"] tbody tr {
            background: rgba(255,255,255,0.10);
            transition: background 0.22s, color 0.22s;
        }
        [data-testid="stDataFrame"] tbody tr:hover {
            background: #2ecc40 !important;
            color: #145c1b !important;
            font-weight: 700;
        }
        [data-testid="stDataFrame"] td {
            font-size: 1.07em;
            font-weight: 500;
            color: #fff;
            text-align: center;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )
    st.subheader("📋 Gestionar Planes - KM MOTOS")
    render_bloque_operador_tigo()

    if not (permiso_editar or permiso_eliminar):
        st.info("👀 Modo solo lectura: puedes visualizar información, pero no modificarla.")
    
    if st.session_state.planes:
        _planes_serializados_tab3 = _serializar_planes_cache(st.session_state.planes)
        df = _construir_dataframe_planes_cache(_planes_serializados_tab3)

        # Filtros
        col_filter1, col_filter2, col_filter3, col_filter4 = st.columns(4)
        
        with col_filter1:
            filtro_area = st.multiselect("Filtrar por Área:", options=df['area'].unique())
        
        with col_filter2:
            filtro_dept = st.multiselect("Filtrar por Departamento:", options=df['departamento'].unique())
        
        with col_filter3:
            valor_min = st.number_input("Valor mínimo (USD):", min_value=0.0, step=0.01)

        with col_filter4:
            filtro_busqueda = st.text_input("Buscar # o nombre:", placeholder="Ej: +504 o Juan")
        
        # Aplicar filtros
        df_filtrado = df.copy()
        
        if filtro_area:
            df_filtrado = df_filtrado[df_filtrado['area'].isin(filtro_area)]
        
        if filtro_dept:
            df_filtrado = df_filtrado[df_filtrado['departamento'].isin(filtro_dept)]
        
        if valor_min > 0:
            df_filtrado = df_filtrado[df_filtrado['valor_usd'] >= valor_min]

        if filtro_busqueda.strip():
            patron = re.escape(filtro_busqueda.strip().lower())
            # Versión normalizada: elimina separadores comunes para comparar solo dígitos/letras
            _sep = re.compile(r'[\s\-().+]')
            patron_norm = re.escape(_sep.sub('', filtro_busqueda.strip().lower()))
            num_norm = df_filtrado['numero'].astype(str).str.lower().apply(lambda x: _sep.sub('', x))
            nom_norm = df_filtrado['nombre_personal'].astype(str).str.lower().apply(lambda x: _sep.sub('', x))
            df_filtrado = df_filtrado[
                df_filtrado['numero'].astype(str).str.lower().str.contains(patron, na=False, regex=True)
                | df_filtrado['nombre_personal'].astype(str).str.lower().str.contains(patron, na=False, regex=True)
                | num_norm.str.contains(patron_norm, na=False, regex=True)
                | nom_norm.str.contains(patron_norm, na=False, regex=True)
            ]

        aplicar_estilo_tabla_profesional()
        df_vista_planes = construir_tabla_planes_profesional(df_filtrado)
        _column_config_planes = {
            "item": st.column_config.NumberColumn("Item", format="%d", width="small"),
            "numero": st.column_config.TextColumn("Número", width="medium"),
            "operador": st.column_config.TextColumn("Operador", width="small"),
            "estado_linea": st.column_config.TextColumn("Estado", width="small"),
            "nombre_personal": st.column_config.TextColumn("Asignado a", width="medium"),
            "area": st.column_config.TextColumn("Área", width="medium"),
            "departamento": st.column_config.TextColumn("Departamento", width="medium"),
            "tienda": st.column_config.TextColumn("Tienda", width="medium"),
            "zona": st.column_config.TextColumn("Zona", width="medium"),
            "empresa": st.column_config.TextColumn("Empresa", width="medium"),
            "perfil_profesional": st.column_config.TextColumn("Perfil", width="large"),
            "dispositivo_asignado": st.column_config.TextColumn("Dispositivo", width="medium"),
            "marca": st.column_config.TextColumn("Marca", width="small"),
            "modelo": st.column_config.TextColumn("Modelo", width="small"),
            "serie_dispositivo": st.column_config.TextColumn("Serie", width="medium"),
            "precio_dispositivo": st.column_config.NumberColumn("Precio del Dispositivo", format="L %.2f"),
            "valor_usd": st.column_config.NumberColumn("Valor Plan USD", format="$ %.2f"),
            "valor_hnl": st.column_config.NumberColumn("Valor Plan HNL", format="L %.2f"),
            "fecha_creacion": st.column_config.TextColumn("Fecha", width="medium"),
        }
        _selected_plan_idx = None
        _df_tabla_click = df_vista_planes.copy()
        _df_tabla_click["_plan_idx"] = df_filtrado.index.to_list()
        _indices_filtrados = [
            int(i) for i in df_filtrado.index.to_list()
            if 0 <= int(i) < len(st.session_state.planes)
        ]
        st.session_state["indices_filtrados_tab3"] = _indices_filtrados

        # Reiniciar selección visual cuando cambian filtros
        _firma_filtros = (
            tuple(sorted(str(x) for x in filtro_area)),
            tuple(sorted(str(x) for x in filtro_dept)),
            round(float(valor_min), 2),
            filtro_busqueda.strip().lower(),
        )
        _firma_prev = st.session_state.get("tabla_planes_click_firma")
        if _firma_prev != _firma_filtros:
            st.session_state["tabla_planes_click_firma"] = _firma_filtros
            st.session_state["tabla_planes_click_version"] = st.session_state.get("tabla_planes_click_version", 0) + 1
        _tabla_click_key = f"tabla_planes_click_editar_v{st.session_state.get('tabla_planes_click_version', 0)}"

        st.caption("Haz clic en una fila y luego usa el botón ✏️ para abrir la edición abajo.")
        try:
            _evt = st.dataframe(
                _df_tabla_click.drop(columns=["_plan_idx"]),
                width="stretch",
                hide_index=True,
                height=430,
                column_config=_column_config_planes,
                on_select="rerun",
                selection_mode="single-row",
                key=_tabla_click_key,
            )
            _rows_sel = getattr(getattr(_evt, "selection", None), "rows", [])
            if _rows_sel:
                _row_sel = int(_rows_sel[0])
                if 0 <= _row_sel < len(_df_tabla_click):
                    _selected_plan_idx = int(_df_tabla_click.iloc[_row_sel]["_plan_idx"])
                else:
                    # Si el filtro cambia, Streamlit puede conservar una selección vieja fuera de rango.
                    _selected_plan_idx = None
        except TypeError:
            # Compatibilidad para versiones antiguas sin selección de filas en st.dataframe
            st.dataframe(
                df_vista_planes,
                width="stretch",
                hide_index=True,
                height=430,
                column_config=_column_config_planes,
            )
        except IndexError:
            # Protección extra por si la selección cambia durante el rerun.
            _selected_plan_idx = None

        _edit_cols = st.columns([4.0, 1.2])
        with _edit_cols[0]:
            if _selected_plan_idx is not None and 0 <= _selected_plan_idx < len(st.session_state.planes):
                _plan_sel_info = st.session_state.planes[_selected_plan_idx]
                st.success(
                    f"Fila seleccionada: {_plan_sel_info.get('numero', '')} - {_plan_sel_info.get('nombre_personal', '')}"
                )
            else:
                st.info("No hay fila seleccionada. Haz clic en una fila de la tabla para editarla.")
        with _edit_cols[1]:
            _can_edit_selected = _selected_plan_idx is not None and 0 <= _selected_plan_idx < len(st.session_state.planes)
            if st.button("✏️ Editar fila", key="btn_editar_fila_tabla", use_container_width=True, disabled=not _can_edit_selected):
                st.session_state['selector_plan_editar_tab3'] = _selected_plan_idx
                st.rerun()

        df_resumen_lineas = construir_resumen_lineas(df_filtrado, filtro_area, filtro_dept)
        st.dataframe(
            df_resumen_lineas,
            width="stretch",
            hide_index=True,
            column_config={
                "Resumen": st.column_config.TextColumn("Resumen", width="medium"),
                "Detalle": st.column_config.TextColumn("Detalle", width="large"),
                "Lineas": st.column_config.NumberColumn("Cantidad de lineas", format="%d"),
                "Total HNL": st.column_config.NumberColumn("Total HNL", format="L %.2f"),
            },
        )
        
        st.markdown("---")
        
        # Opciones de gestión
        col_gestion1, col_gestion2, col_gestion3 = st.columns(3)
        
        with col_gestion1:
            st.markdown("**📥 Descargar Reporte:**")
            formato_descarga = st.selectbox(
                "Formato de descarga:",
                ["Excel", "PDF", "CSV"],
                label_visibility="collapsed",
                key="formato_descarga_gestionar"
            )
            
            n_filtrados = len(df_filtrado)
            n_total = len(df)
            etiqueta_filtro = (
                f"({n_filtrados} de {n_total} registros — filtrados)"
                if n_filtrados < n_total
                else f"({n_total} registros — todos)"
            )
            if st.button(f"📥 Descargar {etiqueta_filtro}", disabled=not permiso_exportar):
                from io import BytesIO as _BytesIO
                # Siempre exportar el conjunto filtrado visible
                _df_export = preparar_tabla_exportacion(df_filtrado, incluir_total=True)
                _df_export_detalle = construir_tabla_planes_profesional(df_filtrado)
                _sufijo = datetime.now().strftime('%Y%m%d')
                _nombre_base = f"planes_corporativos_{_sufijo}"
                _usuario_reporte = st.session_state.get("usuario_actual", "Usuario")
                _fecha_hora_reporte = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                _filtros_aplicados = []
                if filtro_area:
                    _filtros_aplicados.append(f"Áreas: {', '.join(filtro_area)}")
                if filtro_dept:
                    _filtros_aplicados.append(f"Departamentos: {', '.join(filtro_dept)}")
                _desc_filtros = " | ".join(_filtros_aplicados) if _filtros_aplicados else "Todos"
                _titulo_reporte = f"REPORTE FILTRADO - PLANES CORPORATIVOS KM MOTOS ({_sufijo})"
                _texto_filtros = f"FILTROS: {_desc_filtros.upper()}"
                _texto_registros = (
                    f"REGISTROS EXPORTADOS: {n_filtrados}"
                    + (f" DE {n_total} (RESULTADO DE FILTRO APLICADO)" if n_filtrados < n_total else f" DE {n_total}")
                )
                _texto_firma = f"REPORTE GENERADO POR: {str(_usuario_reporte).upper()}"

                _cols_export = list(_df_export.columns)
                _fila_base = {c: "" for c in _cols_export}
                _header_rows = []
                _r1 = _fila_base.copy()
                _r1["numero"] = _titulo_reporte
                _header_rows.append(_r1)
                _r2 = _fila_base.copy()
                _r2["numero"] = _texto_filtros
                _header_rows.append(_r2)
                _r3 = _fila_base.copy()
                _r3["numero"] = _texto_registros
                _header_rows.append(_r3)
                _header_rows.append(_fila_base.copy())

                _footer_rows = []
                _footer_rows.append(_fila_base.copy())
                _rf = _fila_base.copy()
                _rf["numero"] = _texto_firma
                if "fecha_creacion" in _rf:
                    _rf["fecha_creacion"] = _fecha_hora_reporte
                _footer_rows.append(_rf)

                if formato_descarga == "CSV":
                    # Encabezados personalizados como filas, tabla con headers reales
                    _csv_lines = []
                    # Título, filtros y registros como filas independientes
                    _csv_lines.append(f'"{_titulo_reporte}"')
                    _csv_lines.append(f'"{_texto_filtros}"')
                    _csv_lines.append(f'"{_texto_registros}"')
                    _csv_lines.append("")
                    # Tabla principal con encabezados reales
                    _csv_lines.append(_df_export.to_csv(index=False, lineterminator='\n').strip())
                    # Footer (sin encabezado)
                    if _footer_rows:
                        _csv_lines.append(pd.DataFrame(_footer_rows).to_csv(index=False, header=False, lineterminator='\n').strip())
                    _csv_data = ("\n".join(_csv_lines)).encode("utf-8")
                    st.download_button(
                        label="Descargar CSV",
                        data=_csv_data,
                        file_name=f"{_nombre_base}.csv",
                        mime="text/csv",
                        key="btn_csv_gestionar",
                    )

                elif formato_descarga == "Excel":
                    _excel_buffer = _BytesIO()
                    _excel_ok = False
                    for _engine in ("xlsxwriter", "openpyxl"):
                        # (Eliminado bloque try sin except/finally)
                            with pd.ExcelWriter(_excel_buffer, engine=_engine) as _writer:
                                _sheet_name = "Planes"
                                _last_col = max(len(_cols_export), 1)
                                pd.DataFrame(_header_rows).to_excel(
                                    _writer,
                                    index=False,
                                    header=False,
                                    sheet_name=_sheet_name,
                                    startrow=0,
                                )
                                _df_export.to_excel(
                                    _writer,
                                    index=False,
                                    sheet_name=_sheet_name,
                                    startrow=len(_header_rows),
                                )
                                pd.DataFrame(_footer_rows).to_excel(
                                    _writer,
                                    index=False,
                                    header=False,
                                    sheet_name=_sheet_name,
                                    startrow=len(_header_rows) + len(_df_export) + 2,
                                )

                                if _engine == "xlsxwriter":
                                    _workbook = _writer.book
                                    _worksheet = _writer.sheets[_sheet_name]
                                    _title_fmt = _workbook.add_format({
                                        "bold": True,
                                        "font_size": 14,
                                        "align": "center",
                                        "valign": "vcenter",
                                    })
                                    _meta_fmt = _workbook.add_format({
                                        "bold": True,
                                        "font_size": 10,
                                        "align": "center",
                                        "valign": "vcenter",
                                    })
                                    _footer_fmt = _workbook.add_format({
                                        "bold": True,
                                        "font_size": 10,
                                    })
                                    _total_fmt = _workbook.add_format({
                                        "bold": True,
                                        "bg_color": "#D9EAF7",
                                        "border": 1,
                                    })
                                    _worksheet.merge_range(0, 0, 0, _last_col - 1, _titulo_reporte, _title_fmt)
                                    _worksheet.merge_range(1, 0, 1, _last_col - 1, _texto_filtros, _meta_fmt)
                                    _worksheet.merge_range(2, 0, 2, _last_col - 1, _texto_registros, _meta_fmt)
                                    _total_row_idx = len(_header_rows) + len(_df_export)
                                    _total_row_values = _df_export.iloc[-1].tolist()
                                    for _col_idx, _value in enumerate(_total_row_values):
                                        _worksheet.write(_total_row_idx, _col_idx, _value, _total_fmt)
                                    _worksheet.write(len(_header_rows) + len(_df_export) + 3, 0, _texto_firma, _footer_fmt)
                                else:
                                    try:
                                        from openpyxl.styles import Alignment, Font, PatternFill
                                        from openpyxl.utils import get_column_letter
                                    except ImportError:
                                        pass

                                    _worksheet = _writer.sheets[_sheet_name]
                                    _last_col_letter = get_column_letter(_last_col)
                                    for _row_idx, _text, _size in ((1, _titulo_reporte, 14), (2, _texto_filtros, 10), (3, _texto_registros, 10)):
                                        _worksheet.merge_cells(f"A{_row_idx}:{_last_col_letter}{_row_idx}")
                                        _cell = _worksheet.cell(row=_row_idx, column=1)
                                        _cell.value = _text
                                        _cell.font = Font(bold=True, size=_size)
                                        _cell.alignment = Alignment(horizontal="center", vertical="center")
                                    _total_row_excel = len(_header_rows) + len(_df_export) + 1
                                    _total_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
                                    for _col_idx in range(1, _last_col + 1):
                                        _cell_total = _worksheet.cell(row=_total_row_excel, column=_col_idx)
                                        _cell_total.font = Font(bold=True)
                                        _cell_total.fill = _total_fill
                                    _footer_cell = _worksheet.cell(row=len(_header_rows) + len(_df_export) + 4, column=1)
                                    _footer_cell.value = _texto_firma
                                    _footer_cell.font = Font(bold=True, size=10)
                            _excel_ok = True
                            break
                        # except Exception:
                        #     _excel_buffer = _BytesIO()
                    if _excel_ok:
                        _excel_buffer.seek(0)
                        st.download_button(
                            label="Descargar Excel",
                            data=_excel_buffer,
                            file_name=f"{_nombre_base}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="btn_excel_gestionar",
                        )
                    else:
                        st.warning("⚠️ Instala xlsxwriter o openpyxl: `pip install openpyxl`")

                elif formato_descarga == "PDF":
                    if FPDF is not None:
                        _pdf = FPDF(format="letter")
                        _pdf.set_auto_page_break(auto=True, margin=15)
                        _pdf.add_page()
                        _pdf.set_font("Arial", "B", 13)
                        _pdf.cell(
                            0, 10,
                            _texto_seguro_pdf(f"Reporte de PLANES CORPORATIVOS KM MOTOS - {_sufijo}"),
                            ln=True, align="C",
                        )
                        # Agregar filtro aplicado debajo del título
                        _pdf.set_font("Arial", "B", 11)
                        _pdf.cell(
                            0, 8,
                            _texto_seguro_pdf(f"FILTRO: {_desc_filtros.upper()}"),
                            ln=True, align="C",
                        )
                        _pdf.set_font("Arial", size=10)
                        for _, _row in _df_export_detalle.iterrows():
                            _line = _texto_seguro_pdf(
                                f"{_row.get('item','')}. {_row.get('numero','')} | {_row.get('nombre_personal','')} | "
                                f"{_row.get('area','')} | {_row.get('departamento','')} | "
                                f"USD {float(_row.get('valor_usd', 0)):,.2f} | "
                                f"HNL {float(_row.get('valor_hnl', 0)):,.2f}"
                            )
                            _pdf.multi_cell(180, 7, _line)
                            _obs = _texto_seguro_pdf(_row.get("observaciones", ""))
                            if _obs:
                                _pdf.multi_cell(180, 6, f"Obs: {_obs}")
                            # Agregar historial de asignaciones
                            plan_obj = next((p for p in st.session_state.planes if p.get('numero') == _row.get('numero', '')), None)
                            if plan_obj:
                                asignaciones = plan_obj.get('asignaciones_historial', [])
                                if asignaciones:
                                    _pdf.set_font("Arial", size=8)
                                    _pdf.set_x(_pdf.l_margin)
                                    _pdf.cell(0, 5, "Historial de Asignaciones:", ln=True)
                                    for _asig in sorted(asignaciones, key=lambda x: x.get('fecha', ''), reverse=True):
                                        precio_cel = _asig.get('precio_celular', None)
                                        _hist_line = _texto_seguro_pdf(
                                            f"  - {_asig.get('fecha')} | {_asig.get('nombre_personal')} | Por: {_asig.get('usuario')}"
                                            + (f" | Dispositivo: {_asig.get('dispositivo_asignado','')}" if _asig.get('dispositivo_asignado') else "")
                                            + (f" | Marca: {_asig.get('marca','')}" if _asig.get('marca') else "")
                                            + (f" | Modelo: {_asig.get('modelo','')}" if _asig.get('modelo') else "")
                                            + (f" | Serie: {_asig.get('serie_dispositivo','')}" if _asig.get('serie_dispositivo') else "")
                                            + (f" | Precio Celular: L {precio_cel:,.2f}" if precio_cel is not None else "")
                                        )
                                        _pdf.cell(0, 4, _hist_line, ln=True)
                                    _pdf.set_font("Arial", size=9)
                            _pdf.ln(1)
                        _pdf.ln(3)
                        _pdf.set_font("Arial", "B", 10)
                        _pdf.cell(0, 8, _texto_seguro_pdf(f"TOTAL DE LINEAS: {n_filtrados}"), ln=True)
                        _total_hnl_pdf = float(pd.to_numeric(_df_export_detalle.get('valor_hnl', 0), errors='coerce').fillna(0).sum())
                        _pdf.cell(0, 8, _texto_seguro_pdf(f"TOTAL EN LEMPIRAS: HNL {_total_hnl_pdf:,.2f}"), ln=True)
                        # Agregar firma del usuario al final
                        _pdf.ln(4)
                        _pdf.set_font("Arial", size=9)
                        _ancho_util = max(10, _pdf.w - _pdf.l_margin - _pdf.r_margin)
                        _pdf.set_x(_pdf.l_margin)
                        firma_texto = f"Reporte generado por: {_usuario_reporte} | Fecha: {_fecha_hora_reporte}"
                        _pdf.multi_cell(_ancho_util, 6, _texto_seguro_pdf(firma_texto))
                        _pdf_bytes = _pdf_a_bytes(_pdf)
                        st.download_button(
                            label="Descargar PDF",
                            data=_pdf_bytes,
                            file_name=f"{_nombre_base}.pdf",
                            mime="application/pdf",
                            key="btn_pdf_gestionar",
                        )

        
        with col_gestion2:
            if st.button("🗑️ Limpiar todos los datos", disabled=not permiso_eliminar):
                st.session_state.confirmar_limpiar_datos = True

            if st.session_state.get("confirmar_limpiar_datos", False):
                st.warning("Esta acción eliminará todos los planes registrados.")
                col_conf_1, col_conf_2 = st.columns(2)
                with col_conf_1:
                    if st.button("✅ Sí, eliminar", key="confirmar_eliminar_todo"):
                        st.session_state.planes = []
                        guardar_planes()
                        st.session_state.confirmar_limpiar_datos = False
                        st.rerun()
                with col_conf_2:
                    if st.button("❌ Cancelar", key="cancelar_eliminar_todo"):
                        st.session_state.confirmar_limpiar_datos = False
                        st.rerun()
        
        with col_gestion3:
            if st.button("🔄 Actualizar vista"):
                st.rerun()

        st.markdown("---")
        st.subheader("✏️ Edición de Planes")

        _indices_nav = st.session_state.get("indices_filtrados_tab3", [])
        _indices_nav = [i for i in _indices_nav if 0 <= i < len(st.session_state.planes)]
        if not _indices_nav:
            _indices_nav = list(range(len(st.session_state.planes)))

        _idx_actual = st.session_state.get("selector_plan_editar_tab3", 0)
        if _idx_actual not in _indices_nav and _indices_nav:
            st.session_state["selector_plan_editar_tab3"] = _indices_nav[0]
            _idx_actual = _indices_nav[0]

        _pos_actual = _indices_nav.index(_idx_actual) if _idx_actual in _indices_nav else None
        _nav_cols = st.columns([1.2, 1.2, 2.6])
        with _nav_cols[0]:
            if st.button("⬅️ Anterior", key="btn_nav_anterior_tab3", use_container_width=True, disabled=_pos_actual in [None, 0]):
                st.session_state["selector_plan_editar_tab3"] = _indices_nav[_pos_actual - 1]
                st.rerun()
        with _nav_cols[1]:
            if st.button("Siguiente ➡️", key="btn_nav_siguiente_tab3", use_container_width=True, disabled=_pos_actual is None or _pos_actual >= len(_indices_nav) - 1):
                st.session_state["selector_plan_editar_tab3"] = _indices_nav[_pos_actual + 1]
                st.rerun()
        with _nav_cols[2]:
            if _pos_actual is not None:
                st.caption(f"Navegando filtrados: {_pos_actual + 1} de {len(_indices_nav)}")
            else:
                st.caption("No hay selección activa en la lista filtrada.")

        # Edición rápida de un plan existente (se muestra debajo de la tabla)
        indice_seleccionado = st.selectbox(
            "Selecciona un plan para editar:",
            options=_indices_nav,
            format_func=lambda i: f"{st.session_state.planes[i]['numero']} - {st.session_state.planes[i]['nombre_personal']} ({st.session_state.planes[i]['area']})",
            key="selector_plan_editar_tab3",
        )

        plan_sel = st.session_state.planes[indice_seleccionado]

        # ---- Selector de empleado desde catálogo (fuera del form para permitir rerun) ----
        st.markdown("#### 🔍 Asignar empleado desde catálogo")
        _sel_emp_raw = 0
        _patron_emp = ""
        # Inicializar valores por defecto con los datos actuales del plan
        _nombre_default = plan_sel.get("nombre_personal", "")
        _area_default = plan_sel.get("area", "")
        _dept_default = plan_sel.get("departamento", "")
        _perfil_default = plan_sel.get("perfil_profesional", "")
        if st.session_state.empleados:
            _buscar_emp = st.text_input(
                "Buscar empleado (nombre, área, perfil o departamento):",
                placeholder="Escribe para filtrar...",
                key=f"buscar_emp_edit_{indice_seleccionado}"
            )
            _patron_emp = _buscar_emp.strip().lower()
            _emps_filtrados = [
                e for e in st.session_state.empleados
                if not _patron_emp
                or _patron_emp in str(e.get("nombre", "")).lower()
                or _patron_emp in str(e.get("area", "")).lower()
                or _patron_emp in str(e.get("perfil_profesional", "")).lower()
                or _patron_emp in str(e.get("departamento", "")).lower()
            ]
            _opciones_emp_labels = ["— Sin selección (mantener datos actuales) —"] + [
                f"{e.get('nombre', '')}  |  {e.get('perfil_profesional', '')}  |  {e.get('area', '')}  |  {e.get('departamento', '')}"
                for e in _emps_filtrados
            ]
            _sel_emp_raw = st.selectbox(
                "Empleado:",
                options=range(len(_opciones_emp_labels)),
                format_func=lambda i: _opciones_emp_labels[i],
                key=f"sel_emp_edit_{indice_seleccionado}_{_patron_emp}",
                label_visibility="collapsed",
            )
            if _sel_emp_raw > 0:
                _emp_activo = _emps_filtrados[_sel_emp_raw - 1]
                _nombre_default = _emp_activo.get("nombre", plan_sel.get("nombre_personal", ""))
                _area_default = _emp_activo.get("area", plan_sel.get("area", ""))
                _dept_default = _emp_activo.get("departamento", plan_sel.get("departamento", ""))
                _perfil_default = _emp_activo.get("perfil_profesional", plan_sel.get("perfil_profesional", ""))
                st.success(
                    f"✅ **{_nombre_default}** — {_perfil_default} | {_area_default} | {_dept_default}"
                )
            else:
                _nombre_default = plan_sel.get("nombre_personal", "")
                _area_default = plan_sel.get("area", "")
                _dept_default = plan_sel.get("departamento", "")
                _perfil_default = plan_sel.get("perfil_profesional", "")

        # La clave del form cambia con el empleado seleccionado para refrescar los campos pre-rellenos
        _form_key_edit = f"form_editar_plan_{indice_seleccionado}_{_sel_emp_raw}_{_patron_emp if st.session_state.empleados else 0}"

        # Inicializar historiales para evitar errores de variable no definida
        dispositivo_historial_actual = list(plan_sel.get('dispositivo_historial', []))
        asignaciones_historial_actual = list(plan_sel.get('asignaciones_historial', []))

        with st.form(key=_form_key_edit):
            numero_edit = st.text_input("📞 Número Corporativo", value=plan_sel.get('numero', ''))
            operador_edit = st.text_input("📶 Operador", value=plan_sel.get('operador', 'TIGO'))

            # Filtro de empleados para reasignar el número
            empleados = st.session_state.get('empleados', [])
            empleados_nombres = [e.get('nombre', '') for e in empleados if e.get('nombre', '').strip()]
            empleados_nombres = sorted(set(empleados_nombres))
            empleado_seleccionado = st.selectbox(
                "Seleccionar empleado para reasignar:",
                options=["— Sin cambio —"] + empleados_nombres,
                index=0,
                help="Selecciona un empleado para reasignar este número."
            )
            # CSS para poner el selectbox en verde
            st.markdown(
                """
                <style>
                div[data-baseweb="select"] > div {
                    background-color: #1b5e20 !important;
                    color: #fff !important;
                }
                div[data-baseweb="select"] input {
                    color: #fff !important;
                }
                div[data-baseweb="select"] svg {
                    color: #fff !important;
                }
                </style>
                """,
                unsafe_allow_html=True
            )
            if empleado_seleccionado != "— Sin cambio —":
                nombre_edit = empleado_seleccionado
            else:
                nombre_edit = st.text_input("👤 Nombre del Personal", value=_nombre_default)
            area_edit = st.text_input("🏢 Área", value=_area_default)
            departamento_edit = st.text_input("🏛️ Departamento", value=_dept_default)
            perfil_edit = st.text_input("🎓 Perfil Profesional", value=_perfil_default)
            valor_usd_edit = st.number_input("💵 Valor del Plan (USD)", min_value=0.0, step=0.01, value=float(plan_sel.get('valor_usd', 0)))
            observaciones_edit = st.text_area("📝 Observaciones", value=plan_sel.get('observaciones', ''), height=100)
            tasa_manual_edit = st.number_input("💱 Tasa USD a Lempira (HNL)", min_value=0.0, value=float(plan_sel.get('tasa_usd_hnl', 24.0)), step=0.01)

            dispositivo_asignado_edit = st.text_input("📱 Dispositivo Asignado", value=plan_sel.get('dispositivo_asignado', ''))
            marca_dispositivo_edit = st.text_input("🏷️ Marca", value=plan_sel.get('marca', ''))
            modelo_dispositivo_edit = st.text_input("🆔 Modelo", value=plan_sel.get('modelo', ''))
            serie_dispositivo_edit = st.text_input("🔢 Serie del Dispositivo", value=plan_sel.get('serie_dispositivo', ''))
            imei1_edit = st.text_input("📳 IMEI 1", value=plan_sel.get('imei1', ''))
            imei2_edit = st.text_input("📳 IMEI 2", value=plan_sel.get('imei2', ''))
            motivo_cambio_dispositivo_edit = st.text_input("📝 Motivo de cambio de dispositivo (si aplica)", "")


            guardar_edicion = st.form_submit_button("💾 Guardar cambios", disabled=not permiso_editar)

        # Fuera del formulario: botón de descarga de responsiva PDF
        pdf_bytes = generar_responsiva_completa_pdf(
            empleado=nombre_edit,
            departamento=area_edit,
            dispositivo=dispositivo_asignado_edit,
            marca=marca_dispositivo_edit,
            imei=f"{imei1_edit} / {imei2_edit}",
            modelo=modelo_dispositivo_edit,
            plan=perfil_edit,
            serie=serie_dispositivo_edit,
            area=area_edit,
            estado="NUEVO",
            cargador="Original",
            funcionalidad="100%",
            numero_corporativo=numero_edit,
            valor_equipo=valor_usd_edit,
            ciudad="COMAYAGUA",
            fecha=str(datetime.now().year),
            logo_path="km_motos_banner.png"
        )
        st.download_button(
            label="Generar responsiva PDF",
            data=pdf_bytes,
            file_name=f"responsiva_{nombre_edit}.pdf",
            mime="application/pdf",
            key="btn_responsiva_pdf"
        )

        if (dispositivo_asignado_edit != plan_sel.get('dispositivo_asignado', '') or
            marca_dispositivo_edit != plan_sel.get('marca', '') or
            modelo_dispositivo_edit != plan_sel.get('modelo', '') or
            serie_dispositivo_edit != plan_sel.get('serie_dispositivo', '') or
            imei1_edit != plan_sel.get('imei1', '') or
            imei2_edit != plan_sel.get('imei2', '')):
            if motivo_cambio_dispositivo_edit:
                cambio_entry = {
                    'fecha': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    'usuario': st.session_state.usuario_actual,
                    'motivo': motivo_cambio_dispositivo_edit,
                    'dispositivo_asignado': dispositivo_asignado_edit,
                    'marca': marca_dispositivo_edit,
                    'modelo': modelo_dispositivo_edit,
                    'serie_dispositivo': serie_dispositivo_edit,
                    'imei1': imei1_edit,
                    'imei2': imei2_edit
                }
                dispositivo_historial_actual.append(cambio_entry)
                guardar_planes()
                registrar_movimiento("Cambio de Dispositivo", f"{numero_edit} - {nombre_edit} - motivo: {motivo_cambio_dispositivo_edit}")
                st.info("ℹ️ Historial de dispositivo actualizado.")
            else:
                st.warning("⚠️ Has modificado información de dispositivo, por favor proporciona un motivo de cambio para registrar el historial.")

                # Registrar cambio de asignación si cambió nombre_personal
                asignaciones_historial_actual = plan_sel.get('asignaciones_historial', [])
                nombre_anterior = plan_sel.get('nombre_personal', '')

                # Si el historial está vacío y hay un nombre anterior, crear entrada inicial
                if not asignaciones_historial_actual and nombre_anterior:
                    asignaciones_historial_actual.append({
                        'fecha': plan_sel.get('fecha_creacion', datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                        'usuario': 'Sistema',
                        'nombre_personal': nombre_anterior
                    })
                
                # Registrar el cambio si el nombre cambió
                if nombre_edit != nombre_anterior:
                    cambio_asignacion = {
                        'fecha': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        'usuario': st.session_state.usuario_actual,
                        'nombre_personal': nombre_edit,
                        'nombre_anterior': nombre_anterior
                    }
                    asignaciones_historial_actual.append(cambio_asignacion)
                    registrar_movimiento("Cambio de Asignación", f"{numero_edit} - de: {nombre_anterior} a: {nombre_edit}")
                    st.info(f"✅ Asignación registrada: {nombre_anterior} → {nombre_edit}")

                st.session_state.planes[indice_seleccionado] = {
                    'numero': numero_edit,
                    'operador': (str(operador_edit).strip() or 'TIGO').upper(),
                    'nombre_personal': nombre_edit,
                    'area': area_edit,
                    'departamento': departamento_edit,
                    'perfil_profesional': perfil_edit,
                    'valor_usd': valor_usd_edit,
                    'tasa_usd_hnl': tasa_manual_edit,
                    'valor_hnl': valor_usd_edit * tasa_manual_edit,
                    'observaciones': observaciones_edit,
                    'dispositivo_asignado': dispositivo_asignado_edit,
                    'marca': marca_dispositivo_edit,
                    'modelo': modelo_dispositivo_edit,
                    'serie_dispositivo': serie_dispositivo_edit,
                    'imei1': imei1_edit,
                    'imei2': imei2_edit,
                    'dispositivo_historial': dispositivo_historial_actual,
                    'asignaciones_historial': asignaciones_historial_actual,
                    'fecha_creacion': plan_sel.get('fecha_creacion', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                }
                guardar_planes()
                registrar_movimiento("Editar Plan", f"{numero_edit} - {nombre_edit} - {area_edit} - {departamento_edit}")
                st.success("✅ Plan actualizado correctamente")
                try:
                    st.rerun()
                except Exception:
                    pass

        if st.button("🗑️ Eliminar plan seleccionado", disabled=not permiso_eliminar, key="btn_eliminar_plan_tab3"):
            eliminado = st.session_state.planes.pop(indice_seleccionado)
            guardar_planes()
            registrar_movimiento("Eliminar Plan", f"{eliminado.get('numero')} - {eliminado.get('nombre_personal')} - {eliminado.get('area')} - {eliminado.get('departamento')}")
            st.success("✅ Plan eliminado")
            try:
                st.rerun()
            except Exception:
                pass

        st.markdown("---")
        st.subheader("� Historial de Asignaciones")
        asignaciones = plan_sel.get('asignaciones_historial', [])
        if asignaciones:
            for evento in sorted(asignaciones, key=lambda x: x.get('fecha', ''), reverse=True):
                nombre_ant = evento.get('nombre_anterior', '(Sin asignación anterior)')
                precio_disp = evento.get('precio_dispositivo', None)
                detalles = f"- {evento.get('fecha')} | {evento.get('usuario')} | Asignación: {evento.get('nombre_personal')}"
                if precio_disp is not None:
                    detalles += f" | Precio Dispositivo: L {precio_disp:,.2f}"
                st.write(detalles)
                if evento.get('nombre_anterior'):
                    st.write(f"  Cambio de: {nombre_ant}")
        else:
            st.info("No hay historial de asignaciones registrado para este plan.")
        
        st.markdown("---")
        st.subheader("�📌 Historial de Dispositivo Asignado")
        historial = plan_sel.get('dispositivo_historial', [])
        if historial:
            for evento in sorted(historial, key=lambda x: x.get('fecha', ''), reverse=True):
                st.write(f"- {evento.get('fecha')} / {evento.get('usuario')} / Motivo: {evento.get('motivo')}")
                st.write(f"  Dispositivo: {evento.get('dispositivo_asignado')} | Marca: {evento.get('marca')} | Modelo: {evento.get('modelo')} | Serie: {evento.get('serie_dispositivo')} | IMEI1: {evento.get('imei1')} | IMEI2: {evento.get('imei2')}")
        else:
            st.info("Aún no hay cambios de dispositivo registrados para este plan.")
    
    else:
        st.info("📌 No hay planes registrados aún")

# ============ TAB 4: EMPLEADOS ============

if vista_actual == "empleados":
    # --- Estilo visual azul para la tabla de empleados ---
    st.markdown(
        """
        <style>
        [data-testid="stDataFrame"] {
            background: #2daafc;
            border: 3px solid #156fa3;
            border-radius: 18px;
            box-shadow: 0 6px 24px 0 rgba(45,170,252,0.18), 0 0 0 2px #2daafc55;
            overflow: hidden;
        }
        [data-testid="stDataFrame"] thead tr th {
            background: #156fa3;
            color: #fff;
            font-weight: 900;
            font-size: 1.13em;
            letter-spacing: 0.7px;
            border-bottom: 3px solid #2daafc;
            text-shadow: 0 1px 0 #fff8;
            text-align: center;
            box-shadow: 0 4px 16px 0 rgba(21,111,163,0.25), 0 1.5px 0 #2daafc;
        }
        [data-testid="stDataFrame"] tbody tr {
            background: rgba(255,255,255,0.10);
            transition: background 0.22s, color 0.22s;
        }
        [data-testid="stDataFrame"] tbody tr:hover {
            background: #b3e5fc !important;
            color: #156fa3 !important;
            font-weight: 700;
            border-left: 4px solid #156fa3 !important;
            border-right: 4px solid #156fa3 !important;
            box-shadow: 0 0 8px 2px #2daafc55;
        }
        [data-testid="stDataFrame"] td {
            font-size: 1.07em;
            font-weight: 500;
            color: #fff;
            text-align: center;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )
    st.subheader("👥 Catálogo de Empleados Activos")
    st.caption("Importa la base del personal activo para usarla como catálogo al editar asignaciones de líneas.")


    # --- Menú para importar empleados (CSV/Excel) ---
    st.markdown("### 📥 Importar empleados (CSV/Excel)")
    archivo_empleados = st.file_uploader(
        "Selecciona un archivo de hoja de cálculo con empleados",
        type=['csv', 'xlsx', 'xls'],
        help="Permite cargar CSV o Excel con columnas: nombre, perfil_profesional, area, departamento, tienda, zona, empresa",
        key="archivo_import_empleados",
    )

    hoja_empleados = None
    df_preview_emp = None
    if archivo_empleados is not None:
        nombre_archivo_emp = archivo_empleados.name.lower()
        if nombre_archivo_emp.endswith(('.xlsx', '.xls')):
            try:
                excel_file_emp = pd.ExcelFile(archivo_empleados)
                hoja_empleados = st.selectbox(
                    "Selecciona la hoja del Excel (empleados)",
                    options=excel_file_emp.sheet_names,
                    key="hoja_excel_empleados",
                )
                st.caption(f"Hojas detectadas: {', '.join(excel_file_emp.sheet_names)}")
            except Exception as e:
                st.error(f"❌ No pude leer las hojas del Excel: {e}")

        try:
            archivo_empleados.seek(0)
            if nombre_archivo_emp.endswith('.csv'):
                df_preview_emp = pd.read_csv(archivo_empleados).head(10)
            elif nombre_archivo_emp.endswith(('.xlsx', '.xls')) and hoja_empleados:
                df_preview_emp = pd.read_excel(archivo_empleados, sheet_name=hoja_empleados).head(10)

            if df_preview_emp is not None:
                df_preview_emp = preparar_dataframe_importacion(df_preview_emp)
                st.markdown("#### 👀 Vista previa (primeras 10 filas)")
                st.dataframe(df_preview_emp, width="stretch", hide_index=True)
                st.caption(f"Columnas detectadas: {', '.join([str(c) for c in df_preview_emp.columns])}")
        except Exception as e:
            st.warning(f"No se pudo mostrar vista previa: {e}")

    if archivo_empleados is not None:
        if st.button("🔄 Cargar empleados", disabled=not permiso_importar):
            try:
                nombre_archivo_emp = archivo_empleados.name.lower()
                archivo_empleados.seek(0)
                try:
                    if nombre_archivo_emp.endswith('.csv'):
                        df_emp = pd.read_csv(archivo_empleados, dtype=str)
                        registrar_debug_importacion("read_empleados", f"CSV rows={len(df_emp.index)} cols={list(df_emp.columns)}")
                    elif nombre_archivo_emp.endswith(('.xlsx', '.xls')):
                        if not hoja_empleados:
                            st.error("❌ Selecciona una hoja del Excel para continuar.")
                            st.stop()
                        archivo_empleados.seek(0)
                        try:
                            df_emp = pd.read_excel(archivo_empleados, sheet_name=hoja_empleados, engine='openpyxl', dtype=str)
                        except Exception:
                            # fallback a lectura por defecto si openpyxl falla
                            df_emp = pd.read_excel(archivo_empleados, sheet_name=hoja_empleados, dtype=str)
                        registrar_debug_importacion("read_empleados", f"Excel rows={len(df_emp.index)} cols={list(df_emp.columns)} sheet={hoja_empleados}")
                    else:
                        st.error("Formato no soportado para empleados.")
                        st.stop()
                except Exception as e:
                    registrar_debug_importacion("read_error", f"file={archivo_empleados.name}; error={e}")
                    st.error(f"❌ Error leyendo el archivo: {e}")
                    st.stop()

                # --- DEBUG: mostrar encabezados y primeras filas antes y después de normalizar ---
                try:
                    st.markdown("**DEBUG: Encabezados originales detectados**")
                    st.write(list(df_emp.columns))
                    st.markdown("**DEBUG: Primeras filas (raw)**")
                    st.dataframe(df_emp.head(5))
                except Exception:
                    pass

                # Aplicar normalización (misma función que en local)
                df_emp_preparado = None
                try:
                    # Para empleados usamos un flujo de normalización más simple (no eliminar filas)
                    df_emp_preparado = preparar_dataframe_empleados(df_emp.copy())
                except Exception as e:
                    st.error(f"❌ Error al preparar/normalizar dataframe: {e}")

                if df_emp_preparado is not None:
                    try:
                        st.markdown("**DEBUG: Encabezados después de preparar_dataframe_importacion()**")
                        st.write(list(df_emp_preparado.columns))
                        st.markdown("**DEBUG: Primeras filas (preparado)**")
                        st.dataframe(df_emp_preparado.head(5))
                    except Exception:
                        pass

                # Reasignar df_emp al dataframe preparado para el flujo existente
                if df_emp_preparado is not None:
                    df_emp = df_emp_preparado
                else:
                    # Si la preparación falló, continuar con el original para que el usuario vea lo que pasó
                    df_emp = df_emp

                existentes = {e.get('nombre', '').strip().upper() for e in st.session_state.get('empleados', [])}
                nuevos = 0
                duplicados = 0
                invalidas = 0

                for _, row in df_emp.iterrows():
                    # Acepta columna 'nombre' o 'nombre_personal'
                    nombre_raw = row.get('nombre') if 'nombre' in row.index else row.get('nombre_personal') if 'nombre_personal' in row.index else ''
                    nombre = str(nombre_raw).strip()
                    if not nombre:
                        invalidas += 1
                        continue
                    nombre_up = nombre.upper()
                    if nombre_up in existentes:
                        duplicados += 1
                        continue

                    nuevo_emp = {
                        'nombre': nombre_up,
                        'perfil_profesional': str(row.get('perfil_profesional', '')).strip().upper(),
                        'area': str(row.get('area', '')).strip().upper(),
                        'departamento': str(row.get('departamento', '')).strip().upper(),
                        'tienda': str(row.get('tienda', '')).strip().upper(),
                        'zona': str(row.get('zona', '')).strip().upper(),
                        'empresa': str(row.get('empresa', '')).strip().upper(),
                        'historial_cambios': [
                            {
                                'fecha': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                'usuario': st.session_state.get('usuario_actual', 'Sistema'),
                                'cambios': ['Creación de empleado (import)']
                            }
                        ]
                    }
                    st.session_state.empleados.append(nuevo_emp)
                    existentes.add(nombre_up)
                    nuevos += 1

                guardar_empleados()
                registrar_movimiento("Importar empleados", f"{nuevos} nuevos; {duplicados} duplicados; {invalidas} invalidas")
                st.success(f"Importación completada: {nuevos} nuevos; {duplicados} duplicados; {invalidas} inválidas.")
                st.rerun()
            except Exception as e:
                st.error(f"❌ Error importando empleados: {e}")


    # ---- Catálogo actual con edición profesional ----
    st.markdown(f"### 📋 Catálogo actual — {len(st.session_state.empleados)} empleados")
    if st.session_state.empleados:
        _df_cat = pd.DataFrame(st.session_state.empleados)
        for _c in ("nombre", "perfil_profesional", "area", "departamento", "tienda", "zona", "empresa"):
            if _c not in _df_cat.columns:
                _df_cat[_c] = ""
        _buscar_cat = st.text_input("🔍 Buscar en catálogo:", placeholder="Nombre, área, perfil...", key="buscar_emp_catalogo")
        if _buscar_cat.strip():
            _p = _buscar_cat.strip().lower()
            _mask = (
                _df_cat["nombre"].str.lower().str.contains(_p, na=False)
                | _df_cat["area"].str.lower().str.contains(_p, na=False)
                | _df_cat["departamento"].str.lower().str.contains(_p, na=False)
                | _df_cat["perfil_profesional"].str.lower().str.contains(_p, na=False)
            )
            _df_cat = _df_cat[_mask]

        columnas_vista = [c for c in ["nombre", "perfil_profesional", "area", "departamento", "tienda", "zona", "empresa"] if c in _df_cat.columns]
        _df_cat_vista = _df_cat[columnas_vista].copy()
        _df_cat_vista.insert(0, "No.", range(1, len(_df_cat_vista) + 1))

        # Tabla interactiva con selección de fila
        st.data_editor(
            _df_cat_vista,
            hide_index=True,
            use_container_width=True,
            height=350,
            column_config={
                "No.": st.column_config.NumberColumn("No.", format="%d", width="small"),
                "nombre": st.column_config.TextColumn("Nombre", width="large"),
                "perfil_profesional": st.column_config.TextColumn("Perfil", width="large"),
                "area": st.column_config.TextColumn("Área", width="medium"),
                "departamento": st.column_config.TextColumn("Departamento", width="medium"),
                "tienda": st.column_config.TextColumn("Tienda", width="medium"),
                "zona": st.column_config.TextColumn("Zona", width="medium"),
                "empresa": st.column_config.TextColumn("Empresa", width="medium"),
            },
            key="emp_table_editor"
        )

        # Selección profesional de empleado debajo de la tabla
        nombres_vista = _df_cat_vista["nombre"].tolist()
        idx_ver = st.selectbox(
            "Selecciona un empleado para ver o editar:",
            options=range(len(nombres_vista)),
            format_func=lambda i: f"{_df_cat_vista.iloc[i]['No.']}. {nombres_vista[i]}",
            key="ver_editar_emp_idx_selectbox"
        )
        if 0 <= idx_ver < len(_df_cat_vista):
            nombre_sel = _df_cat_vista.iloc[idx_ver]["nombre"]
            emp_sel = next((e for e in st.session_state.empleados if e.get("nombre", "") == nombre_sel), None)
            if emp_sel:
                # st.markdown(f"#### Información completa de: **{emp_sel.get('nombre','')}**")
                # st.write(emp_sel)
                form_key = f"form_edit_emp_{idx_ver}_{emp_sel.get('nombre','').replace(' ','_')}"
                with st.form(key=form_key):
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        nombre_edit = st.text_input("👤 Nombre completo", value=emp_sel.get("nombre", ""))
                        area_edit = st.text_input("🏢 Área", value=emp_sel.get("area", ""))
                        tienda_edit = st.text_input("🏬 Tienda", value=emp_sel.get("tienda", ""))
                    with col2:
                        perfil_edit = st.text_input("🎓 Perfil profesional", value=emp_sel.get("perfil_profesional", ""))
                        dept_edit = st.text_input("🏛️ Departamento", value=emp_sel.get("departamento", ""))
                        zona_edit = st.text_input("🌎 Zona", value=emp_sel.get("zona", ""))
                    with col3:
                        empresa_edit = st.text_input("🏢 Empresa", value=emp_sel.get("empresa", ""))
                    guardar_cambios = st.form_submit_button("💾 Guardar cambios")
                    eliminar_emp = st.form_submit_button("🗑️ Eliminar empleado")
                    if guardar_cambios:
                        cambios = []
                        for campo, nuevo, viejo in [
                            ("nombre", nombre_edit.strip().upper(), emp_sel.get("nombre", "")),
                            ("perfil_profesional", perfil_edit.strip().upper(), emp_sel.get("perfil_profesional", "")),
                            ("area", area_edit.strip().upper(), emp_sel.get("area", "")),
                            ("departamento", dept_edit.strip().upper(), emp_sel.get("departamento", "")),
                            ("tienda", tienda_edit.strip().upper(), emp_sel.get("tienda", "")),
                            ("zona", zona_edit.strip().upper(), emp_sel.get("zona", "")),
                            ("empresa", empresa_edit.strip().upper(), emp_sel.get("empresa", "")),
                        ]:
                            if nuevo != viejo:
                                cambios.append(f"{campo}: '{viejo}' → '{nuevo}'")
                                emp_sel[campo] = nuevo
                        if cambios:
                            if "historial_cambios" not in emp_sel or not isinstance(emp_sel["historial_cambios"], list):
                                emp_sel["historial_cambios"] = []
                            emp_sel["historial_cambios"].append({
                                "fecha": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                "usuario": st.session_state.get("usuario_actual", "Sistema"),
                                "cambios": cambios
                            })
                            guardar_empleados()
                            registrar_movimiento("Editar empleado", f"{emp_sel.get('nombre','')} | {'; '.join(cambios)}")
                            st.success("✅ Cambios guardados.")
                            st.rerun()
                        else:
                            st.info("No hay cambios para guardar.")
                    if eliminar_emp:
                        if st.session_state.get("confirmar_eliminar_emp", None) == idx_ver:
                            st.session_state.empleados.pop(idx_ver)
                            guardar_empleados()
                            registrar_movimiento("Eliminar empleado", f"{emp_sel.get('nombre','')}")
                            st.success("✅ Empleado eliminado.")
                            st.session_state["confirmar_eliminar_emp"] = None
                            st.rerun()
                        else:
                            st.session_state["confirmar_eliminar_emp"] = idx_ver
                            st.warning("Presiona de nuevo para confirmar la eliminación.")

                # Mostrar historial de cambios
                st.markdown("#### 🕓 Historial de cambios de este empleado")
                historial = emp_sel.get("historial_cambios", [])
                if historial:
                    for evento in sorted(historial, key=lambda x: x.get('fecha', ''), reverse=True):
                        st.write(f"- {evento.get('fecha')} / {evento.get('usuario')} / Cambios: {', '.join(evento.get('cambios', []))}")
                else:
                    st.info("No hay historial de cambios registrado para este empleado.")

        # --- Selección de empleado para editar/eliminar ---
        st.markdown("### ✏️ Editar o eliminar empleado")
        empleados_nombres = _df_cat_vista["nombre"].tolist()
        if empleados_nombres:
            idx_sel = st.selectbox("Selecciona un empleado:", options=range(len(empleados_nombres)), format_func=lambda i: empleados_nombres[i], key="sel_emp_edit")
            emp_sel = None
            if 0 <= idx_sel < len(st.session_state.empleados):
                emp_sel = st.session_state.empleados[idx_sel]
            if emp_sel:
                with st.form(key=f"form_edit_emp_{idx_sel}"):
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        nombre_edit = st.text_input("👤 Nombre completo", value=emp_sel.get("nombre", ""))
                        area_edit = st.text_input("🏢 Área", value=emp_sel.get("area", ""))
                        tienda_edit = st.text_input("🏬 Tienda", value=emp_sel.get("tienda", ""))
                    with col2:
                        perfil_edit = st.text_input("🎓 Perfil profesional", value=emp_sel.get("perfil_profesional", ""))
                        dept_edit = st.text_input("🏛️ Departamento", value=emp_sel.get("departamento", ""))
                        zona_edit = st.text_input("🌎 Zona", value=emp_sel.get("zona", ""))
                    with col3:
                        empresa_edit = st.text_input("🏢 Empresa", value=emp_sel.get("empresa", ""))
                    guardar_cambios = st.form_submit_button("💾 Guardar cambios")
                    eliminar_emp = st.form_submit_button("🗑️ Eliminar empleado")
                    if guardar_cambios:
                        cambios = []
                        for campo, nuevo, viejo in [
                            ("nombre", nombre_edit.strip().upper(), emp_sel.get("nombre", "")),
                            ("perfil_profesional", perfil_edit.strip().upper(), emp_sel.get("perfil_profesional", "")),
                            ("area", area_edit.strip().upper(), emp_sel.get("area", "")),
                            ("departamento", dept_edit.strip().upper(), emp_sel.get("departamento", "")),
                            ("tienda", tienda_edit.strip().upper(), emp_sel.get("tienda", "")),
                            ("zona", zona_edit.strip().upper(), emp_sel.get("zona", "")),
                            ("empresa", empresa_edit.strip().upper(), emp_sel.get("empresa", "")),
                        ]:
                            if nuevo != viejo:
                                cambios.append(f"{campo}: '{viejo}' → '{nuevo}'")
                                emp_sel[campo] = nuevo
                        if cambios:
                            if "historial_cambios" not in emp_sel or not isinstance(emp_sel["historial_cambios"], list):
                                emp_sel["historial_cambios"] = []
                            emp_sel["historial_cambios"].append({
                                "fecha": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                "usuario": st.session_state.get("usuario_actual", "Sistema"),
                                "cambios": cambios
                            })
                            guardar_empleados()
                            registrar_movimiento("Editar empleado", f"{emp_sel.get('nombre','')} | {'; '.join(cambios)}")
                            st.success("✅ Cambios guardados.")
                            st.rerun()
                        else:
                            st.info("No hay cambios para guardar.")
                    if eliminar_emp:
                        if st.session_state.get("confirmar_eliminar_emp", None) == idx_sel:
                            st.session_state.empleados.pop(idx_sel)
                            guardar_empleados()
                            registrar_movimiento("Eliminar empleado", f"{emp_sel.get('nombre','')}")
                            st.success("✅ Empleado eliminado.")
                            st.session_state["confirmar_eliminar_emp"] = None
                            st.rerun()
                        else:
                            st.session_state["confirmar_eliminar_emp"] = idx_sel
                            st.warning("Presiona de nuevo para confirmar la eliminación.")

                # Mostrar historial de cambios
                st.markdown("#### 🕓 Historial de cambios de este empleado")
                historial = emp_sel.get("historial_cambios", [])
                if historial:
                    for evento in sorted(historial, key=lambda x: x.get('fecha', ''), reverse=True):
                        st.write(f"- {evento.get('fecha')} / {evento.get('usuario')} / Cambios: {', '.join(evento.get('cambios', []))}")
                else:
                    st.info("No hay historial de cambios registrado para este empleado.")

        st.markdown("---")
        st.markdown("### ➕ Agregar empleado manualmente")
        with st.form(key="form_agregar_empleado"):
            _c1, _c2, _c3 = st.columns(3)
            with _c1:
                _emp_nombre = st.text_input("👤 Nombre completo")
                _emp_area = st.text_input("🏢 Área")
                _emp_tienda = st.text_input("🏬 Tienda")
            with _c2:
                _emp_perfil = st.text_input("🎓 Perfil profesional")
                _emp_dept = st.text_input("🏛️ Departamento")
                _emp_zona = st.text_input("🌎 Zona")
            with _c3:
                _emp_empresa = st.text_input("🏢 Empresa")
            if st.form_submit_button("💾 Agregar al catálogo"):
                _nombre_nuevo = _emp_nombre.strip().upper()
                if not _nombre_nuevo:
                    st.error("❌ El nombre es obligatorio.")
                else:
                    _nombres_ya = {e.get("nombre", "") for e in st.session_state.empleados}
                    if _nombre_nuevo in _nombres_ya:
                        st.warning(f"⚠️ Ya existe un empleado con el nombre '{_nombre_nuevo}'.")
                    else:
                        nuevo_emp = {
                            "nombre": _nombre_nuevo,
                            "perfil_profesional": _emp_perfil.strip().upper(),
                            "area": _emp_area.strip().upper(),
                            "departamento": _emp_dept.strip().upper(),
                            "tienda": _emp_tienda.strip().upper(),
                            "zona": _emp_zona.strip().upper(),
                            "empresa": _emp_empresa.strip().upper(),
                            "historial_cambios": [
                                {
                                    "fecha": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                    "usuario": st.session_state.get("usuario_actual", "Sistema"),
                                    "cambios": ["Creación de empleado"]
                                }
                            ]
                        }
                        st.session_state.empleados.append(nuevo_emp)
                        guardar_empleados()
                        st.success(f"✅ Empleado '{_nombre_nuevo}' agregado al catálogo.")
                        st.rerun()

        st.markdown("---")
        if st.button("🗑️ Eliminar catálogo completo", key="btn_limpiar_empleados"):
            st.session_state["confirmar_limpiar_empleados"] = True
        if st.session_state.get("confirmar_limpiar_empleados", False):
            st.warning("⚠️ Esto eliminará todos los empleados del catálogo.")
            _col_e1, _col_e2 = st.columns(2)
            with _col_e1:
                if st.button("✅ Sí, eliminar catálogo", key="btn_confirmar_limpiar_emp"):
                    st.session_state.empleados = []
                    guardar_empleados()
                    st.session_state["confirmar_limpiar_empleados"] = False
                    st.rerun()
            with _col_e2:
                if st.button("❌ Cancelar", key="btn_cancelar_limpiar_emp"):
                    st.session_state["confirmar_limpiar_empleados"] = False
                    st.rerun()
    else:
        st.info("📌 El catálogo está vacío. Importa un archivo CSV o Excel con los datos del personal.")


# ============ TAB 5: CONFIGURACIÓN ============
if vista_actual == "configuracion":
    st.subheader("⚙️ Configuración y Datos")

    # ---- Recalcular planes con tasa actual ----
    st.markdown("### 💱 Recalcular Planes con Tasa Actual")
    tasa_vigente_tab4 = float(st.session_state.get("tasa_usd_hnl", 24.0))
    st.info(
        f"Tasa vigente: **{tasa_vigente_tab4:.4f} HNL/USD** | "
        f"Fuente: {st.session_state.get('tasa_fuente', 'N/D')} | "
        f"Actualizada: {st.session_state.get('tasa_actualizada_en', 'N/D')}"
    )
    if st.button(
        f"🔁 Recalcular valor HNL de todos los planes ({len(st.session_state.planes)} registros)",
        disabled=not st.session_state.get("puede_editar", False),
    ):
        if st.session_state.planes:
            actualizados = 0
            for plan in st.session_state.planes:
                try:
                    plan["tasa_usd_hnl"] = tasa_vigente_tab4
                    plan["valor_hnl"] = round(float(plan.get("valor_usd", 0)) * tasa_vigente_tab4, 2)
                    actualizados += 1
                except Exception:
                    pass
            guardar_planes()
            st.success(
                f"✅ {actualizados} planes recalculados. "
                f"Valor HNL = USD x {tasa_vigente_tab4:.4f}"
            )
        else:
            st.warning("No hay planes registrados para recalcular.")

    st.markdown("---")

    # ---- Historial de tasa USD/HNL ----
    st.markdown("### 📈 Historial de Tasa USD/HNL")
    if os.path.exists(ARCHIVO_TASA_HISTORIAL):
        try:
            with open(ARCHIVO_TASA_HISTORIAL, "r", encoding="utf-8") as _fh:
                _hist = json.load(_fh)
            if _hist:
                _df_hist = pd.DataFrame(_hist)
                _df_hist["datetime"] = pd.to_datetime(
                    _df_hist["fecha"] + " " + _df_hist["hora"], errors="coerce"
                )
                _df_hist = _df_hist.sort_values("datetime", ascending=False).reset_index(drop=True)
                _df_hist["variacion"] = (
                    _df_hist["tasa"].diff(-1)
                    .apply(lambda x: f"+{x:.4f}" if isinstance(x, float) and x > 0 else (f"{x:.4f}" if isinstance(x, float) else ""))
                )
                col_hist_graf, col_hist_tabla = st.columns([3, 2])
                with col_hist_graf:
                    import copy as _copy
                    _df_graf = _df_hist[["datetime", "tasa"]].dropna().sort_values("datetime")
                    if not _df_graf.empty:
                        st.line_chart(_df_graf.set_index("datetime")["tasa"])
                with col_hist_tabla:
                    st.dataframe(
                        _df_hist[["fecha", "hora", "tasa", "fuente", "variacion"]].head(30),
                        use_container_width=True,
                        hide_index=True,
                    )
                # Boton para descargar historial CSV
                csv_hist = _df_hist[["fecha", "hora", "tasa", "fuente"]].to_csv(index=False)
                st.download_button(
                    label="📥 Descargar historial CSV",
                    data=csv_hist.encode("utf-8"),
                    file_name=f"historial_tasa_usd_hnl_{datetime.now().strftime('%Y%m%d')}.csv",
                    mime="text/csv",
                )
            else:
                st.caption("Sin registros en el historial aun.")
        except Exception as _e:
            st.caption(f"No se pudo leer el historial: {_e}")
    else:
        st.caption("El historial se creara automaticamente al actualizar la tasa.")

    st.markdown("---")
    
    col_config1, col_config2 = st.columns(2)
    
    with col_config1:
        st.info(f"**Total de Planes:** {len(st.session_state.planes)}")
    
    with col_config2:
        if st.session_state.planes:
            df_temp = pd.DataFrame(st.session_state.planes)
            st.info(f"**Inversión Total:** ${df_temp['valor_usd'].sum():,.2f}")
            st.info(f"**Última Actualización:** {df_temp['fecha_creacion'].iloc[-1]}")

    st.markdown("### 🧹 Mantenimiento de Números")
    st.caption("Normaliza formato (+504...) y elimina duplicados exactos por número.")
    if st.button("🧹 Normalizar y deduplicar números", disabled=not permiso_editar):
        antes = len(st.session_state.planes)
        planes_limpios, cambios_numero, duplicados_eliminados = normalizar_y_deduplicar_planes(st.session_state.planes)
        st.session_state.planes = planes_limpios
        guardar_planes()
        registrar_movimiento(
            "Mantenimiento Números",
            f"cambios_formato={cambios_numero}; duplicados_eliminados={duplicados_eliminados}; total_antes={antes}; total_despues={len(planes_limpios)}",
        )
        st.success(
            f"Mantenimiento completado. Formatos corregidos: {cambios_numero}. Duplicados eliminados: {duplicados_eliminados}."
        )
        st.rerun()
    
    st.markdown("---")
    
    st.subheader("📤 Importar/Exportar Datos")
    
    col_import, col_export = st.columns(2)
    
    with col_import:
        st.markdown("### 📥 Importar datos masivos (CSV/Excel)")
        archivo_datos = st.file_uploader(
            "Selecciona un archivo de hoja de calculo",
            type=['csv', 'xlsx', 'xls'],
            help="Formatos permitidos: CSV, Excel (.xlsx, .xls)",
        )

        hoja_excel = None
        df_preview = None
        if archivo_datos is not None:
            nombre_archivo = archivo_datos.name.lower()
            if nombre_archivo.endswith('.xlsx') or nombre_archivo.endswith('.xls'):
                try:
                    excel_file = pd.ExcelFile(archivo_datos)
                    hoja_excel = st.selectbox(
                        "Selecciona la hoja del Excel",
                        options=excel_file.sheet_names,
                        key="hoja_excel_importacion",
                    )
                    st.caption(f"Hojas detectadas: {', '.join(excel_file.sheet_names)}")
                except ImportError:
                    st.error("❌ Falta una libreria para leer Excel. Instala openpyxl y vuelve a intentar.")
                except Exception as e:
                    st.error(f"❌ No pude leer las hojas del Excel: {e}")

            # Vista previa antes de importar (10 filas)
            try:
                archivo_datos.seek(0)
                if nombre_archivo.endswith('.csv'):
                    df_preview = pd.read_csv(archivo_datos).head(10)
                elif (nombre_archivo.endswith('.xlsx') or nombre_archivo.endswith('.xls')) and hoja_excel:
                    df_preview = pd.read_excel(archivo_datos, sheet_name=hoja_excel).head(10)

                if df_preview is not None:
                    df_preview = preparar_dataframe_importacion(df_preview)
                    st.markdown("#### 👀 Vista previa (primeras 10 filas)")
                    st.dataframe(df_preview, width="stretch", hide_index=True)
                    st.caption(f"Columnas detectadas: {', '.join([str(c) for c in df_preview.columns])}")
            except Exception as e:
                st.warning(f"No se pudo mostrar vista previa: {e}")

        if archivo_datos is not None:
            if st.button("🔄 Cargar datos masivos", disabled=not permiso_importar):
                try:
                    nombre_archivo = archivo_datos.name.lower()
                    if nombre_archivo.endswith('.csv'):
                        df_importado = pd.read_csv(archivo_datos)
                        tipo_fuente = "CSV"
                    elif nombre_archivo.endswith('.xlsx') or nombre_archivo.endswith('.xls'):
                        if not hoja_excel:
                            st.error("❌ Selecciona una hoja del Excel para continuar.")
                            st.stop()
                        archivo_datos.seek(0)
                        df_importado = pd.read_excel(archivo_datos, sheet_name=hoja_excel)
                        tipo_fuente = f"Excel ({hoja_excel})"
                    else:
                        st.error("❌ Formato no soportado. Usa CSV, XLSX o XLS.")
                        st.stop()

                    # Normaliza nombres de columnas para soportar variaciones de encabezados.
                    df_importado = preparar_dataframe_importacion(df_importado)

                    columnas_requeridas = ['numero', 'nombre_personal', 'area', 'departamento', 'valor_usd']
                    if not all(col in df_importado.columns for col in columnas_requeridas):
                        st.error(f"❌ El archivo debe tener las columnas: {', '.join(columnas_requeridas)}")
                    else:
                        df_importado['numero'] = df_importado['numero'].apply(normalizar_numero_telefonico)
                        if 'operador' not in df_importado.columns:
                            df_importado['operador'] = 'TIGO'
                        else:
                            df_importado['operador'] = (
                                df_importado['operador']
                                .fillna('TIGO')
                                .astype(str)
                                .str.strip()
                                .replace('', 'TIGO')
                                .str.upper()
                            )
                        df_importado = df_importado[df_importado['numero'].astype(str).str.strip() != ""].copy()
                        df_importado['valor_usd'] = pd.to_numeric(df_importado['valor_usd'], errors='coerce').fillna(0.0)

                        if 'tasa_usd_hnl' not in df_importado.columns:
                            df_importado['tasa_usd_hnl'] = float(st.session_state.tasa_usd_hnl)
                        else:
                            df_importado['tasa_usd_hnl'] = pd.to_numeric(df_importado['tasa_usd_hnl'], errors='coerce').fillna(float(st.session_state.tasa_usd_hnl))

                        if 'valor_hnl' not in df_importado.columns:
                            df_importado['valor_hnl'] = df_importado['valor_usd'] * df_importado['tasa_usd_hnl']
                        else:
                            df_importado['valor_hnl'] = pd.to_numeric(df_importado['valor_hnl'], errors='coerce').fillna(df_importado['valor_usd'] * df_importado['tasa_usd_hnl'])

                        if 'fecha_creacion' not in df_importado.columns:
                            fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            df_importado['fecha_creacion'] = fecha_actual

                        planes_limpios, cambios_numero, duplicados_eliminados = normalizar_y_deduplicar_planes(df_importado.to_dict('records'))
                        st.session_state.planes = planes_limpios
                        guardar_planes()
                        registrar_movimiento(
                            "Importar datos",
                            f"{len(st.session_state.planes)} planes importados desde {tipo_fuente}; formato_corregido={cambios_numero}; duplicados_eliminados={duplicados_eliminados}",
                        )
                        st.success(
                            f"✅ {len(st.session_state.planes)} planes importados desde {tipo_fuente}. "
                            f"Formato corregido: {cambios_numero}. Duplicados eliminados: {duplicados_eliminados}."
                        )
                        st.rerun()
                except ImportError:
                    st.error("❌ Falta una libreria para leer Excel. Instala openpyxl y vuelve a intentar.")
                except Exception as e:
                    st.error(f"❌ Error al cargar el archivo: {str(e)}")
    
    with col_export:
        st.markdown("### 📤 Exportar Datos")
        if st.session_state.planes:
            _planes_serializados_cfg = _serializar_planes_cache(st.session_state.planes)
            _signature_cfg = hashlib.sha1(_planes_serializados_cfg.encode("utf-8", errors="ignore")).hexdigest()[:16]
            _export_ready = st.session_state.get("config_export_ready")
            if isinstance(_export_ready, dict) and _export_ready.get("signature") != _signature_cfg:
                st.session_state.pop("config_export_ready", None)
                _export_ready = None

            formato_cfg = st.selectbox(
                "Formato de exportación",
                ["CSV", "Excel", "JSON", "PDF"],
                key="formato_export_configuracion",
            )

            if st.button("📦 Preparar archivo", disabled=not permiso_exportar, use_container_width=True):
                from io import BytesIO

                df = _construir_dataframe_planes_cache(_planes_serializados_cfg)
                df_export_full = preparar_tabla_exportacion(df, incluir_total=True)
                df_export_detalle = construir_tabla_planes_profesional(df)
                _sufijo = datetime.now().strftime('%Y%m%d')
                _nombre_base = f"planes_corporativos_{_sufijo}"

                payload = {
                    "signature": _signature_cfg,
                    "format": formato_cfg,
                    "label": "",
                    "data": None,
                    "file_name": "",
                    "mime": "application/octet-stream",
                }

                if formato_cfg == "CSV":
                    payload.update({
                        "label": "📄 Descargar como CSV",
                        "data": df_export_full.to_csv(index=False).encode("utf-8"),
                        "file_name": f"{_nombre_base}.csv",
                        "mime": "text/csv",
                    })
                elif formato_cfg == "JSON":
                    payload.update({
                        "label": "📋 Descargar como JSON",
                        "data": json.dumps(st.session_state.planes, indent=2, ensure_ascii=False, default=str),
                        "file_name": f"{_nombre_base}.json",
                        "mime": "application/json",
                    })
                elif formato_cfg == "Excel":
                    excel_buffer = BytesIO()
                    excel_ok = False
                    for engine in ("openpyxl", "xlsxwriter"):
                        try:
                            with pd.ExcelWriter(excel_buffer, engine=engine) as writer:
                                df_export_full.to_excel(writer, index=False, sheet_name='Planes')
                                resumen_area = (
                                    df.groupby('area', dropna=False)['valor_usd']
                                    .sum()
                                    .reset_index()
                                    .sort_values('valor_usd', ascending=False)
                                )
                                resumen_area.to_excel(writer, index=False, sheet_name='Resumen_Area')
                                resumen_depto = (
                                    df.groupby('departamento', dropna=False)
                                    .size()
                                    .reset_index(name='cantidad_lineas')
                                    .sort_values('cantidad_lineas', ascending=False)
                                )
                                resumen_depto.to_excel(writer, index=False, sheet_name='Resumen_Departamento')
                            excel_ok = True
                            break
                        except ModuleNotFoundError:
                            excel_buffer = BytesIO()

                    if excel_ok:
                        excel_buffer.seek(0)
                        payload.update({
                            "label": "📊 Descargar como Excel",
                            "data": excel_buffer.getvalue(),
                            "file_name": f"{_nombre_base}.xlsx",
                            "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        })
                    else:
                        st.warning("⚠️ No se encontró biblioteca para generar Excel (xlsxwriter/openpyxl).")
                elif formato_cfg == "PDF":
                    if FPDF is not None:
                        _pdf = FPDF(format='letter')
                        _pdf.set_auto_page_break(auto=True, margin=15)
                        _pdf.add_page()
                        _pdf.set_font("Arial", size=12)
                        _pdf.cell(0, 10, _texto_seguro_pdf("Reporte de PLANES CORPORATIVOS KM MOTOS"), ln=True, align='C')
                        _pdf.set_font("Arial", size=10)

                        areas_incluidas = df_export_detalle['area'].unique()
                        areas_str = ", ".join(areas_incluidas) if len(areas_incluidas) <= 5 else f"{len(areas_incluidas)} áreas"
                        _pdf.cell(0, 8, _texto_seguro_pdf(f"Áreas: {areas_str}"), ln=True, align='C')
                        _pdf.cell(0, 8, _texto_seguro_pdf(f"Total de lineas: {len(df_export_detalle)}"), ln=True, align='C')
                        _pdf.cell(0, 10, "", ln=True)

                        mapa_planes = {
                            str(p.get('numero', '')): p
                            for p in st.session_state.planes
                        }

                        for _, row in df_export_detalle.iterrows():
                            line = _texto_seguro_pdf(
                                f"{row.get('item', '')}. {row.get('numero', '')} | {row.get('nombre_personal', '')} | {row.get('perfil_profesional', '')} | "
                                f"{row.get('area', '')} | {row.get('departamento', '')} | USD {float(row.get('valor_usd', 0)):,.2f} | "
                                f"HNL {float(row.get('valor_hnl', 0)):,.2f} | Tasa {float(row.get('tasa_usd_hnl', 0)):,.2f}"
                            )
                            _pdf.multi_cell(180, 8, line)
                            observ = _texto_seguro_pdf(row.get('observaciones', ''))
                            if observ:
                                _pdf.multi_cell(180, 8, f"Observaciones: {observ}")

                            plan_obj = mapa_planes.get(str(row.get('numero', '')))
                            if plan_obj:
                                asignaciones = plan_obj.get('asignaciones_historial', [])
                                if asignaciones:
                                    _pdf.set_font("Arial", size=8)
                                    _pdf.set_x(_pdf.l_margin)
                                    _pdf.cell(0, 4, "Historial de Asignaciones:", ln=True)
                                    for asig in sorted(asignaciones, key=lambda x: x.get('fecha', ''), reverse=True):
                                        hist_line = _texto_seguro_pdf(
                                            f"  - {asig.get('fecha')} | {asig.get('nombre_personal')} | Por: {asig.get('usuario')}"
                                        )
                                        _pdf.cell(0, 4, hist_line, ln=True)
                                    _pdf.set_font("Arial", size=10)

                            _pdf.ln(1)

                        _pdf.set_font("Arial", 'B', 10)
                        _pdf.cell(0, 8, _texto_seguro_pdf(f"TOTAL DE LINEAS: {len(df_export_detalle)}"), ln=True)
                        _total_hnl_pdf = float(pd.to_numeric(_df_export_detalle.get('valor_hnl', 0), errors='coerce').fillna(0).sum())
                        _pdf.cell(0, 8, _texto_seguro_pdf(f"TOTAL EN LEMPIRAS: HNL {_total_hnl_pdf:,.2f}"), ln=True)
                        
                        # Agregar firma del usuario al final
                        _pdf.ln(4)
                        _pdf.set_font("Arial", size=9)
                        _ancho_util = max(10, _pdf.w - _pdf.l_margin - _pdf.r_margin)
                        _pdf.set_x(_pdf.l_margin)
                        firma_texto = f"Reporte generado por: {_usuario_reporte} | Fecha: {_fecha_hora_reporte}"
                        _pdf.multi_cell(_ancho_util, 6, _texto_seguro_pdf(firma_texto))

                        payload.update({
                            "label": "📕 Descargar como PDF",
                            "data": _pdf_a_bytes(_pdf),
                            "file_name": f"{_nombre_base}.pdf",
                            "mime": "application/pdf",
                        })


            _export_ready = st.session_state.get("config_export_ready")
            if isinstance(_export_ready, dict) and _export_ready.get("signature") == _signature_cfg:
                st.download_button(
                    label=_export_ready.get("label", "Descargar"),
                    data=_export_ready.get("data"),
                    file_name=_export_ready.get("file_name", "reporte"),
                    mime=_export_ready.get("mime", "application/octet-stream"),
                    disabled=not permiso_exportar,
                    use_container_width=True,
                )
            else:
                st.caption("Selecciona formato y presiona 'Preparar archivo' para generar la descarga.")
        else:
            st.info("No hay datos para exportar")
    
    st.markdown("---")
    st.subheader("👥 Gestión de Usuarios")

    if st.session_state.rol in ["superadministrador", "administrador"]:
        st.info("Crea usuarios, define rol y decide si cada cuenta puede editar o solo visualizar.")
        usuarios = gestor_usuarios.obtener_usuarios()

        with st.form(key='form_usuario'):
            st.markdown("### ➕ Crear Usuario")
            usuario_nuevo = st.text_input("Usuario:")
            contrasena_nuevo = st.text_input("Contraseña:", type='password')
            email_nuevo = st.text_input("Email (opcional):")
            puede_editar_nuevo = st.checkbox("Puede editar datos", value=True)

            if st.session_state.rol == "superadministrador":
                rol_nuevo = st.selectbox("Rol:", ["usuario", "administrador", "superadministrador"])
            else:
                rol_nuevo = "usuario"
                st.info("Como administrador, solo puedes crear usuarios con rol usuario.")

            crear = st.form_submit_button("🔐 Crear usuario")
            if crear:
                if not usuario_nuevo or not contrasena_nuevo:
                    st.error("Usuario y contraseña son obligatorios")
                elif st.session_state.rol != "superadministrador" and rol_nuevo != "usuario":
                    st.error("Solo superadministrador puede crear administradores o superadministradores")
                else:
                    exito, msg = gestor_usuarios.crear_usuario(
                        usuario_nuevo,
                        contrasena_nuevo,
                        rol_nuevo,
                        email_nuevo,
                        puede_editar_nuevo,
                    )
                    if exito:
                        registrar_movimiento(
                            "Crear Usuario",
                            f"{usuario_nuevo} - rol={rol_nuevo} - puede_editar={puede_editar_nuevo}",
                        )
                        st.success(msg)
                        st.rerun()
                    else:
                        st.error(msg)

        st.markdown("---")
        st.subheader("Usuarios registrados")
        tabla = []
        for u in usuarios:
            data_u = gestor_usuarios.usuarios.get(u, {})
            permisos_u = data_u.get("permisos", {})
            tabla.append({
                "Usuario": u,
                "Rol": data_u.get("rol", "usuario"),
                "Puede editar": "Sí" if data_u.get("puede_editar", False) else "No",
                "Crear": "Sí" if permisos_u.get("crear", False) else "No",
                "Editar": "Sí" if permisos_u.get("editar", False) else "No",
                "Eliminar": "Sí" if permisos_u.get("eliminar", False) else "No",
                "Importar": "Sí" if permisos_u.get("importar", False) else "No",
                "Exportar": "Sí" if permisos_u.get("exportar", False) else "No",
                "Email": data_u.get("email", ""),
            })
        st.dataframe(pd.DataFrame(tabla), width="stretch", hide_index=True)

        st.markdown("### ✏️ Editar Permisos / Contraseña")
        usuario_obj = st.selectbox("Selecciona un usuario", options=usuarios, key="usuario_objetivo")
        data_u = gestor_usuarios.usuarios.get(usuario_obj, {})
        rol_actual_obj = data_u.get("rol", "usuario")
        puede_gestionar_obj = st.session_state.rol == "superadministrador" or rol_actual_obj == "usuario"

        if not puede_gestionar_obj:
            st.warning("Solo superadministrador puede editar cuentas administrador/superadministrador.")
        else:
            with st.form(key="form_editar_usuario"):
                nueva_password = st.text_input("Nueva contraseña (opcional)", type="password")
                nuevo_email = st.text_input("Email", value=data_u.get("email", ""))
                puede_editar_obj = st.checkbox("Puede editar datos", value=bool(data_u.get("puede_editar", False)))

                permisos_actuales = data_u.get("permisos", {})
                st.markdown("Permisos finos")
                colp1, colp2, colp3 = st.columns(3)
                with colp1:
                    perm_crear = st.checkbox("Crear", value=bool(permisos_actuales.get("crear", False)))
                    perm_editar = st.checkbox("Editar", value=bool(permisos_actuales.get("editar", False)))
                with colp2:
                    perm_eliminar = st.checkbox("Eliminar", value=bool(permisos_actuales.get("eliminar", False)))
                    perm_importar = st.checkbox("Importar", value=bool(permisos_actuales.get("importar", False)))
                with colp3:
                    perm_exportar = st.checkbox("Exportar", value=bool(permisos_actuales.get("exportar", True)))

                permisos_nuevos = {
                    "crear": perm_crear,
                    "editar": perm_editar,
                    "eliminar": perm_eliminar,
                    "importar": perm_importar,
                    "exportar": perm_exportar,
                }

                if st.session_state.rol == "superadministrador":
                    nuevo_rol = st.selectbox(
                        "Rol",
                        ["usuario", "administrador", "superadministrador"],
                        index=["usuario", "administrador", "superadministrador"].index(rol_actual_obj)
                        if rol_actual_obj in ["usuario", "administrador", "superadministrador"]
                        else 0,
                    )
                else:
                    nuevo_rol = None
                    st.info(f"Rol actual: {rol_actual_obj}")

                guardar_cambios = st.form_submit_button("💾 Guardar cambios del usuario")
                if guardar_cambios:
                    exito, msg = gestor_usuarios.actualizar_usuario(
                        usuario_obj,
                        contrasena=nueva_password if nueva_password else None,
                        rol=nuevo_rol,
                        email=nuevo_email,
                        puede_editar=puede_editar_obj,
                        permisos=permisos_nuevos,
                    )
                    if exito:
                        registrar_movimiento(
                            "Editar Usuario",
                            f"{usuario_obj} - rol={nuevo_rol or rol_actual_obj} - editar={puede_editar_obj} - permisos={permisos_nuevos}",
                        )
                        st.success(msg)
                        if usuario_obj == st.session_state.usuario_actual:
                            st.session_state.puede_editar = bool(puede_editar_obj)
                            if nuevo_rol:
                                st.session_state.rol = nuevo_rol
                        st.rerun()
                    else:
                        st.error(msg)

            if usuario_obj == st.session_state.usuario_actual:
                st.caption("No puedes eliminar tu propio usuario mientras estás en sesión.")
            elif usuario_obj in ["superadmin", "admin"]:
                st.caption("Las cuentas administrativas base están protegidas y no se pueden eliminar.")
            elif st.button("🗑️ Eliminar usuario seleccionado"):
                exito, msg = gestor_usuarios.eliminar_usuario(usuario_obj)
                if exito:
                    registrar_movimiento("Eliminar Usuario", usuario_obj)
                    st.success(msg)
                    st.rerun()
                else:
                    st.error(msg)
    else:
        st.info("Solo administradores o superadministradores pueden gestionar usuarios")

    st.markdown("---")
    st.subheader("� Historial de Movimientos")
    if st.session_state.movimientos:
        df_mov = pd.DataFrame(st.session_state.movimientos)
        col_mov1, col_mov2 = st.columns([2, 1])
        with col_mov1:
            fecha_desde = st.date_input("Desde", value=datetime.now().date(), key="mov_desde")
        with col_mov2:
            fecha_hasta = st.date_input("Hasta", value=datetime.now().date(), key="mov_hasta")

        df_mov['_fecha_dt'] = pd.to_datetime(df_mov['fecha'], errors='coerce')
        inicio_dt = pd.to_datetime(fecha_desde)
        fin_dt = pd.to_datetime(fecha_hasta) + pd.Timedelta(days=1)
        df_mov_filtrado = df_mov[(df_mov['_fecha_dt'] >= inicio_dt) & (df_mov['_fecha_dt'] < fin_dt)].copy()
        df_mov_filtrado = df_mov_filtrado.drop(columns=['_fecha_dt']).sort_values(by='fecha', ascending=False).reset_index(drop=True)

        st.dataframe(df_mov_filtrado, width="stretch")

        if FPDF is not None and not df_mov_filtrado.empty:
            pdf_audit = FPDF(format='letter')
            pdf_audit.set_auto_page_break(auto=True, margin=15)
            pdf_audit.add_page()
            pdf_audit.set_font("Arial", "B", 12)
            pdf_audit.cell(0, 10, "Bitacora de Auditoria", ln=True, align='C')
            pdf_audit.set_font("Arial", size=10)
            pdf_audit.cell(0, 8, f"Rango: {fecha_desde} a {fecha_hasta}", ln=True)
            pdf_audit.ln(2)
            ancho_util = max(10, pdf_audit.w - pdf_audit.l_margin - pdf_audit.r_margin)

            for _, row in df_mov_filtrado.iterrows():
                linea = _texto_seguro_pdf(
                    f"{row.get('fecha')} | {row.get('usuario')} | {row.get('tipo')} | {row.get('detalle')}"
                )
                # Asegura iniciar cada renglon desde el margen izquierdo para evitar errores de ancho.
                pdf_audit.set_x(pdf_audit.l_margin)
                pdf_audit.multi_cell(ancho_util, 7, linea)

            pdf_audit_bytes = _pdf_a_bytes(pdf_audit)
            st.download_button(
                label="📄 Descargar Auditoria PDF",
                data=pdf_audit_bytes,
                file_name=f"auditoria_{fecha_desde}_{fecha_hasta}.pdf",
                mime="application/pdf",
                disabled=not permiso_exportar,
            )
    else:
        st.info("No hay movimientos registrados aún")

# Footer
# Footer
import datetime
import os
import hashlib

st.markdown("---")
anio_actual = datetime.datetime.now().year

# Intentar resolver información del commit desplegado (varias fuentes posibles)
commit_sha = None
commit_source = None
commit_date = None

# 1) Variables de entorno (Streamlit/GitHub Actions/CI pueden exponerlas)
for env_var in ("GIT_COMMIT", "GITHUB_SHA", "STREAMLIT_GIT_COMMIT", "COMMIT_SHA"):
    val = os.getenv(env_var)
    if val:
        commit_sha = val.strip()
        commit_source = f"env:{env_var}"
        break

# 2) Si hay un repo .git, intentar leer HEAD
if not commit_sha:
    try:
        git_head = os.path.join(os.path.dirname(__file__), ".git", "HEAD")
        if os.path.exists(git_head):
            with open(git_head, "r", encoding="utf-8") as f:
                head = f.read().strip()
            if head.startswith("ref:"):
                ref = head.split(" ", 1)[1].strip()
                ref_path = os.path.join(os.path.dirname(__file__), ".git", ref)
                if os.path.exists(ref_path):
                    with open(ref_path, "r", encoding="utf-8") as f:
                        commit_sha = f.read().strip()
                        commit_source = f".git:{ref}"
            else:
                # detached HEAD
                commit_sha = head
                commit_source = ".git:HEAD"
    except Exception:
        commit_sha = commit_sha

# 3) Fallback: hash del contenido del archivo para identificar versión desplegada
if not commit_sha:
    try:
        with open(__file__, "rb") as f:
            data = f.read()
        commit_sha = hashlib.sha1(data).hexdigest()
        commit_source = "file:sha1"
        commit_date = datetime.datetime.fromtimestamp(os.path.getmtime(__file__)).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        commit_sha = None

short_sha = commit_sha[:10] if commit_sha else "desconocido"

# Mostrar footer con información de commit/version
st.markdown(f"""
<div style='text-align: center; color: gray; font-size: 12px;'>
    <p>Sistema de Gestión de Planes Telefónicos Corporativos</p>
    <p>Última actualización (año): {anio_actual}</p>
    <p style='font-size:11px; margin-top:6px;'>Versión desplegada: <strong>{short_sha}</strong> <span style='color:#9aa0a6;'>({commit_source or 'desconocido'})</span></p>
""", unsafe_allow_html=True)

if commit_date:
    st.markdown(f"<div style='text-align:center; color:gray; font-size:11px;'>Commit date: {commit_date}</div>", unsafe_allow_html=True)

# --- Banner visual del celular al final de la app (última línea) ---
if st.session_state.get("usuario_actual"):
    render_hero_principal(st.session_state.usuario_actual, st.session_state.rol)
